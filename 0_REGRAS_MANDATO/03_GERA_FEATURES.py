# -*- coding: utf-8 -*-
"""
ARCHANGEL v1 - 3_GERA_FEATURES_V3_FAST_AUDITED.py

Objetivo:
    Gerar features técnicas, estatísticas, risco, regime, liquidez,
    contexto temporal e auditoria pós-geração para séries OHLCV.

Melhorias vs versão anterior:
    - Mais rápido para timeframes curtos: 1min e 3min.
    - Report Excel mais limpo e orientado a decisão.
    - Pós-auditoria de sincronização entre OHLCV e features.
    - Preparação para ML/DL, risk management e backtesting.
    - Menos explosão de linhas no Excel.
    - Mais governança anti-leakage.

Entradas:
    <PROJECT_ROOT>\\0_REGRAS_MANDATO\\BASE_JSON\\01_MAPA_ATIVOS_LATEST.json

Saídas:
    <PROJECT_ROOT>\\3_FEATURES\\FEATURES_PARQUET\\...
    <PROJECT_ROOT>\\0_REGRAS_MANDATO\\BASE_JSON\\03_FEATURES_CATALOG_LATEST.json
    <PROJECT_ROOT>\\0_REGRAS_MANDATO\\BASE_EXCEL\\3_FEATURES.xlsx
    <PROJECT_ROOT>\\3_FEATURES\\_logs\\...

Notas:
    - Este script NÃO cria labels.
    - Este script NÃO usa dados futuros.
    - DateTime é preservado como naive, conforme política Asia/Dubai.
"""

from __future__ import annotations

import os
import gc
import re
import sys
import json
import time
import math
import hashlib
import traceback
import warnings
from pathlib import Path
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple
from concurrent.futures import FIRST_COMPLETED, ProcessPoolExecutor, wait

import numpy as np
import pandas as pd
from pandas.errors import PerformanceWarning


def env_bool(name: str, default: bool) -> bool:
    raw = os.environ.get(name)
    if raw is None:
        return default
    return str(raw).strip().lower() in {"1", "true", "yes", "y", "sim", "on"}


# =============================================================================
# 0. WARNINGS
# =============================================================================

warnings.simplefilter("ignore", PerformanceWarning)
warnings.filterwarnings("ignore", category=RuntimeWarning)


# =============================================================================
# 1. PATHS E CONFIGURAÇÕES
# =============================================================================

# A raiz é inferida do local deste script, sem depender da letra da unidade.
ROOT_DIR = Path(__file__).resolve().parent.parent
os.environ.setdefault("CUPY_CACHE_DIR", str(ROOT_DIR / "_CACHE" / "cupy_kernel_cache"))
os.environ.setdefault("NUMBA_CACHE_DIR", str(ROOT_DIR / "_CACHE" / "numba_cache"))

FEATURES_DIR = ROOT_DIR / "3_FEATURES"
FEATURES_PARQUET_DIR = FEATURES_DIR / "FEATURES_PARQUET"
FEATURES_LOG_DIR = FEATURES_DIR / "_logs"

BASE_REGRAS_DIR = ROOT_DIR / "0_REGRAS_MANDATO"
BASE_JSON_DIR = BASE_REGRAS_DIR / "BASE_JSON"
BASE_EXCEL_DIR = BASE_REGRAS_DIR / "BASE_EXCEL"


MAPA_ATIVOS_PATH = BASE_JSON_DIR / "01_MAPA_ATIVOS_LATEST.json"
FEATURES_JSON_PATH = BASE_JSON_DIR / "03_FEATURES_CATALOG_LATEST.json"
FEATURES_EXCEL_PATH = BASE_EXCEL_DIR / "3_FEATURES.xlsx"

SCRIPT_NAME = "3_GERA_FEATURES_V3_FAST_AUDITED.py"
SYSTEM_NAME = "ARCHANGEL"
SYSTEM_VERSION = "v1"

SCHEMA_VERSION = "ARCHANGEL_FEATURE_STORE_3.3_CUDA_CONTROLLED_ROLLING"



RUN_ID = os.environ.get("ARCHANGEL_RUN_ID") or datetime.now().strftime("%Y%m%d_%H%M%S")
os.environ.setdefault("ARCHANGEL_RUN_ID", RUN_ID)
INCREMENTAL_AUDIT_PATH = FEATURES_LOG_DIR / f"3_FEATURES_RUN_AUDIT_INCREMENTAL_{RUN_ID}.jsonl"
RUN_REPORT_PATH = FEATURES_LOG_DIR / f"3_FEATURES_RUN_REPORT_{RUN_ID}.json"
RUN_REPORT_LATEST_PATH = BASE_JSON_DIR / "03_FEATURES_RUN_REPORT_LATEST.json"
RUN_REPORT_BASE_JSON_PATH = RUN_REPORT_LATEST_PATH

TIMEZONE_LOCAL = "Asia/Dubai"
DATETIME_COL = "DateTime"
TIMESTAMP_UTC_MS_COL = "timestamp_utc_ms"
BAR_TIMESTAMP_POLICY = "close_time"

TIMEZONE_POLICY = {
    "reference_timezone": TIMEZONE_LOCAL,
    "datetime_column": DATETIME_COL,
    "timestamp_utc_ms_column": TIMESTAMP_UTC_MS_COL,
    "datetime_assumption": "DateTime salvo como Asia/Dubai local time naive.",
    "meta_timezone": TIMEZONE_LOCAL,
    "meta_datetime_is_naive": True,
    "meta_timestamp_utc_ms_present": True,
    "bar_timestamp_policy": BAR_TIMESTAMP_POLICY,
    "warning": (
        "Este script preserva DateTime como Dubai naive e usa timestamp_utc_ms "
        "como chave temporal universal interna quando disponível."
    ),
}


# =============================================================================
# 1B. DATA QUALITY, CONTRATO E ML READINESS
# =============================================================================

BASES_DIR = ROOT_DIR / "2_BASES"
BASES_QUALITY_DIR = BASES_DIR / "_quality"

DATA_QUALITY_REPORT_PATH = BASE_JSON_DIR / "02_01_DATA_QUALITY_REPORT_LATEST.json"
COST_MODEL_PATH = BASE_JSON_DIR / "00_COST_MODEL.json"
PYTHON_ENVIRONMENT_PATH = BASE_JSON_DIR / "00_02_PYTHON_ENVIRONMENT_LATEST.json"

INPUT_CONTRACT_VERSION = "ARCHANGEL_OHLCV_INPUT_CONTRACT_1.0"
QUALITY_GATE_VERSION = "ARCHANGEL_DATA_QUALITY_GATE_1.0"
ML_READINESS_VERSION = "ARCHANGEL_FEATURE_ML_READY_1.0"

ENABLE_PRE_FEATURE_QUALITY_GATE = True
ALLOW_WARNING_QUALITY_TO_PROCESS = True
DEFAULT_QUALITY_STATUS_IF_MISSING = "WARNING"

BLOCK_ON_QUALITY_FAIL = True
BLOCK_ON_INPUT_CONTRACT_FAIL = True
BLOCK_ON_TIME_GRID_FAIL = False

ENABLE_INPUT_CONTRACT_VALIDATION = True
ENABLE_TIME_GRID_VALIDATION = True
ENABLE_TIMEZONE_CONSISTENCY_CHECK = True
BLOCK_ON_TIMEZONE_CONSISTENCY_FAIL = True


ENABLE_FEATURE_READY_TIMESTAMP = True
ENABLE_FEATURE_VALIDITY_FLAGS = True
ENABLE_ML_READY_SCHEMA_FLAGS = True
MIN_FEATURE_NON_NULL_RATIO_FOR_VALID_ROW = 0.95


MAX_TIMESTAMP_DATETIME_DRIFT_MS = 1_000
TIME_GRID_IRREGULAR_TOLERANCE_MS = 1_000

TIME_GRID_FAIL_LARGE_GAP_RATIO_THRESHOLD = 0.01
TIME_GRID_FAIL_IRREGULAR_RATIO_THRESHOLD = 0.05

REQUIRED_INPUT_CONTRACT_COLUMNS = {
    DATETIME_COL,
    TIMESTAMP_UTC_MS_COL,
    "Open",
    "High",
    "Low",
    "Close",
    "Volume",
}


QUALITY_WARNING_STATUSES = {
    "WARNING",
    "WARN",
    "UNKNOWN",
    "MISSING_REPORT",
    "CHECK_WARNING",
}

QUALITY_PASS_STATUSES = {"PASS", "OK", "CHECK_OK"}
QUALITY_FAIL_STATUSES = {
    "FAIL",
    "ERROR",
    "CRITICAL",
    "CHECK_FAIL",
}



QUALITY_CAN_PROCESS_STATUSES = (
    QUALITY_PASS_STATUSES
    | (QUALITY_WARNING_STATUSES if ALLOW_WARNING_QUALITY_TO_PROCESS else set())
)

ML_ALLOWED_FEATURE_PREFIXES = ("feat_", "xasset_", "regime_")
ML_FORBIDDEN_FEATURE_PREFIXES = ("label_", "meta_", "quality_")

FILTER_ASSETS: Optional[set[str]] = None
FILTER_SOURCES: Optional[set[str]] = None
FILTER_TIMEFRAMES: Optional[set[str]] = None

DATASET_KIND_ALLOWED = {"ohlcv"}
ONLY_QUALITY_OK = False
MIN_ROWS = 300

OVERWRITE_EXISTING = True

INCLUDE_OHLCV_IN_OUTPUT = True
INCLUDE_METADATA_COLUMNS = True
METADATA_OUTPUT_MODE = os.environ.get("ARCHANGEL_FEATURE_METADATA_OUTPUT_MODE", "MINIMAL").strip().upper()
READ_OPTIONAL_MARKET_COLUMNS = env_bool("ARCHANGEL_READ_OPTIONAL_MARKET_COLUMNS", False)
INCLUDE_OPTIONAL_MARKET_COLUMNS_IN_OUTPUT = env_bool("ARCHANGEL_INCLUDE_OPTIONAL_MARKET_COLUMNS", False)

DOWNCAST_FLOATS_TO_FLOAT32 = True
FEATURE_EXECUTION_PROFILE = os.environ.get("ARCHANGEL_FEATURE_PROFILE", "FAST_FIRST").strip().upper()

PARQUET_COMPRESSION = (
    os.environ.get("ARCHANGEL_PARQUET_COMPRESSION")
    or ("snappy" if FEATURE_EXECUTION_PROFILE in {"FAST_FIRST", "FAST"} else "zstd")
)
PARQUET_ROW_GROUP_SIZE = int(os.environ.get("ARCHANGEL_PARQUET_ROW_GROUP_SIZE", "1000000"))
PARQUET_ENGINE = "pyarrow"

HASH_SAMPLE_ROWS = 100
QUALITY_SUMMARY_FAST_FOR_LARGE_SERIES = env_bool("ARCHANGEL_QUALITY_SUMMARY_FAST", True)
QUALITY_SUMMARY_FULL_NULLS_ROW_THRESHOLD = int(os.environ.get("ARCHANGEL_QUALITY_FULL_NULLS_ROW_THRESHOLD", "750000"))
QUALITY_SUMMARY_SAMPLE_ROWS = int(os.environ.get("ARCHANGEL_QUALITY_SUMMARY_SAMPLE_ROWS", "250000"))
PRIORITIZE_TIMEFRAME_FROM_PATH = True

DROP_FEATURES_WITH_NULL_RATIO_ABOVE = 0.985
DROP_CONSTANT_FEATURES = True

ENABLE_POST_AUDIT = True
POST_AUDIT_MAX_GAP_MULTIPLIER = 3.5

READ_PARQUET_ONLY_OHLCV_WHEN_POSSIBLE = True

ENABLE_PARALLEL_PROCESSING = True

FEATURE_CUDA_MODE = os.environ.get("ARCHANGEL_FEATURE_CUDA_MODE", "auto").strip().lower()
FEATURE_CUDA_AUTO_MIN_GPU_MEMORY_MB = float(os.environ.get("ARCHANGEL_FEATURE_CUDA_AUTO_MIN_GPU_MEMORY_MB", "4096"))
FEATURE_CUDA_AUTO_MIN_ROWS = int(os.environ.get("ARCHANGEL_FEATURE_CUDA_AUTO_MIN_ROWS", "1500000"))
FEATURE_CUDA_MAX_WORKERS = int(os.environ.get("ARCHANGEL_FEATURE_CUDA_MAX_WORKERS", "1"))
FEATURE_CUDA_ACCELERATED_BLOCKS = (
    "returns_vector_core",
    "rolling_return",
    "volatility",
    "vol_of_vol",
    "sma",
    "volume",
    "atr_range_volatility",
    "range_volatility_estimators",
    "trend_strength",
    "regime",
    "volume_pressure",
    "shock_features",
)
FEATURE_CUDA_NOTE = (
    "CUDA na etapa 3 é aplicado somente em operações vetoriais e rolling cumulativas seguras "
    "(sum/mean/std/corr), com fallback CPU e benchmark de equivalência numérica. "
    "Rolling min/max, quantile, skew/kurt e apply(MAD) continuam no Pandas/CPU nesta fase."
)

# Seu hardware tem 12 cores físicos / 24 threads,
# mas features + pandas + parquet consomem RAM.
# O perfil padrão atual é ultra-agressivo-controlado: mira alto uso de CPU/RAM
# enquanto preserva reserva para Windows, VS Code, cache de disco e escrita Parquet.
FEATURE_RESOURCE_PROFILE = os.environ.get("ARCHANGEL_FEATURE_RESOURCE_PROFILE", "ULTRA").strip().upper()
MAX_WORKERS_FEATURES = int(os.environ.get("ARCHANGEL_MAX_WORKERS_FEATURES", str(os.cpu_count() or 1)))
ENABLE_MEMORY_AWARE_WORKERS = True
ARCHANGEL_RAM_CAP_GB = float(os.environ.get("ARCHANGEL_RAM_CAP_GB", "110"))
TARGET_RAM_USED_GB = float(os.environ.get("ARCHANGEL_TARGET_RAM_USED_GB", "100"))
TARGET_CPU_PERCENT = float(os.environ.get("ARCHANGEL_TARGET_CPU_PERCENT", "85"))
MIN_FREE_RAM_GB_TO_START_BATCH = float(os.environ.get("ARCHANGEL_MIN_FREE_RAM_GB", "10"))
ESTIMATED_RAM_GB_PER_WORKER_STANDARD = float(os.environ.get("ARCHANGEL_ESTIMATED_RAM_GB_PER_WORKER_STANDARD", "4"))
ESTIMATED_RAM_GB_PER_WORKER_TURBO = float(os.environ.get("ARCHANGEL_ESTIMATED_RAM_GB_PER_WORKER_TURBO", "5"))

PROCESS_TURBO_TIMEFRAMES_FIRST = True
ENABLE_TIMEFRAME_BATCH_EXECUTION = True
MAX_WORKERS_TURBO_BATCH = int(os.environ.get("ARCHANGEL_MAX_WORKERS_TURBO_BATCH", str(min(22, MAX_WORKERS_FEATURES))))

ENABLE_ADAPTIVE_WORKERS = env_bool("ARCHANGEL_ENABLE_ADAPTIVE_WORKERS", True)
ADAPTIVE_WORKERS_INITIAL_TURBO = int(os.environ.get("ARCHANGEL_ADAPTIVE_INITIAL_TURBO_WORKERS", "14"))
ADAPTIVE_WORKERS_INITIAL_STANDARD = int(os.environ.get("ARCHANGEL_ADAPTIVE_INITIAL_STANDARD_WORKERS", "18"))
ADAPTIVE_SCALE_UP_FREE_RAM_GB = float(os.environ.get("ARCHANGEL_ADAPTIVE_SCALE_UP_FREE_RAM_GB", "26"))
ADAPTIVE_HOLD_FREE_RAM_GB = float(os.environ.get("ARCHANGEL_ADAPTIVE_HOLD_FREE_RAM_GB", "12"))
ADAPTIVE_SCALE_DOWN_FREE_RAM_GB = float(os.environ.get("ARCHANGEL_ADAPTIVE_SCALE_DOWN_FREE_RAM_GB", "8"))
ADAPTIVE_SCALE_UP_STABLE_HEARTBEATS = int(os.environ.get("ARCHANGEL_ADAPTIVE_SCALE_UP_STABLE_HEARTBEATS", "1"))
ADAPTIVE_SCALE_UP_WORKER_STEP = int(os.environ.get("ARCHANGEL_ADAPTIVE_SCALE_UP_WORKER_STEP", "4"))
ADAPTIVE_SCALE_DOWN_WORKER_STEP = int(os.environ.get("ARCHANGEL_ADAPTIVE_SCALE_DOWN_WORKER_STEP", "3"))
ADAPTIVE_FAILURE_COOLDOWN_HEARTBEATS = int(os.environ.get("ARCHANGEL_ADAPTIVE_FAILURE_COOLDOWN_HEARTBEATS", "1"))
ADAPTIVE_CPU_LOW_SCALE_UP_THRESHOLD = float(os.environ.get("ARCHANGEL_ADAPTIVE_CPU_LOW_SCALE_UP_THRESHOLD", "70"))

ENABLE_RETRY_FAILED_SERIES = env_bool("ARCHANGEL_RETRY_FAILED_SERIES", True)
MAX_RETRY_ATTEMPTS_PER_SERIES = int(os.environ.get("ARCHANGEL_MAX_RETRY_ATTEMPTS_PER_SERIES", "1"))
RETRY_FAILED_ONLY_FROM_REPORT = os.environ.get("ARCHANGEL_RETRY_FAILED_ONLY_FROM_REPORT")
RETRY_OUTPUT_JSON_PATH = FEATURES_LOG_DIR / f"3_FEATURES_RETRY_PLAN_{RUN_ID}.json"
RETRY_OUTPUT_CSV_PATH = FEATURES_LOG_DIR / f"3_FEATURES_RETRY_PLAN_{RUN_ID}.csv"
RETRY_OUTPUT_LATEST_PATH = BASE_JSON_DIR / "03_FEATURES_RETRY_PLAN_LATEST.json"
RETRY_OUTPUT_BASE_JSON_PATH = RETRY_OUTPUT_LATEST_PATH

ENABLE_PROGRESS_HEARTBEAT = True
PROGRESS_HEARTBEAT_SECONDS = max(1.0, float(os.environ.get("ARCHANGEL_PROGRESS_HEARTBEAT_SECONDS", "10")))
PROGRESS_SHOW_ACTIVE_LIMIT = max(1, int(os.environ.get("ARCHANGEL_PROGRESS_SHOW_ACTIVE_LIMIT", "6")))
PROGRESS_SHOW_RECENT_LIMIT = max(0, int(os.environ.get("ARCHANGEL_PROGRESS_SHOW_RECENT_LIMIT", "5")))

RAM_SAFE_ROW_THRESHOLD = int(os.environ.get("ARCHANGEL_RAM_SAFE_ROW_THRESHOLD", "500000"))
FAST_FIRST_ROW_THRESHOLD = int(os.environ.get("ARCHANGEL_FAST_FIRST_ROW_THRESHOLD", "250000"))
FAST_FIRST_SKIP_FAMILIES = {
    "autocorrelation",
    "donchian",
    "drawdown_distance",
    "higher_moments",
    "liquidity_proxy",
    "market_structure",
    "oscillators",
    "range_volatility_estimators",
    "regime",
    "risk_proxy",
    "shock_features",
    "vol_of_vol",
    "slopes",
    "trend_strength",
    "volume_pressure",
}
RAM_SAFE_SKIP_FAMILIES_FOR_TURBO = {
    "autocorrelation",
    "higher_moments",
    "risk_proxy",
    "vol_of_vol",
}
RAM_SAFE_SKIP_FAMILIES_FOR_LONG_SERIES = {
    "autocorrelation",
    "higher_moments",
}

ENABLE_FEATURE_BLOCK_CONSTRUCTION = True
FEATURE_BLOCK_COLUMN_COUNT = int(os.environ.get("ARCHANGEL_FEATURE_BLOCK_COLUMNS", "96"))
ENABLE_PROCESS_MEMORY_TELEMETRY = True

_CPU_TIMES_PREVIOUS: Optional[Tuple[int, int, int]] = None

FORCE_GC_EACH_SERIES = True
FORCE_GC_EACH_FEATURE_BLOCK = False
VERBOSE_SERIES_LOG = False

# Excel mais leve por default.
WRITE_FULL_FEATURE_DETAILS_TO_EXCEL = False
SAVE_FEATURE_QUALITY_DETAILS = False
SAVE_FEATURE_DISTRIBUTION_STATS = False

# Modo turbo para séries muito longas ou timeframes curtos.
ENABLE_TIMEFRAME_TURBO_MODE = True
TURBO_TIMEFRAMES = {"1min", "3min", "2min"}
TURBO_DISABLE_AUTOCORR = True
TURBO_DISABLE_HIGHER_MOMENTS = True
TURBO_DISABLE_ROLLING_VAR_QUANTILE = True
TURBO_DISABLE_VOLUME_ROLLING_CORR = env_bool("ARCHANGEL_TURBO_DISABLE_VOLUME_ROLLING_CORR", True)
TURBO_LIMIT_LONG_WINDOWS = True
TURBO_MAX_WINDOW = 200

FAST_MODE_DISABLE_ENTROPY = True
FAST_MODE_DISABLE_SLOPES = True
FEATURE_VALIDITY_FINITE_CHECK = env_bool("ARCHANGEL_FEATURE_VALIDITY_FINITE_CHECK", False)


# =============================================================================
# 2. FEATURE CONFIG
# =============================================================================

FEATURE_CONFIG = {
    "returns": {
        "enabled": True,
        "windows": [1, 2, 3, 5, 7, 11, 15, 27, 41, 77, 137],
    },
    "return_lags": {
        "enabled": True,
        "lags": [1, 2, 3, 5, 10, 20],
    },
    "rolling_return": {
        "enabled": True,
        "windows": [3, 5, 7, 11, 15, 27, 41, 77, 137],
    },
    "volatility": {
        "enabled": True,
        "windows": [5, 7, 11, 15, 27, 41, 77, 137],
        "annualization": True,
    },
    "vol_of_vol": {
        "enabled": True,
        "windows": [20, 50, 100, 200],
    },
    "ema": {
        "enabled": True,
        "windows": [3, 5, 7, 11, 15, 27, 41, 77, 137],
    },
    "sma": {
        "enabled": True,
        "windows": [3, 5, 7, 11, 15, 27, 41, 77, 137],
    },
    "bollinger": {
        "enabled": True,
        "windows": [20, 50, 100],
        "num_std": 2.0,
    },
    "rsi": {
        "enabled": True,
        "windows": [7, 14, 21, 30],
    },
    "macd": {
        "enabled": True,
        "fast": 12,
        "slow": 26,
        "signal": 9,
    },
    "range_candle": {
        "enabled": True,
    },
    "volume": {
        "enabled": True,
        "windows": [5, 10, 20, 50, 100, 200],
    },
    "drawdown_distance": {
        "enabled": True,
        "windows": [20, 50, 100, 200, 288],
    },
    "trend_strength": {
        "enabled": True,
        "windows": [10, 20, 50, 100, 200],
    },
    "autocorrelation": {
        "enabled": True,
        "windows": [20, 50, 100],
        "lags": [1, 2, 3, 5],
    },
    "atr_range_volatility": {
        "enabled": True,
        "windows": [7, 14, 21, 50],
    },
    "range_volatility_estimators": {
        "enabled": True,
        "windows": [5, 10, 20, 50, 100],
    },
    "donchian": {
        "enabled": True,
        "windows": [20, 50, 100],
    },
    "oscillators": {
        "enabled": True,
        "stoch_windows": [14, 21, 50],
        "cci_windows": [20, 50],
    },
    "slopes": {
        "enabled": not FAST_MODE_DISABLE_SLOPES,
        "windows": [5, 10, 20, 50, 100],
    },
    "higher_moments": {
        "enabled": True,
        "windows": [20, 50, 100, 200],
    },
    "regime": {
        "enabled": True,
        "windows": [20, 50, 100, 200],
        "entropy_enabled": not FAST_MODE_DISABLE_ENTROPY,
    },
    "risk_proxy": {
        "enabled": True,
        "windows": [20, 50, 100, 200],
        "var_quantiles": [0.01, 0.05],
    },
    "volume_pressure": {
        "enabled": True,
        "windows": [10, 20, 50, 100],
    },
    "liquidity_proxy": {
        "enabled": True,
        "windows": [10, 20, 50, 100],
    },
    "shock_features": {
        "enabled": True,
        "windows": [20, 50, 100],
    },
    "market_structure": {
        "enabled": True,
        "windows": [20, 50, 100],
    },
    "time_context": {
        "enabled": True,
    },
    "ml_safety": {
        "enabled": True,
    },
}


FEATURE_FAMILY_DESCRIPTIONS = {
    "returns": "Retornos logarítmicos e simples.",
    "return_lags": "Lags explícitos de retorno para modelos ML.",
    "rolling_return": "Retornos acumulados e médios móveis.",
    "volatility": "Volatilidade realizada e anualizada.",
    "vol_of_vol": "Volatilidade da volatilidade.",
    "ema": "EMAs, distâncias e spreads.",
    "sma": "SMAs e distâncias relativas.",
    "bollinger": "Bandas de Bollinger.",
    "rsi": "RSI bruto e normalizado.",
    "macd": "MACD clássico.",
    "range_candle": "Geometria do candle.",
    "volume": "Volume, z-score e correlação volume-retorno.",
    "drawdown_distance": "Distância contra máximas/mínimas móveis.",
    "trend_strength": "Força de tendência normalizada.",
    "autocorrelation": "Autocorrelação móvel.",
    "atr_range_volatility": "True Range e ATR.",
    "range_volatility_estimators": "Parkinson e Garman-Klass.",
    "donchian": "Canais de Donchian.",
    "oscillators": "Stochastic, Williams %R e CCI.",
    "slopes": "Inclinação rolling do preço.",
    "higher_moments": "Skew e kurtosis dos retornos.",
    "regime": "Eficiência, regime de volatilidade e entropia opcional.",
    "risk_proxy": "Sharpe, Sortino e VaR histórico.",
    "volume_pressure": "Pressão de volume.",
    "liquidity_proxy": "Dollar volume e proxies de liquidez.",
    "shock_features": "Choques de retorno, range e volume.",
    "market_structure": "Breakout, posição estrutural e higher-high/lower-low simples.",
    "time_context": "Hora, dia da semana e contexto temporal.",
    "ml_safety": "Flags auxiliares para treino e filtragem.",
}


# =============================================================================
# 3. UTILITÁRIOS
# =============================================================================


def min_rows_for_timeframe(timeframe: Optional[str]) -> int:
    tf = str(timeframe)

    if tf == "7D":
        return 40

    if tf == "3D":
        return 80

    if tf == "1D":
        return 180

    if tf in {"47min", "37min", "23min", "13min"}:
        return 200

    return MIN_ROWS



def now_iso() -> str:
    return datetime.now().isoformat(timespec="seconds")


def ensure_dir(path: Path) -> Path:
    path.mkdir(parents=True, exist_ok=True)
    return path


def path_to_str(path: Path) -> str:
    return str(path)


def load_json(path: Path) -> Dict[str, Any]:
    if not path.exists():
        raise FileNotFoundError(f"JSON não encontrado: {path}")
    return json.loads(path.read_text(encoding="utf-8"))


def write_json_atomic(payload: Dict[str, Any], path: Path) -> None:
    ensure_dir(path.parent)
    tmp = path.with_suffix(path.suffix + ".tmp")
    tmp.write_text(json.dumps(payload, ensure_ascii=False, indent=4, default=str), encoding="utf-8")
    tmp.replace(path)


def get_process_memory_mb() -> Optional[float]:
    if not ENABLE_PROCESS_MEMORY_TELEMETRY:
        return None
    try:
        import psutil  # type: ignore
        return round(psutil.Process(os.getpid()).memory_info().rss / (1024 ** 2), 3)
    except Exception:
        pass

    if os.name == "nt":
        try:
            import ctypes
            from ctypes import wintypes

            class PROCESS_MEMORY_COUNTERS_EX(ctypes.Structure):
                _fields_ = [
                    ("cb", wintypes.DWORD),
                    ("PageFaultCount", wintypes.DWORD),
                    ("PeakWorkingSetSize", ctypes.c_size_t),
                    ("WorkingSetSize", ctypes.c_size_t),
                    ("QuotaPeakPagedPoolUsage", ctypes.c_size_t),
                    ("QuotaPagedPoolUsage", ctypes.c_size_t),
                    ("QuotaPeakNonPagedPoolUsage", ctypes.c_size_t),
                    ("QuotaNonPagedPoolUsage", ctypes.c_size_t),
                    ("PagefileUsage", ctypes.c_size_t),
                    ("PeakPagefileUsage", ctypes.c_size_t),
                    ("PrivateUsage", ctypes.c_size_t),
                ]

            psapi = ctypes.WinDLL("psapi.dll")
            counters = PROCESS_MEMORY_COUNTERS_EX()
            counters.cb = ctypes.sizeof(PROCESS_MEMORY_COUNTERS_EX)
            handle = ctypes.windll.kernel32.GetCurrentProcess()
            psapi.GetProcessMemoryInfo.argtypes = [
                wintypes.HANDLE,
                ctypes.POINTER(PROCESS_MEMORY_COUNTERS_EX),
                wintypes.DWORD,
            ]
            psapi.GetProcessMemoryInfo.restype = wintypes.BOOL
            ok = psapi.GetProcessMemoryInfo(handle, ctypes.byref(counters), counters.cb)
            if ok:
                return round(float(counters.WorkingSetSize) / (1024 ** 2), 3)
        except Exception:
            return None

    return None


def get_available_ram_gb() -> Optional[float]:
    try:
        import psutil  # type: ignore
        return round(psutil.virtual_memory().available / (1024 ** 3), 3)
    except Exception:
        pass

    if os.name == "nt":
        try:
            import ctypes

            class MEMORYSTATUSEX(ctypes.Structure):
                _fields_ = [
                    ("dwLength", ctypes.c_ulong),
                    ("dwMemoryLoad", ctypes.c_ulong),
                    ("ullTotalPhys", ctypes.c_ulonglong),
                    ("ullAvailPhys", ctypes.c_ulonglong),
                    ("ullTotalPageFile", ctypes.c_ulonglong),
                    ("ullAvailPageFile", ctypes.c_ulonglong),
                    ("ullTotalVirtual", ctypes.c_ulonglong),
                    ("ullAvailVirtual", ctypes.c_ulonglong),
                    ("sullAvailExtendedVirtual", ctypes.c_ulonglong),
                ]

            status = MEMORYSTATUSEX()
            status.dwLength = ctypes.sizeof(MEMORYSTATUSEX)
            if ctypes.windll.kernel32.GlobalMemoryStatusEx(ctypes.byref(status)):
                return round(float(status.ullAvailPhys) / (1024 ** 3), 3)
        except Exception:
            return None

    return None


def get_used_ram_gb() -> Optional[float]:
    total_ram_gb = get_total_ram_gb()
    available_ram_gb = get_available_ram_gb()
    if total_ram_gb is None or available_ram_gb is None:
        return None
    return round(max(0.0, float(total_ram_gb) - float(available_ram_gb)), 3)


def get_target_free_ram_gb() -> float:
    total_ram_gb = get_total_ram_gb()
    if total_ram_gb is None:
        return get_effective_min_free_ram_gb()

    free_from_target_used = max(0.0, float(total_ram_gb) - float(TARGET_RAM_USED_GB))
    free_from_cap = max(0.0, float(total_ram_gb) - float(ARCHANGEL_RAM_CAP_GB))
    return round(max(float(MIN_FREE_RAM_GB_TO_START_BATCH), free_from_target_used, free_from_cap), 3)


def get_system_cpu_percent() -> Optional[float]:
    try:
        import psutil  # type: ignore
        return round(float(psutil.cpu_percent(interval=None)), 2)
    except Exception:
        pass

    if os.name == "nt":
        try:
            import ctypes
            from ctypes import wintypes

            class FILETIME(ctypes.Structure):
                _fields_ = [
                    ("dwLowDateTime", wintypes.DWORD),
                    ("dwHighDateTime", wintypes.DWORD),
                ]

            def filetime_to_int(value: FILETIME) -> int:
                return (int(value.dwHighDateTime) << 32) + int(value.dwLowDateTime)

            idle = FILETIME()
            kernel = FILETIME()
            user = FILETIME()
            ok = ctypes.windll.kernel32.GetSystemTimes(
                ctypes.byref(idle),
                ctypes.byref(kernel),
                ctypes.byref(user),
            )
            if not ok:
                return None

            current = (
                filetime_to_int(idle),
                filetime_to_int(kernel),
                filetime_to_int(user),
            )

            global _CPU_TIMES_PREVIOUS
            previous = _CPU_TIMES_PREVIOUS
            _CPU_TIMES_PREVIOUS = current
            if previous is None:
                return None

            idle_delta = max(0, current[0] - previous[0])
            kernel_delta = max(0, current[1] - previous[1])
            user_delta = max(0, current[2] - previous[2])
            total_delta = kernel_delta + user_delta
            if total_delta <= 0:
                return None

            busy_ratio = 1.0 - (idle_delta / total_delta)
            return round(max(0.0, min(100.0, busy_ratio * 100.0)), 2)
        except Exception:
            return None

    return None


def get_total_ram_gb() -> Optional[float]:
    try:
        import psutil  # type: ignore
        return round(psutil.virtual_memory().total / (1024 ** 3), 3)
    except Exception:
        pass

    if os.name == "nt":
        try:
            import ctypes

            class MEMORYSTATUSEX(ctypes.Structure):
                _fields_ = [
                    ("dwLength", ctypes.c_ulong),
                    ("dwMemoryLoad", ctypes.c_ulong),
                    ("ullTotalPhys", ctypes.c_ulonglong),
                    ("ullAvailPhys", ctypes.c_ulonglong),
                    ("ullTotalPageFile", ctypes.c_ulonglong),
                    ("ullAvailPageFile", ctypes.c_ulonglong),
                    ("ullTotalVirtual", ctypes.c_ulonglong),
                    ("ullAvailVirtual", ctypes.c_ulonglong),
                    ("sullAvailExtendedVirtual", ctypes.c_ulonglong),
                ]

            status = MEMORYSTATUSEX()
            status.dwLength = ctypes.sizeof(MEMORYSTATUSEX)
            if ctypes.windll.kernel32.GlobalMemoryStatusEx(ctypes.byref(status)):
                return round(float(status.ullTotalPhys) / (1024 ** 3), 3)
        except Exception:
            return None

    return None


def get_effective_min_free_ram_gb() -> float:
    total_ram_gb = get_total_ram_gb()
    if total_ram_gb is None:
        return float(MIN_FREE_RAM_GB_TO_START_BATCH)

    reserve_from_cap = max(0.0, float(total_ram_gb) - float(ARCHANGEL_RAM_CAP_GB))
    return round(max(float(MIN_FREE_RAM_GB_TO_START_BATCH), reserve_from_cap), 3)


def timeframe_sort_key(timeframe: Optional[str]) -> Tuple[int, int, str]:
    seconds = infer_timeframe_seconds(timeframe)
    if seconds is None:
        return (1, 10**12, str(timeframe or "unknown"))
    return (0, int(seconds), str(timeframe))


def is_ram_heavy_timeframe(timeframe: Optional[str]) -> bool:
    seconds = infer_timeframe_seconds(timeframe)
    return bool(seconds is not None and seconds <= 180)


def should_apply_ram_safe_profile(timeframe: Optional[str], row_count: int) -> bool:
    if FEATURE_EXECUTION_PROFILE in {"FULL", "ALL"}:
        return False
    if FEATURE_EXECUTION_PROFILE in {"FAST_FIRST", "FAST", "RAM_SAFE", "SAFE"}:
        return True
    return bool(is_turbo_timeframe(timeframe) or int(row_count) >= RAM_SAFE_ROW_THRESHOLD)


def resolve_feature_config_for_series(
    timeframe: Optional[str],
    row_count: int,
) -> Tuple[Dict[str, Dict[str, Any]], Dict[str, Any]]:
    active = {
        family: dict(config)
        for family, config in FEATURE_CONFIG.items()
    }

    profile_notes: list[str] = []
    disabled_families: set[str] = set()
    ram_safe = should_apply_ram_safe_profile(timeframe, row_count)

    if FEATURE_EXECUTION_PROFILE in {"FAST_FIRST", "FAST"}:
        if is_turbo_timeframe(timeframe) or int(row_count) >= FAST_FIRST_ROW_THRESHOLD:
            disabled_families.update(FAST_FIRST_SKIP_FAMILIES)
            profile_notes.append("FAST_FIRST_HEAVY_SERIES_FAMILY_SKIP")

    if ram_safe and is_turbo_timeframe(timeframe):
        disabled_families.update(RAM_SAFE_SKIP_FAMILIES_FOR_TURBO)
        profile_notes.append("RAM_SAFE_TURBO_FAMILY_SKIP")

    if ram_safe and int(row_count) >= RAM_SAFE_ROW_THRESHOLD:
        disabled_families.update(RAM_SAFE_SKIP_FAMILIES_FOR_LONG_SERIES)
        profile_notes.append("RAM_SAFE_LONG_SERIES_FAMILY_SKIP")

    for family in disabled_families:
        if family in active:
            active[family]["enabled"] = False

    return active, {
        "feature_execution_profile": FEATURE_EXECUTION_PROFILE,
        "ram_safe_applied": ram_safe,
        "disabled_families": sorted(disabled_families),
        "profile_notes": profile_notes,
        "row_count": int(row_count),
        "timeframe": timeframe,
    }


def downcast_feature_block(block: pd.DataFrame) -> pd.DataFrame:
    if not DOWNCAST_FLOATS_TO_FLOAT32 or block.empty:
        return block
    for col in block.columns:
        if pd.api.types.is_float_dtype(block[col]):
            block[col] = block[col].astype("float32")
    return block


def build_feature_df_from_blocks(features: Dict[str, Any], index: pd.Index) -> pd.DataFrame:
    if not ENABLE_FEATURE_BLOCK_CONSTRUCTION:
        return pd.DataFrame(features, index=index)

    blocks: list[pd.DataFrame] = []
    keys = list(features.keys())
    block_size = max(1, int(FEATURE_BLOCK_COLUMN_COUNT))

    for start in range(0, len(keys), block_size):
        block_keys = keys[start:start + block_size]
        block_data = {key: features.pop(key) for key in block_keys}
        block = pd.DataFrame(block_data, index=index)
        block = block.replace([np.inf, -np.inf], np.nan)
        block = downcast_feature_block(block)
        blocks.append(block)
        del block_data
        if FORCE_GC_EACH_FEATURE_BLOCK:
            gc.collect()

    if not blocks:
        return pd.DataFrame(index=index)

    return pd.concat(blocks, axis=1)


def append_jsonl(path: Path, payload: Dict[str, Any]) -> None:
    ensure_dir(path.parent)
    with path.open("a", encoding="utf-8") as f:
        f.write(json.dumps(payload, ensure_ascii=False, default=str) + "\n")


def append_incremental_audit(record: Dict[str, Any]) -> None:
    payload = {
        "event_time": now_iso(),
        "run_id": RUN_ID,
        "script": SCRIPT_NAME,
        **record,
    }
    append_jsonl(INCREMENTAL_AUDIT_PATH, payload)


def safe_filename_token(value: Any) -> str:
    text = str(value)
    text = text.replace("\\", "_").replace("/", "_").replace(":", "_").replace(" ", "_").replace(".", "_")
    return re.sub(r"[^A-Za-z0-9_\-]+", "_", text)


def make_hash_from_dataframe(df: pd.DataFrame, rows: int = HASH_SAMPLE_ROWS) -> str:
    try:
        if df.empty:
            return "empty"
        sample = pd.concat([df.head(rows), df.tail(rows)], axis=0)
        return hashlib.sha256(sample.to_csv(index=False).encode("utf-8", errors="ignore")).hexdigest()
    except Exception:
        return "hash_error"


def sample_df_for_quality_summary(df: pd.DataFrame) -> Tuple[pd.DataFrame, bool]:
    if (
        not QUALITY_SUMMARY_FAST_FOR_LARGE_SERIES
        or df.empty
        or len(df) <= QUALITY_SUMMARY_FULL_NULLS_ROW_THRESHOLD
    ):
        return df, False

    sample_rows = max(1, min(int(QUALITY_SUMMARY_SAMPLE_ROWS), len(df)))
    if sample_rows >= len(df):
        return df, False

    positions = np.linspace(0, len(df) - 1, sample_rows, dtype=np.int64)
    positions = np.unique(positions)
    return df.iloc[positions], True


def short_series_hash(series_id: str, input_path: str) -> str:
    base = f"{series_id}|{input_path}"
    return hashlib.sha256(base.encode("utf-8", errors="ignore")).hexdigest()[:10]


def safe_get_nested(data: Dict[str, Any], keys: List[str], default: Any = None) -> Any:
    current = data
    for key in keys:
        if not isinstance(current, dict):
            return default
        current = current.get(key)
    return current if current is not None else default


def safe_div(numerator: Any, denominator: Any) -> Any:
    with np.errstate(divide="ignore", invalid="ignore"):
        return numerator / denominator


def safe_log(series: pd.Series) -> pd.Series:
    return np.log(series.where(series > 0))


def infer_timeframe_seconds(timeframe: str | None) -> Optional[int]:
    if not timeframe:
        return None

    mapping = {
        "1min": 60, "2min": 120, "3min": 180, "5min": 300, "7min": 420,
        "13min": 780, "15min": 900, "23min": 1380, "37min": 2220,
        "47min": 2820, "1h": 3600, "4h": 14400, "8h": 28800,
        "10h": 36000, "11h": 39600, "12h": 43200,
        "1D": 86400, "3D": 259200, "7D": 604800,
    }

    tf = str(timeframe).strip()

    if tf in mapping:
        return mapping[tf]

    lower = tf.lower()

    try:
        if lower.endswith("s"):
            return int(lower[:-1])
        if lower.endswith("min"):
            return int(lower.replace("min", "")) * 60
        if lower.endswith("h"):
            return int(lower.replace("h", "")) * 3600
        if lower.endswith("d"):
            return int(lower.replace("d", "")) * 86400
    except Exception:
        return None

    return None


def annualization_factor_from_timeframe(timeframe: Optional[str]) -> float:
    seconds = infer_timeframe_seconds(timeframe)
    if seconds is None or seconds <= 0:
        return 365.0
    return float(365 * 24 * 3600) / float(seconds)


def infer_timeframe_from_path(path_str: str | None) -> Optional[str]:
    if not path_str:
        return None

    text = str(path_str).replace("\\", "/").lower()
    filename = Path(str(path_str)).name.lower()

    patterns = [
        (r"(^|/|_)(\d+)_min($|/|_)", "min"),
        (r"(^|/|_)(\d+)min($|/|_|\.parquet)", "min"),
        (r"(^|/|_)(\d+)_hour($|/|_)", "h"),
        (r"(^|/|_)(\d+)h($|/|_|\.parquet)", "h"),
        (r"(^|/|_)(\d+)_day($|/|_)", "D"),
        (r"(^|/|_)(\d+)d($|/|_|\.parquet)", "D"),
        (r"(^|/|_)(\d+)s($|/|_|\.parquet)", "s"),
    ]

    for pattern, suffix in patterns:
        match = re.search(pattern, text)
        if match:
            value = int(match.group(2))
            return f"{value}{suffix}"

    for suffix, label in [("min", "min"), ("h", "h"), ("d", "D"), ("s", "s")]:
        match = re.search(rf"_(\d+){suffix}\.parquet$", filename)
        if match:
            return f"{int(match.group(1))}{label}"

    return None


def is_turbo_timeframe(timeframe: Optional[str]) -> bool:
    if not ENABLE_TIMEFRAME_TURBO_MODE:
        return False
    return str(timeframe) in TURBO_TIMEFRAMES


def filter_windows_for_timeframe(windows: List[int], timeframe: Optional[str]) -> List[int]:
    if is_turbo_timeframe(timeframe) and TURBO_LIMIT_LONG_WINDOWS:
        return [w for w in windows if int(w) <= TURBO_MAX_WINDOW]
    return windows



def normalize_timestamp_to_utc_ms(series: pd.Series) -> pd.Series:
    """
    Normaliza uma série temporal numérica para timestamp UTC em milissegundos.

    Aceita automaticamente:
        - segundos      ~ 1e9
        - milissegundos ~ 1e12
        - microssegundos ~ 1e15
        - nanosegundos  ~ 1e18

    Retorna:
        pd.Series dtype Int64 com timestamps em ms.
    """
    s = pd.to_numeric(series, errors="coerce")

    if s.dropna().empty:
        return pd.Series(pd.NA, index=series.index, dtype="Int64")

    median_abs = float(s.dropna().abs().median())

    # Nanosegundos Unix, comum em pandas datetime64[ns].astype(int)
    if median_abs >= 1e17:
        out = (s / 1_000_000).round()

    # Microssegundos Unix
    elif median_abs >= 1e14:
        out = (s / 1_000).round()

    # Milissegundos Unix
    elif median_abs >= 1e11:
        out = s.round()

    # Segundos Unix
    elif median_abs >= 1e8:
        out = (s * 1_000).round()

    else:
        # Valor pequeno demais para timestamp Unix confiável.
        out = pd.Series(pd.NA, index=series.index)

    return out.astype("Int64")




def utc_ms_to_dubai_naive(ms_series: pd.Series) -> pd.Series:
    """
    Converte timestamp UTC em milissegundos para DateTime Dubai naive.

    Antes de converter, normaliza a escala para ms:
        segundos / ms / micros / nanos -> ms.
    """
    ms_norm = normalize_timestamp_to_utc_ms(ms_series)
    dt_utc = pd.to_datetime(ms_norm, unit="ms", utc=True, errors="coerce")
    return dt_utc.dt.tz_convert(TIMEZONE_LOCAL).dt.tz_localize(None)





def datetime_dubai_naive_to_utc_ms(series: pd.Series) -> pd.Series:
    """
    Converte DateTime Dubai naive para timestamp UTC em milissegundos.

    Compatibilidade:
        - Se DateTime já vier timezone-aware, converte para UTC.
        - Se vier naive, assume Asia/Dubai.
    """
    dt = pd.to_datetime(series, errors="coerce")

    try:
        if getattr(dt.dt, "tz", None) is not None:
            dt_utc = dt.dt.tz_convert("UTC")
        else:
            dt_utc = dt.dt.tz_localize(TIMEZONE_LOCAL).dt.tz_convert("UTC")

        # Pandas pode representar datetimes em ns, us ou ms. Normalizar pela
        # magnitude evita assumir incorretamente que o epoch interno é sempre ns.
        return normalize_timestamp_to_utc_ms(dt_utc.astype("int64"))

    except Exception:
        return pd.Series(pd.NA, index=series.index, dtype="Int64")



def ensure_archangel_time_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    Garante compatibilidade entre bases antigas e novas.

    Política ARCHANGEL time-aware:
        - timestamp_utc_ms é a fonte da verdade quando existir.
        - Antes de usar timestamp_utc_ms, normaliza escala:
          segundos / ms / micros / nanos -> ms.
        - DateTime é derivado de timestamp_utc_ms como Asia/Dubai naive.
        - Se timestamp_utc_ms não existir, ele é reconstruído a partir de DateTime,
          assumindo que DateTime está em Asia/Dubai naive.
    """
    if df is None or df.empty:
        return df

    out = df.copy()

    timestamp_is_usable = False

    if TIMESTAMP_UTC_MS_COL in out.columns:
        normalized_ts = normalize_timestamp_to_utc_ms(out[TIMESTAMP_UTC_MS_COL])
        timestamp_is_usable = bool(normalized_ts.notna().mean() >= 0.90)
        if timestamp_is_usable:
            out[TIMESTAMP_UTC_MS_COL] = normalized_ts

    if DATETIME_COL in out.columns:
        out[DATETIME_COL] = pd.to_datetime(out[DATETIME_COL], errors="coerce")

    if DATETIME_COL in out.columns and not timestamp_is_usable:
        rebuilt_ts = datetime_dubai_naive_to_utc_ms(out[DATETIME_COL])
        if rebuilt_ts.notna().mean() >= 0.90:
            out[TIMESTAMP_UTC_MS_COL] = rebuilt_ts
            timestamp_is_usable = True

    if timestamp_is_usable:
        out[DATETIME_COL] = utc_ms_to_dubai_naive(out[TIMESTAMP_UTC_MS_COL])

    elif DATETIME_COL in out.columns:
        out[DATETIME_COL] = pd.to_datetime(out[DATETIME_COL], errors="coerce")

        try:
            if getattr(out[DATETIME_COL].dt, "tz", None) is not None:
                out[DATETIME_COL] = (
                    out[DATETIME_COL]
                    .dt.tz_convert(TIMEZONE_LOCAL)
                    .dt.tz_localize(None)
                )
        except Exception:
            pass

        out[TIMESTAMP_UTC_MS_COL] = datetime_dubai_naive_to_utc_ms(out[DATETIME_COL])

    if DATETIME_COL in out.columns and TIMESTAMP_UTC_MS_COL in out.columns:
        out = out.dropna(subset=[DATETIME_COL, TIMESTAMP_UTC_MS_COL]).copy()
        out[TIMESTAMP_UTC_MS_COL] = normalize_timestamp_to_utc_ms(
            out[TIMESTAMP_UTC_MS_COL]
        ).astype("int64")

    return out






def get_archangel_time_metadata() -> Dict[str, Any]:
    return {
        "meta_timezone": TIMEZONE_LOCAL,
        "meta_datetime_is_naive": True,
        "meta_timestamp_utc_ms_present": True,
        "meta_datetime_column": DATETIME_COL,
        "meta_timestamp_utc_ms_column": TIMESTAMP_UTC_MS_COL,
        "meta_bar_timestamp_policy": BAR_TIMESTAMP_POLICY,
    }


# =============================================================================
# 4. PREPARAÇÃO OHLCV
# =============================================================================
def normalize_ohlcv_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    Normaliza nomes de colunas vindos da camada 2.

    Padrão interno da camada 3:
        DateTime
        timestamp_utc_ms
        Open
        High
        Low
        Close
        Volume

    Observação:
        Mantemos OHLCV em PascalCase para compatibilidade com a função
        generate_all_features() existente.
    """
    rename = {}

    for col in df.columns:
        lower = str(col).strip().lower()

        if lower in {"datetime", "date_time", "date", "open_time", "close_time", "time"}:
            rename[col] = DATETIME_COL

        elif lower in {
            "timestamp_utc_ms",
            "utc_timestamp_ms",
            "meta_timestamp_utc_ms",
            "timestamp_ms",
            "close_time_utc_ms",
        }:
            rename[col] = TIMESTAMP_UTC_MS_COL

        elif lower == "timestamp":
            rename[col] = "timestamp"

        elif lower in {"open", "o"}:
            rename[col] = "Open"

        elif lower in {"high", "h"}:
            rename[col] = "High"

        elif lower in {"low", "l"}:
            rename[col] = "Low"

        elif lower in {"close", "c"}:
            rename[col] = "Close"

        elif lower in {"volume", "vol", "base_volume"}:
            rename[col] = "Volume"

        elif lower in {"quote_volume", "quote_asset_volume"}:
            rename[col] = "quote_volume"

        elif lower in {"number_of_trades", "trades", "trade_count"}:
            rename[col] = "number_of_trades"

        elif lower in {"taker_buy_base_volume", "taker_buy_volume"}:
            rename[col] = "taker_buy_base_volume"

        elif lower in {"taker_buy_quote_volume"}:
            rename[col] = "taker_buy_quote_volume"

    return df.rename(columns=rename)






def prepare_ohlcv(df_raw: pd.DataFrame) -> pd.DataFrame:
    """
    Prepara OHLCV sem esconder problemas críticos.

    Responsabilidades:
        - Normalizar nomes.
        - Garantir DateTime/timestamp_utc_ms.
        - Converter OHLCV para numérico.
        - Ordenar por timestamp_utc_ms.
        - Deduplicar mantendo último registro.

    Validação profunda fica em validate_input_contract().
    """
    df = normalize_ohlcv_columns(df_raw)


    if TIMESTAMP_UTC_MS_COL not in df.columns and "timestamp" in df.columns:
        ts_numeric = pd.to_numeric(df["timestamp"], errors="coerce")

        if ts_numeric.notna().mean() > 0.90:
            normalized_ts = normalize_timestamp_to_utc_ms(ts_numeric)

            if normalized_ts.notna().mean() > 0.90:
                df[TIMESTAMP_UTC_MS_COL] = normalized_ts
            else:
                df[DATETIME_COL] = pd.to_datetime(df["timestamp"], errors="coerce")
        else:
            df[DATETIME_COL] = pd.to_datetime(df["timestamp"], errors="coerce")







    required_price_cols = ["Open", "High", "Low", "Close", "Volume"]
    missing_price = [c for c in required_price_cols if c not in df.columns]

    if missing_price:
        raise ValueError(f"Colunas OHLCV obrigatórias ausentes: {missing_price}")

    if DATETIME_COL not in df.columns and TIMESTAMP_UTC_MS_COL not in df.columns:
        raise ValueError(
            f"Coluna temporal ausente. Esperado '{DATETIME_COL}' ou '{TIMESTAMP_UTC_MS_COL}'."
        )

    keep_cols = []

    if DATETIME_COL in df.columns:
        keep_cols.append(DATETIME_COL)

    if TIMESTAMP_UTC_MS_COL in df.columns:
        keep_cols.append(TIMESTAMP_UTC_MS_COL)

    keep_cols += required_price_cols

    optional_cols = [
        "quote_volume",
        "number_of_trades",
        "taker_buy_base_volume",
        "taker_buy_quote_volume",
    ] if READ_OPTIONAL_MARKET_COLUMNS else []

    keep_cols += [c for c in optional_cols if c in df.columns]
    keep_cols = list(dict.fromkeys(keep_cols))

    df = df.loc[:, keep_cols].copy()
    df = ensure_archangel_time_columns(df)

    for col in required_price_cols:
        df[col] = pd.to_numeric(df[col], errors="coerce")

    for col in optional_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    df = df.dropna(subset=[DATETIME_COL, TIMESTAMP_UTC_MS_COL])
    df[TIMESTAMP_UTC_MS_COL] = pd.to_numeric(
        df[TIMESTAMP_UTC_MS_COL],
        errors="coerce",
    ).astype("int64")

    df = df.sort_values(TIMESTAMP_UTC_MS_COL, kind="mergesort")
    df = df.drop_duplicates(subset=[TIMESTAMP_UTC_MS_COL], keep="last")
    df = df.reset_index(drop=True)

    final_cols = [
        DATETIME_COL,
        TIMESTAMP_UTC_MS_COL,
        "Open",
        "High",
        "Low",
        "Close",
        "Volume",
    ]

    final_cols += [c for c in optional_cols if c in df.columns]

    return df.loc[:, final_cols].copy()




# =============================================================================
# 4B. DATA QUALITY GATE, CONTRATO E VALIDAÇÃO TEMPORAL
# =============================================================================

def normalize_quality_status(status: Any) -> str:
    if status is None:
        return "UNKNOWN"

    s = str(status).strip().upper()

    if s in QUALITY_PASS_STATUSES:
        return "PASS"

    if s in QUALITY_WARNING_STATUSES:
        return "WARNING"

    if s in QUALITY_FAIL_STATUSES:
        return "FAIL"

    return "WARNING"


def load_quality_report(path: Path = DATA_QUALITY_REPORT_PATH) -> Dict[str, Any]:
    """
    Carrega relatório de qualidade da camada 2.

    Formatos aceitos:
        1. {"series": {"series_id": {...}}}
        2. {"series_by_id": {"series_id": {...}}}
        3. {"results": [{"series_id": "..."}]}
        4. Ausente: retorna relatório vazio.
    """
    if not path.exists():
        return {
            "schema_version": QUALITY_GATE_VERSION,
            "status": "MISSING_REPORT",
            "series": {},
            "path": path_to_str(path),
        }

    try:
        payload = load_json(path)
    except Exception as exc:
        return {
            "schema_version": QUALITY_GATE_VERSION,
            "status": "ERROR",
            "error": str(exc),
            "series": {},
            "path": path_to_str(path),
        }

    series_map = {}

    if isinstance(payload.get("series"), dict):
        series_map = payload["series"]

    elif isinstance(payload.get("series_by_id"), dict):
        series_map = payload["series_by_id"]

    elif isinstance(payload.get("results"), list):
        for item in payload["results"]:
            if isinstance(item, dict) and item.get("series_id"):
                series_map[item["series_id"]] = item

    payload["_normalized_series"] = series_map

    return payload


def get_series_quality_record(
    quality_report: Dict[str, Any],
    series_id: str,
) -> Dict[str, Any]:
    series_map = quality_report.get("_normalized_series") or quality_report.get("series") or {}

    if not isinstance(series_map, dict):
        series_map = {}

    record = series_map.get(series_id)

    if not isinstance(record, dict):
        return {
            "series_id": series_id,
            "status": DEFAULT_QUALITY_STATUS_IF_MISSING,
            "normalized_status": normalize_quality_status(DEFAULT_QUALITY_STATUS_IF_MISSING),
            "source": "DEFAULT_BECAUSE_MISSING_IN_QUALITY_REPORT",
            "checks": {},
        }

    raw_status = (
        record.get("status")
        or record.get("quality_status")
        or record.get("final_status")
        or DEFAULT_QUALITY_STATUS_IF_MISSING
    )

    out = dict(record)
    out["series_id"] = series_id
    out["normalized_status"] = normalize_quality_status(raw_status)

    return out


def can_process_quality_status(normalized_status: str) -> bool:
    if normalized_status == "FAIL" and BLOCK_ON_QUALITY_FAIL:
        return False

    return normalized_status in QUALITY_CAN_PROCESS_STATUSES


def get_ml_quality_status_from_record(quality_record: Optional[Dict[str, Any]]) -> str:
    if not isinstance(quality_record, dict):
        return "ML_CAUTION"
    status = str(quality_record.get("ml_quality_status") or "").strip().upper()
    if status in {"ML_READY", "ML_CAUTION", "ML_BLOCKED", "ML_NOT_APPLICABLE"}:
        return status
    normalized = normalize_quality_status(quality_record.get("status"))
    if normalized == "PASS":
        return "ML_READY"
    if normalized == "FAIL":
        return "ML_BLOCKED"
    return "ML_CAUTION"


def is_quality_usable_for_ml(quality_record: Optional[Dict[str, Any]], normalized_quality: str) -> bool:
    if isinstance(quality_record, dict) and "ml_usable_for_broad_training" in quality_record:
        return bool(quality_record.get("ml_usable_for_broad_training"))
    return normalize_quality_status(normalized_quality) in {"PASS", "WARNING"}


def validate_input_contract(df: pd.DataFrame, info: Dict[str, Any]) -> Dict[str, Any]:
    """
    Fase 1:
        Contrato formal de entrada entre camada 2 e camada 3.

    Não modifica df. Apenas audita.
    """
    report = {
        "contract_version": INPUT_CONTRACT_VERSION,
        "series_id": info.get("series_id"),
        "status": "PASS",
        "errors": [],
        "warnings": [],
        "checks": {},
    }

    missing = sorted(REQUIRED_INPUT_CONTRACT_COLUMNS - set(df.columns))

    report["checks"]["required_columns"] = sorted(REQUIRED_INPUT_CONTRACT_COLUMNS)
    report["checks"]["missing_columns"] = missing

    if missing:
        report["errors"].append(f"Missing required columns: {missing}")

    if df.empty:
        report["errors"].append("Input dataframe is empty.")

    if TIMESTAMP_UTC_MS_COL in df.columns:
        ts = pd.to_numeric(df[TIMESTAMP_UTC_MS_COL], errors="coerce")

        report["checks"]["timestamp_null_count"] = int(ts.isna().sum())
        report["checks"]["timestamp_duplicate_count"] = int(ts.duplicated().sum())
        report["checks"]["timestamp_monotonic_increasing"] = bool(ts.is_monotonic_increasing)

        if ts.isna().any():
            report["errors"].append("timestamp_utc_ms contains null values.")

        if ts.duplicated().any():
            report["errors"].append("timestamp_utc_ms contains duplicated values.")

        if not ts.is_monotonic_increasing:
            report["errors"].append("timestamp_utc_ms is not monotonic increasing.")

    if DATETIME_COL in df.columns:
        dt = pd.to_datetime(df[DATETIME_COL], errors="coerce")

        report["checks"]["datetime_null_count"] = int(dt.isna().sum())
        report["checks"]["datetime_duplicate_count"] = int(dt.duplicated().sum())
        report["checks"]["datetime_monotonic_increasing"] = bool(dt.is_monotonic_increasing)

        if dt.isna().any():
            report["errors"].append("DateTime contains null values.")

        if dt.duplicated().any():
            report["warnings"].append("DateTime contains duplicated values.")

        if not dt.is_monotonic_increasing:
            report["errors"].append("DateTime is not monotonic increasing.")

    for col in ["Open", "High", "Low", "Close", "Volume"]:
        if col not in df.columns:
            continue

        s = pd.to_numeric(df[col], errors="coerce")
        report["checks"][f"{col}_null_count"] = int(s.isna().sum())

        if col in {"Open", "High", "Low", "Close"}:
            non_positive = int((s <= 0).sum())
            report["checks"][f"{col}_non_positive_count"] = non_positive

            if non_positive > 0:
                report["errors"].append(f"{col} contains zero or negative values.")

        if col == "Volume":
            negative_volume = int((s < 0).sum())
            report["checks"]["volume_negative_count"] = negative_volume

            if negative_volume > 0:
                report["errors"].append("Volume contains negative values.")

    if all(c in df.columns for c in ["Open", "High", "Low", "Close"]):
        open_ = pd.to_numeric(df["Open"], errors="coerce")
        high = pd.to_numeric(df["High"], errors="coerce")
        low = pd.to_numeric(df["Low"], errors="coerce")
        close = pd.to_numeric(df["Close"], errors="coerce")

        high_lt_low = int((high < low).sum())
        close_outside = int(((close > high) | (close < low)).sum())
        open_outside = int(((open_ > high) | (open_ < low)).sum())

        report["checks"]["high_lt_low_count"] = high_lt_low
        report["checks"]["close_outside_low_high_count"] = close_outside
        report["checks"]["open_outside_low_high_count"] = open_outside

        if high_lt_low > 0:
            report["errors"].append("Found rows where High < Low.")

        if close_outside > 0:
            report["errors"].append("Found rows where Close is outside Low/High.")

        if open_outside > 0:
            report["errors"].append("Found rows where Open is outside Low/High.")

    if ENABLE_TIMEZONE_CONSISTENCY_CHECK:
        tz_report = validate_datetime_timestamp_consistency(df)
        report["checks"]["timezone_consistency"] = tz_report


        if tz_report.get("status") in {"FAIL", "ERROR"}:
            msg = "DateTime and timestamp_utc_ms timezone consistency failed."

            if BLOCK_ON_TIMEZONE_CONSISTENCY_FAIL:
                report["errors"].append(msg)
            else:
                report["warnings"].append(msg)

        elif tz_report.get("status") == "WARNING":
            report["warnings"].append("DateTime and timestamp_utc_ms timezone consistency warning.")





    if report["errors"]:
        report["status"] = "FAIL"
    elif report["warnings"]:
        report["status"] = "WARNING"
    else:
        report["status"] = "PASS"

    return report


def validate_datetime_timestamp_consistency(df: pd.DataFrame) -> Dict[str, Any]:
    """
    Confirma se DateTime Dubai naive corresponde ao timestamp_utc_ms.
    """
    report = {
        "status": "PASS",
        "max_abs_drift_ms": None,
        "mean_abs_drift_ms": None,
        "sample_size": 0,
        "errors": [],
        "warnings": [],
    }

    if DATETIME_COL not in df.columns or TIMESTAMP_UTC_MS_COL not in df.columns:
        report["status"] = "WARNING"
        report["warnings"].append("DateTime or timestamp_utc_ms missing.")
        return report

    try:
        dt_ms = datetime_dubai_naive_to_utc_ms(df[DATETIME_COL])
        ts_ms = normalize_timestamp_to_utc_ms(df[TIMESTAMP_UTC_MS_COL])


        comp = pd.DataFrame({
            "dt_ms": dt_ms,
            "ts_ms": ts_ms,
        }).dropna()

        if comp.empty:
            report["status"] = "FAIL"
            report["errors"].append("No valid rows to compare DateTime and timestamp_utc_ms.")
            return report

        drift = (comp["dt_ms"].astype("int64") - comp["ts_ms"].astype("int64")).abs()

        report["sample_size"] = int(len(comp))
        report["max_abs_drift_ms"] = int(drift.max())
        report["mean_abs_drift_ms"] = float(drift.mean())

        if report["max_abs_drift_ms"] > MAX_TIMESTAMP_DATETIME_DRIFT_MS:
            report["status"] = "FAIL"
            report["errors"].append(
                f"Max drift {report['max_abs_drift_ms']}ms exceeds "
                f"{MAX_TIMESTAMP_DATETIME_DRIFT_MS}ms."
            )

        return report

    except Exception as exc:
        report["status"] = "ERROR"
        report["errors"].append(str(exc))
        return report


def validate_time_grid(
    df: pd.DataFrame,
    timeframe: Optional[str],
    info: Dict[str, Any],
    max_gap_multiplier: float = POST_AUDIT_MAX_GAP_MULTIPLIER,
) -> Dict[str, Any]:
    """
    Fase 4:
        Validação temporal forte antes da geração de features.
    """
    expected_seconds = infer_timeframe_seconds(timeframe)
    expected_ms = None if expected_seconds is None else expected_seconds * 1000

    report = {
        "series_id": info.get("series_id"),
        "timeframe": timeframe,
        "expected_seconds": expected_seconds,
        "expected_ms": expected_ms,
        "status": "PASS",
        "errors": [],
        "warnings": [],
        "checks": {},
    }

    if expected_ms is None:
        report["status"] = "WARNING"
        report["warnings"].append("Could not infer timeframe seconds.")
        return report

    if TIMESTAMP_UTC_MS_COL not in df.columns:
        report["status"] = "FAIL"
        report["errors"].append("timestamp_utc_ms missing.")
        return report

    ts = normalize_timestamp_to_utc_ms(df[TIMESTAMP_UTC_MS_COL]).dropna().astype("int64")
    

    if len(ts) < 2:
        report["status"] = "FAIL"
        report["errors"].append("Not enough rows to validate time grid.")
        return report

    deltas = ts.diff().dropna()

    tolerance = max(TIME_GRID_IRREGULAR_TOLERANCE_MS, expected_ms * 0.001)
    irregular = (deltas - expected_ms).abs() > tolerance
    large_gaps = deltas > expected_ms * max_gap_multiplier

    irregular_count = int(irregular.sum())
    large_gap_count = int(large_gaps.sum())

    total_deltas = int(len(deltas))
    irregular_ratio = irregular_count / max(1, total_deltas)
    large_gap_ratio = large_gap_count / max(1, total_deltas)

    report["checks"] = {
        "rows": int(len(df)),
        "total_deltas": total_deltas,
        "median_delta_ms": float(deltas.median()),
        "min_delta_ms": int(deltas.min()),
        "max_delta_ms": int(deltas.max()),
        "expected_ms": int(expected_ms),
        "irregular_delta_count": irregular_count,
        "large_gap_count": large_gap_count,
        "irregular_ratio": round(irregular_ratio, 8),
        "large_gap_ratio": round(large_gap_ratio, 8),
        "max_gap_multiplier": max_gap_multiplier,
    }

    if irregular_count > 0:
        report["warnings"].append("Irregular candle deltas detected.")

    if large_gap_count > 0:
        report["warnings"].append("Large candle gaps detected.")

    if large_gap_ratio > TIME_GRID_FAIL_LARGE_GAP_RATIO_THRESHOLD:
        report["errors"].append(
            f"Large gap ratio {large_gap_ratio:.6f} exceeds threshold "
            f"{TIME_GRID_FAIL_LARGE_GAP_RATIO_THRESHOLD:.6f}."
        )

    if irregular_ratio > TIME_GRID_FAIL_IRREGULAR_RATIO_THRESHOLD:
        report["errors"].append(
            f"Irregular ratio {irregular_ratio:.6f} exceeds threshold "
            f"{TIME_GRID_FAIL_IRREGULAR_RATIO_THRESHOLD:.6f}."
        )

    if report["errors"]:
        report["status"] = "FAIL"
    elif report["warnings"]:
        report["status"] = "WARNING"
    else:
        report["status"] = "PASS"

    return report

















# =============================================================================
# 5. INDICADORES AUXILIARES
# =============================================================================

def add_feature_meta(
    metadata: List[Dict[str, Any]],
    feature: str,
    family: str,
    description: str,
    formula: str,
    lookback: int,
    risk_relevance: str = "",
    ml_relevance: str = "",
) -> None:
    metadata.append({
        "feature": feature,
        "family": family,
        "type_feature": family,
        "description": description,
        "formula": formula,
        "lookback": int(lookback) if lookback is not None else None,
        "uses_future_data": False,
        "risk_relevance": risk_relevance,
        "ml_relevance": ml_relevance,
    })


def true_range(high: pd.Series, low: pd.Series, close: pd.Series) -> pd.Series:
    close_prev = close.shift(1)
    a = high - low
    b = (high - close_prev).abs()
    c = (low - close_prev).abs()
    return pd.Series(np.maximum(np.maximum(a.to_numpy(), b.to_numpy()), c.to_numpy()), index=close.index)


def rolling_zscore(series: pd.Series, window: int) -> pd.Series:
    mean = series.rolling(window, min_periods=window).mean()
    std = series.rolling(window, min_periods=window).std()
    return safe_div(series - mean, std)


def stochastic_k(close: pd.Series, high: pd.Series, low: pd.Series, window: int) -> pd.Series:
    lowest = low.rolling(window, min_periods=window).min()
    highest = high.rolling(window, min_periods=window).max()
    return 100.0 * safe_div(close - lowest, highest - lowest)


def williams_r(close: pd.Series, high: pd.Series, low: pd.Series, window: int) -> pd.Series:
    lowest = low.rolling(window, min_periods=window).min()
    highest = high.rolling(window, min_periods=window).max()
    return -100.0 * safe_div(highest - close, highest - lowest)


def cci(close: pd.Series, high: pd.Series, low: pd.Series, window: int) -> pd.Series:
    typical_price = (high + low + close) / 3.0
    ma = typical_price.rolling(window, min_periods=window).mean()

    def _mad(x: np.ndarray) -> float:
        return float(np.mean(np.abs(x - np.mean(x))))

    mad = typical_price.rolling(window, min_periods=window).apply(_mad, raw=True)
    return safe_div(typical_price - ma, 0.015 * mad)


def load_python_environment() -> Dict[str, Any]:
    try:
        return load_json(PYTHON_ENVIRONMENT_PATH)
    except Exception:
        return {}


def resolve_feature_compute_backend(timeframe: Optional[str], row_count: int) -> Dict[str, Any]:
    python_env = load_python_environment()
    cuda = python_env.get("cuda", {}) if isinstance(python_env, dict) else {}
    cupy_status = cuda.get("cupy", {}) if isinstance(cuda, dict) else {}
    nvidia = cuda.get("nvidia_smi", {}) if isinstance(cuda, dict) else {}
    gpus = nvidia.get("gpus", []) if isinstance(nvidia, dict) else []
    memory_total_mb = None
    if gpus and isinstance(gpus[0], dict):
        memory_total_mb = gpus[0].get("memory_total_mb")

    plan = {
        "mode": FEATURE_CUDA_MODE,
        "resolved_backend": "pandas_cpu",
        "resolved_device": "cpu",
        "cuda_enabled": False,
        "reason": "CUDA desabilitado ou não selecionado.",
        "timeframe": timeframe,
        "row_count": int(row_count),
        "accelerated_blocks": [],
        "python_environment_path": path_to_str(PYTHON_ENVIRONMENT_PATH),
        "cupy_version": cupy_status.get("version") if isinstance(cupy_status, dict) else None,
        "nvidia_smi": nvidia,
        "auto_min_gpu_memory_mb": FEATURE_CUDA_AUTO_MIN_GPU_MEMORY_MB,
        "auto_min_rows": FEATURE_CUDA_AUTO_MIN_ROWS,
        "cuda_max_workers": FEATURE_CUDA_MAX_WORKERS,
        "note": FEATURE_CUDA_NOTE,
    }

    if FEATURE_CUDA_MODE in {"off", "false", "0", "cpu", "pandas", "pandas_cpu"}:
        plan["reason"] = "ARCHANGEL_FEATURE_CUDA_MODE força CPU."
        return plan

    cupy_ready = bool(
        isinstance(cuda, dict)
        and cuda.get("ready_for_code_migration")
        and isinstance(cupy_status, dict)
        and cupy_status.get("cuda_available")
    )
    if not cupy_ready:
        plan["reason"] = "CuPy/CUDA não está validado em 00_02_PYTHON_ENVIRONMENT_LATEST.json."
        return plan

    if FEATURE_CUDA_MODE == "auto":
        if int(row_count or 0) < FEATURE_CUDA_AUTO_MIN_ROWS:
            plan["reason"] = (
                "Auto conservador: série abaixo do limite mínimo de linhas para compensar transferência CPU/GPU; benchmark atual só ficou competitivo perto de 1M linhas."
            )
            return plan
        try:
            if memory_total_mb is not None and float(memory_total_mb) < FEATURE_CUDA_AUTO_MIN_GPU_MEMORY_MB:
                plan["reason"] = (
                    "Auto conservador: GPU atual abaixo do limite de memória para etapa 3 paralela; "
                    "use ARCHANGEL_FEATURE_CUDA_MODE=cuda para smoke tests controlados."
                )
                return plan
        except Exception:
            pass

    if FEATURE_CUDA_MODE not in {"auto", "cuda", "cupy", "gpu"}:
        plan["reason"] = f"Modo CUDA desconhecido: {FEATURE_CUDA_MODE}; usando CPU."
        return plan

    try:
        with warnings.catch_warnings():
            warnings.filterwarnings("ignore", message="CUDA path could not be detected.*")
            import cupy as cp

        if int(cp.cuda.runtime.getDeviceCount()) <= 0:
            plan["reason"] = "CuPy importou, mas não encontrou device CUDA."
            return plan
        test = cp.arange(4, dtype=cp.float32)
        _ = float(cp.sum(test).get())
    except Exception as exc:
        plan["reason"] = f"Falha ao validar CuPy no worker: {type(exc).__name__}: {exc}"
        return plan

    plan.update({
        "resolved_backend": "cupy_cuda",
        "resolved_device": "cuda:0",
        "cuda_enabled": True,
        "reason": "CuPy/CUDA validado para blocos vetoriais selecionados.",
        "accelerated_blocks": list(FEATURE_CUDA_ACCELERATED_BLOCKS),
    })
    return plan


def compute_return_core(
    close: pd.Series,
    backend_plan: Dict[str, Any],
) -> Tuple[pd.Series, pd.Series, pd.Series, Dict[str, Any]]:
    if backend_plan.get("resolved_backend") != "cupy_cuda":
        log_close = safe_log(close)
        ret_1 = log_close.diff(1)
        ret_simple_1 = close.pct_change(1)
        return log_close, ret_1, ret_simple_1, {
            "backend": "pandas_cpu",
            "status": "OK",
            "accelerated": False,
        }

    try:
        with warnings.catch_warnings():
            warnings.filterwarnings("ignore", message="CUDA path could not be detected.*")
            import cupy as cp

        index = close.index
        close_np = close.to_numpy(dtype="float64", copy=False)
        c = cp.asarray(close_np)
        log_c = cp.where(c > 0, cp.log(c), cp.nan)

        ret_1_gpu = cp.empty_like(log_c)
        ret_1_gpu[:1] = cp.nan
        ret_1_gpu[1:] = log_c[1:] - log_c[:-1]

        ret_simple_gpu = cp.empty_like(c)
        ret_simple_gpu[:1] = cp.nan
        prev = c[:-1]
        ret_simple_gpu[1:] = cp.where(prev != 0, (c[1:] / prev) - 1.0, cp.nan)

        log_close = pd.Series(cp.asnumpy(log_c), index=index)
        ret_1 = pd.Series(cp.asnumpy(ret_1_gpu), index=index)
        ret_simple_1 = pd.Series(cp.asnumpy(ret_simple_gpu), index=index)
        return log_close, ret_1, ret_simple_1, {
            "backend": "cupy_cuda",
            "status": "OK",
            "accelerated": True,
            "rows": int(len(close)),
        }
    except Exception as exc:
        log_close = safe_log(close)
        ret_1 = log_close.diff(1)
        ret_simple_1 = close.pct_change(1)
        return log_close, ret_1, ret_simple_1, {
            "backend": "pandas_cpu_fallback",
            "status": "FALLBACK",
            "accelerated": False,
            "error": f"{type(exc).__name__}: {exc}",
        }


def compute_log_diff(
    log_close: pd.Series,
    periods: int,
    backend_plan: Dict[str, Any],
) -> Tuple[pd.Series, Dict[str, Any]]:
    if backend_plan.get("resolved_backend") != "cupy_cuda":
        return log_close.diff(periods), {"backend": "pandas_cpu", "status": "OK"}

    try:
        with warnings.catch_warnings():
            warnings.filterwarnings("ignore", message="CUDA path could not be detected.*")
            import cupy as cp

        values = cp.asarray(log_close.to_numpy(dtype="float64", copy=False))
        out = cp.empty_like(values)
        out[:periods] = cp.nan
        out[periods:] = values[periods:] - values[:-periods]
        return pd.Series(cp.asnumpy(out), index=log_close.index), {
            "backend": "cupy_cuda",
            "status": "OK",
            "periods": int(periods),
        }
    except Exception as exc:
        return log_close.diff(periods), {
            "backend": "pandas_cpu_fallback",
            "status": "FALLBACK",
            "periods": int(periods),
            "error": f"{type(exc).__name__}: {exc}",
        }


def _rolling_cpu_report(kind: str, window: int, *, status: str = "OK", rows: Optional[int] = None, reason: Optional[str] = None) -> Dict[str, Any]:
    report = {
        "backend": "pandas_cpu",
        "status": status,
        "kind": kind,
        "window": int(window),
        "accelerated": False,
    }
    if rows is not None:
        report["rows"] = int(rows)
    if reason:
        report["reason"] = reason
    return report


def compute_rolling_stats(
    series: pd.Series,
    window: int,
    backend_plan: Dict[str, Any],
    *,
    need_sum: bool = False,
    need_mean: bool = False,
    need_std: bool = False,
) -> Tuple[Dict[str, pd.Series], Dict[str, Any]]:
    window = int(window)
    if backend_plan.get("resolved_backend") != "cupy_cuda":
        roll = series.rolling(window, min_periods=window)
        output: Dict[str, pd.Series] = {}
        if need_sum:
            output["sum"] = roll.sum()
        if need_mean:
            output["mean"] = roll.mean()
        if need_std:
            output["std"] = roll.std()
        return output, _rolling_cpu_report("stats", window)

    n_local = int(len(series))
    if FEATURE_CUDA_MODE == "auto" and n_local < FEATURE_CUDA_AUTO_MIN_ROWS:
        roll = series.rolling(window, min_periods=window)
        output: Dict[str, pd.Series] = {}
        if need_sum:
            output["sum"] = roll.sum()
        if need_mean:
            output["mean"] = roll.mean()
        if need_std:
            output["std"] = roll.std()
        return output, _rolling_cpu_report(
            "stats",
            window,
            status="CPU_BELOW_CUDA_ROW_THRESHOLD",
            rows=n_local,
            reason=f"rows<{FEATURE_CUDA_AUTO_MIN_ROWS}; avoiding CPU/GPU transfer overhead on current GPU",
        )

    try:
        with warnings.catch_warnings():
            warnings.filterwarnings("ignore", message="CUDA path could not be detected.*")
            import cupy as cp

        index = series.index
        values = cp.asarray(series.to_numpy(dtype="float64", copy=False))
        n = int(values.size)
        output_arrays: Dict[str, Any] = {}

        if n == 0 or window <= 0:
            empty = pd.Series(np.nan, index=index, dtype="float64")
            output = {}
            if need_sum:
                output["sum"] = empty.copy()
            if need_mean:
                output["mean"] = empty.copy()
            if need_std:
                output["std"] = empty.copy()
            return output, {
                "backend": "cupy_cuda",
                "status": "OK",
                "kind": "stats",
                "window": window,
                "rows": n,
                "accelerated": True,
            }

        valid = cp.isfinite(values)
        values0 = cp.where(valid, values, 0.0)
        center = None
        centered0 = None
        centered_sums = None
        if need_std:
            if bool(cp.any(valid).get()):
                center = cp.mean(values[valid])
            else:
                center = cp.asarray(0.0, dtype=cp.float64)
            centered0 = cp.where(valid, values - center, 0.0)
            centered_csum = cp.concatenate([cp.zeros(1, dtype=cp.float64), cp.cumsum(centered0)])
            if n >= window:
                centered_sums = centered_csum[window:] - centered_csum[:-window]
        counts_csum = cp.concatenate([cp.zeros(1, dtype=cp.int64), cp.cumsum(valid.astype(cp.int64))])
        values_csum = cp.concatenate([cp.zeros(1, dtype=cp.float64), cp.cumsum(values0)])
        counts = counts_csum[window:] - counts_csum[:-window] if n >= window else cp.asarray([], dtype=cp.int64)
        sums = values_csum[window:] - values_csum[:-window] if n >= window else cp.asarray([], dtype=cp.float64)
        full_mask = counts == window

        if need_sum or need_mean:
            sum_out = cp.full(n, cp.nan, dtype=cp.float64)
            if n >= window:
                sum_out[window - 1:] = cp.where(full_mask, sums, cp.nan)
            if need_sum:
                output_arrays["sum"] = sum_out
            if need_mean:
                mean_out = cp.full(n, cp.nan, dtype=cp.float64)
                if n >= window:
                    if need_std and centered_sums is not None and center is not None:
                        mean_values = center + (centered_sums / float(window))
                    else:
                        mean_values = sums / float(window)
                    mean_out[window - 1:] = cp.where(full_mask, mean_values, cp.nan)
                output_arrays["mean"] = mean_out

        if need_std:
            std_out = cp.full(n, cp.nan, dtype=cp.float64)
            if n >= window and window > 1 and centered0 is not None and centered_sums is not None:
                centered2_csum = cp.concatenate([cp.zeros(1, dtype=cp.float64), cp.cumsum(centered0 * centered0)])
                centered_sums2 = centered2_csum[window:] - centered2_csum[:-window]
                variance_num = centered_sums2 - (centered_sums * centered_sums / float(window))
                variance = cp.maximum(variance_num / float(window - 1), 0.0)
                std_out[window - 1:] = cp.where(full_mask, cp.sqrt(variance), cp.nan)
            output_arrays["std"] = std_out

        output = {name: pd.Series(cp.asnumpy(array), index=index) for name, array in output_arrays.items()}
        return output, {
            "backend": "cupy_cuda",
            "status": "OK",
            "kind": "stats",
            "window": window,
            "rows": n,
            "accelerated": True,
            "metrics": sorted(output.keys()),
        }
    except Exception as exc:
        roll = series.rolling(window, min_periods=window)
        output = {}
        if need_sum:
            output["sum"] = roll.sum()
        if need_mean:
            output["mean"] = roll.mean()
        if need_std:
            output["std"] = roll.std()
        return output, {
            "backend": "pandas_cpu_fallback",
            "status": "FALLBACK",
            "kind": "stats",
            "window": window,
            "accelerated": False,
            "error": f"{type(exc).__name__}: {exc}",
        }


def compute_rolling_corr(
    left: pd.Series,
    right: pd.Series,
    window: int,
    backend_plan: Dict[str, Any],
) -> Tuple[pd.Series, Dict[str, Any]]:
    window = int(window)
    if backend_plan.get("resolved_backend") != "cupy_cuda":
        return left.rolling(window, min_periods=window).corr(right), _rolling_cpu_report("corr", window)

    n_local = int(len(left))
    if FEATURE_CUDA_MODE == "auto" and n_local < FEATURE_CUDA_AUTO_MIN_ROWS:
        return left.rolling(window, min_periods=window).corr(right), _rolling_cpu_report(
            "corr",
            window,
            status="CPU_BELOW_CUDA_ROW_THRESHOLD",
            rows=n_local,
            reason=f"rows<{FEATURE_CUDA_AUTO_MIN_ROWS}; avoiding CPU/GPU transfer overhead on current GPU",
        )

    try:
        with warnings.catch_warnings():
            warnings.filterwarnings("ignore", message="CUDA path could not be detected.*")
            import cupy as cp

        index = left.index
        x = cp.asarray(left.to_numpy(dtype="float64", copy=False))
        y = cp.asarray(right.to_numpy(dtype="float64", copy=False))
        n = int(x.size)
        out = cp.full(n, cp.nan, dtype=cp.float64)

        if n >= window and window > 1:
            valid = cp.isfinite(x) & cp.isfinite(y)
            if bool(cp.any(valid).get()):
                x_center = cp.mean(x[valid])
                y_center = cp.mean(y[valid])
            else:
                x_center = cp.asarray(0.0, dtype=cp.float64)
                y_center = cp.asarray(0.0, dtype=cp.float64)
            x0 = cp.where(valid, x - x_center, 0.0)
            y0 = cp.where(valid, y - y_center, 0.0)

            def csum(values):
                return cp.concatenate([cp.zeros(1, dtype=values.dtype), cp.cumsum(values)])

            counts_csum = csum(valid.astype(cp.int64))
            x_csum = csum(x0)
            y_csum = csum(y0)
            x2_csum = csum(x0 * x0)
            y2_csum = csum(y0 * y0)
            xy_csum = csum(x0 * y0)

            counts = counts_csum[window:] - counts_csum[:-window]
            sx = x_csum[window:] - x_csum[:-window]
            sy = y_csum[window:] - y_csum[:-window]
            sx2 = x2_csum[window:] - x2_csum[:-window]
            sy2 = y2_csum[window:] - y2_csum[:-window]
            sxy = xy_csum[window:] - xy_csum[:-window]

            cov_num = sxy - (sx * sy / float(window))
            var_x = sx2 - (sx * sx / float(window))
            var_y = sy2 - (sy * sy / float(window))
            denom = cp.sqrt(cp.maximum(var_x, 0.0) * cp.maximum(var_y, 0.0))
            mask = (counts == window) & (denom > 0)
            out[window - 1:] = cp.where(mask, cov_num / denom, cp.nan)

        return pd.Series(cp.asnumpy(out), index=index), {
            "backend": "cupy_cuda",
            "status": "OK",
            "kind": "corr",
            "window": window,
            "rows": n,
            "accelerated": True,
        }
    except Exception as exc:
        return left.rolling(window, min_periods=window).corr(right), {
            "backend": "pandas_cpu_fallback",
            "status": "FALLBACK",
            "kind": "corr",
            "window": window,
            "accelerated": False,
            "error": f"{type(exc).__name__}: {exc}",
        }


def append_cuda_rolling_report(
    execution_steps: List[Dict[str, Any]],
    source_family: str,
    reports: List[Dict[str, Any]],
) -> None:
    if not reports:
        return

    accelerated = [report for report in reports if report.get("backend") == "cupy_cuda"]
    fallbacks = [report for report in reports if report.get("status") == "FALLBACK"]
    cpu = [report for report in reports if report.get("backend") == "pandas_cpu"]

    cpu_policy = [report for report in cpu if report.get("status") != "OK"]
    if not accelerated and not fallbacks and not cpu_policy:
        return

    execution_steps.append({
        "family": "__cuda_rolling__",
        "source_family": source_family,
        "status": "FALLBACK" if fallbacks else ("CPU_POLICY" if cpu_policy and not accelerated else "OK"),
        "elapsed_seconds": 0.0,
        "backend": "cupy_cuda" if accelerated else ("pandas_cpu_fallback" if fallbacks else "pandas_cpu"),
        "accelerated": bool(accelerated),
        "accelerated_count": len(accelerated),
        "fallback_count": len(fallbacks),
        "cpu_count": len(cpu),
        "cpu_policy_count": len(cpu_policy),
        "windows": sorted({int(report.get("window")) for report in reports if report.get("window") is not None}),
        "kinds": sorted({str(report.get("kind")) for report in reports if report.get("kind")}),
        "errors_first_5": [str(report.get("error")) for report in fallbacks if report.get("error")][:5],
        "cpu_policy_reasons_first_5": [str(report.get("reason")) for report in cpu_policy if report.get("reason")][:5],
    })


# =============================================================================
# 6. GERAÇÃO DAS FEATURES
# =============================================================================

def generate_all_features(
    df: pd.DataFrame,
    timeframe: Optional[str] = None,
) -> Tuple[pd.DataFrame, List[Dict[str, Any]], List[Dict[str, Any]]]:

    metadata: List[Dict[str, Any]] = []
    execution_steps: List[Dict[str, Any]] = []
    features: Dict[str, Any] = {}

    FEATURE_CONFIG, profile_report = resolve_feature_config_for_series(timeframe, len(df))
    execution_steps.append({
        "family": "__profile__",
        "status": "OK",
        "elapsed_seconds": 0.0,
        **profile_report,
    })
    backend_plan = resolve_feature_compute_backend(timeframe, len(df))
    execution_steps.append({
        "family": "__compute_backend__",
        "status": "OK",
        "elapsed_seconds": 0.0,
        **backend_plan,
    })

    turbo = is_turbo_timeframe(timeframe)

    close = df["Close"].astype(float)
    open_ = df["Open"].astype(float)
    high = df["High"].astype(float)
    low = df["Low"].astype(float)
    volume = df["Volume"].astype(float)

    ann_factor = annualization_factor_from_timeframe(timeframe)
    sqrt_ann = math.sqrt(ann_factor)

    log_close, ret_1, ret_simple_1, return_core_report = compute_return_core(close, backend_plan)
    execution_steps.append({
        "family": "__cuda_vector_core__",
        "status": return_core_report.get("status", "OK"),
        "elapsed_seconds": 0.0,
        **return_core_report,
    })
    abs_ret_1 = ret_1.abs()
    ret_1_sq = ret_1 ** 2
    if backend_plan.get("resolved_backend") == "cupy_cuda":
        log_close_sensitive_cpu = safe_log(close)
        ret_1_sensitive_cpu = log_close_sensitive_cpu.diff(1)
        abs_ret_1_sensitive_cpu = ret_1_sensitive_cpu.abs()
    else:
        log_close_sensitive_cpu = log_close
        ret_1_sensitive_cpu = ret_1
        abs_ret_1_sensitive_cpu = abs_ret_1

    # -------------------------------------------------------------------------
    # RETURNS
    # -------------------------------------------------------------------------
    family = "returns"
    if FEATURE_CONFIG[family]["enabled"]:
        start = time.time()

        features["feat_log_close"] = log_close
        add_feature_meta(metadata, "feat_log_close", family, "Log natural do Close.", "ln(Close)", 0)

        for w in filter_windows_for_timeframe(FEATURE_CONFIG[family]["windows"], timeframe):
            col = f"feat_ret_log_{w}"
            diff_series, diff_report = compute_log_diff(log_close, w, backend_plan)
            features[col] = diff_series
            if diff_report.get("status") == "FALLBACK":
                execution_steps.append({
                    "family": "__cuda_vector_core__",
                    "status": "FALLBACK",
                    "elapsed_seconds": 0.0,
                    **diff_report,
                })
            add_feature_meta(metadata, col, family, f"Retorno logarítmico em {w} candles.", f"ln(C_t/C_t-{w})", w)

        features["feat_ret_simple_1"] = ret_simple_1
        add_feature_meta(metadata, "feat_ret_simple_1", family, "Retorno simples de 1 candle.", "Close/Close.shift(1)-1", 1)

        execution_steps.append({"family": family, "status": "OK", "elapsed_seconds": round(time.time() - start, 6)})

    # -------------------------------------------------------------------------
    # RETURN LAGS
    # -------------------------------------------------------------------------
    family = "return_lags"
    if FEATURE_CONFIG[family]["enabled"]:
        start = time.time()

        for lag in FEATURE_CONFIG[family]["lags"]:
            col = f"feat_ret_log_1_lag_{lag}"
            features[col] = ret_1.shift(lag)
            add_feature_meta(metadata, col, family, f"Lag {lag} do retorno log 1.", f"ret_1.shift({lag})", lag)

        execution_steps.append({"family": family, "status": "OK", "elapsed_seconds": round(time.time() - start, 6)})

    # -------------------------------------------------------------------------
    # ROLLING RETURN
    # -------------------------------------------------------------------------
    family = "rolling_return"
    if FEATURE_CONFIG[family]["enabled"]:
        start = time.time()
        rolling_reports: List[Dict[str, Any]] = []

        for w in filter_windows_for_timeframe(FEATURE_CONFIG[family]["windows"], timeframe):
            stats, report = compute_rolling_stats(ret_1, w, backend_plan, need_sum=True, need_mean=True)
            rolling_reports.append(report)
            features[f"feat_roll_ret_sum_{w}"] = stats["sum"]
            features[f"feat_roll_ret_mean_{w}"] = stats["mean"]

            add_feature_meta(metadata, f"feat_roll_ret_sum_{w}", family, f"Soma móvel do retorno em {w}.", f"sum(ret,{w})", w)
            add_feature_meta(metadata, f"feat_roll_ret_mean_{w}", family, f"Média móvel do retorno em {w}.", f"mean(ret,{w})", w)

        execution_steps.append({"family": family, "status": "OK", "elapsed_seconds": round(time.time() - start, 6)})
        append_cuda_rolling_report(execution_steps, family, rolling_reports)

    # -------------------------------------------------------------------------
    # VOLATILITY
    # -------------------------------------------------------------------------
    family = "volatility"
    vol_cache: Dict[int, pd.Series] = {}

    if FEATURE_CONFIG[family]["enabled"]:
        start = time.time()
        rolling_reports: List[Dict[str, Any]] = []

        for w in filter_windows_for_timeframe(FEATURE_CONFIG[family]["windows"], timeframe):
            ret_stats, ret_report = compute_rolling_stats(ret_1, w, backend_plan, need_std=True)
            abs_stats, abs_report = compute_rolling_stats(abs_ret_1, w, backend_plan, need_mean=True)
            sq_stats, sq_report = compute_rolling_stats(ret_1_sq, w, backend_plan, need_sum=True)
            rolling_reports.extend([ret_report, abs_report, sq_report])

            rv = ret_stats["std"]
            vol_cache[w] = rv

            features[f"feat_rv_std_{w}"] = rv
            features[f"feat_rv_abs_mean_{w}"] = abs_stats["mean"]
            features[f"feat_rv_quadratic_{w}"] = np.sqrt(sq_stats["sum"])

            add_feature_meta(metadata, f"feat_rv_std_{w}", family, f"Volatilidade realizada {w}.", f"std(ret,{w})", w)
            add_feature_meta(metadata, f"feat_rv_abs_mean_{w}", family, f"Média abs retorno {w}.", f"mean(abs(ret),{w})", w)
            add_feature_meta(metadata, f"feat_rv_quadratic_{w}", family, f"Vol quadrática {w}.", f"sqrt(sum(ret^2,{w}))", w)

            if FEATURE_CONFIG[family].get("annualization", False):
                features[f"feat_rv_std_{w}_ann"] = rv * sqrt_ann
                add_feature_meta(metadata, f"feat_rv_std_{w}_ann", family, f"Vol anualizada {w}.", "rv*sqrt(AF)", w)

        execution_steps.append({"family": family, "status": "OK", "elapsed_seconds": round(time.time() - start, 6)})
        append_cuda_rolling_report(execution_steps, family, rolling_reports)

    # -------------------------------------------------------------------------
    # VOL OF VOL
    # -------------------------------------------------------------------------
    family = "vol_of_vol"
    if FEATURE_CONFIG[family]["enabled"]:
        start = time.time()
        rolling_reports: List[Dict[str, Any]] = []

        for w in filter_windows_for_timeframe(FEATURE_CONFIG[family]["windows"], timeframe):
            base_vol = vol_cache.get(w)
            if base_vol is None:
                base_stats, base_report = compute_rolling_stats(ret_1, w, backend_plan, need_std=True)
                rolling_reports.append(base_report)
                base_vol = base_stats["std"]

            col = f"feat_vol_of_vol_{w}"
            vv_stats, vv_report = compute_rolling_stats(base_vol, w, backend_plan, need_std=True)
            rolling_reports.append(vv_report)
            features[col] = vv_stats["std"]
            add_feature_meta(metadata, col, family, f"Volatilidade da volatilidade {w}.", f"std(rv_{w},{w})", 2 * w)

        execution_steps.append({"family": family, "status": "OK", "elapsed_seconds": round(time.time() - start, 6)})
        append_cuda_rolling_report(execution_steps, family, rolling_reports)

    # -------------------------------------------------------------------------
    # EMA
    # -------------------------------------------------------------------------
    family = "ema"
    if FEATURE_CONFIG[family]["enabled"]:
        start = time.time()

        ema_map = {}
        windows = filter_windows_for_timeframe(sorted(FEATURE_CONFIG[family]["windows"]), timeframe)

        for w in windows:
            ema = close.ewm(span=w, adjust=False, min_periods=max(2, w // 2)).mean()
            ema_map[w] = ema

            features[f"feat_ema_{w}"] = ema
            features[f"feat_close_ema_{w}_ratio"] = safe_div(close, ema) - 1.0

            add_feature_meta(metadata, f"feat_ema_{w}", family, f"EMA {w}.", f"EMA(Close,{w})", w)
            add_feature_meta(metadata, f"feat_close_ema_{w}_ratio", family, f"Distância Close/EMA {w}.", f"Close/EMA_{w}-1", w)

        for fast, slow in [(3, 5), (5, 10), (7, 14), (10, 20), (20, 50), (50, 100), (100, 200)]:
            if fast in ema_map and slow in ema_map:
                col = f"feat_ema_{fast}_{slow}_spread"
                features[col] = safe_div(ema_map[fast], ema_map[slow]) - 1.0
                add_feature_meta(metadata, col, family, f"Spread EMA {fast}/{slow}.", f"EMA_{fast}/EMA_{slow}-1", slow)

        execution_steps.append({"family": family, "status": "OK", "elapsed_seconds": round(time.time() - start, 6)})

    # -------------------------------------------------------------------------
    # SMA
    # -------------------------------------------------------------------------
    family = "sma"
    if FEATURE_CONFIG[family]["enabled"]:
        start = time.time()
        rolling_reports: List[Dict[str, Any]] = []

        for w in filter_windows_for_timeframe(FEATURE_CONFIG[family]["windows"], timeframe):
            stats, report = compute_rolling_stats(close, w, backend_plan, need_mean=True)
            rolling_reports.append(report)
            sma = stats["mean"]

            features[f"feat_sma_{w}"] = sma
            features[f"feat_close_sma_{w}_ratio"] = safe_div(close, sma) - 1.0

            add_feature_meta(metadata, f"feat_sma_{w}", family, f"SMA {w}.", f"SMA(Close,{w})", w)
            add_feature_meta(metadata, f"feat_close_sma_{w}_ratio", family, f"Distância Close/SMA {w}.", f"Close/SMA_{w}-1", w)

        execution_steps.append({"family": family, "status": "OK", "elapsed_seconds": round(time.time() - start, 6)})
        append_cuda_rolling_report(execution_steps, family, rolling_reports)

    # -------------------------------------------------------------------------
    # BOLLINGER
    # -------------------------------------------------------------------------
    family = "bollinger"
    if FEATURE_CONFIG[family]["enabled"]:
        start = time.time()
        num_std = float(FEATURE_CONFIG[family]["num_std"])

        for w in filter_windows_for_timeframe(FEATURE_CONFIG[family]["windows"], timeframe):
            # Mantido em Pandas/CPU: std de preço pode sofrer cancelamento numérico em CUDA cumulativo.
            mean = close.rolling(w, min_periods=w).mean()
            std = close.rolling(w, min_periods=w).std()
            upper = mean + num_std * std
            lower = mean - num_std * std

            col_map = {
                f"feat_bb_z_{w}": safe_div(close - mean, std),
                f"feat_bb_width_{w}": safe_div(upper - lower, mean),
                f"feat_bb_upper_dist_{w}": safe_div(close, upper) - 1.0,
                f"feat_bb_lower_dist_{w}": safe_div(close, lower) - 1.0,
                f"feat_bb_percent_b_{w}": safe_div(close - lower, upper - lower),
            }

            features.update(col_map)

            for col in col_map:
                add_feature_meta(metadata, col, family, f"Bollinger {w}.", f"Bollinger({w})", w)

        execution_steps.append({"family": family, "status": "OK", "elapsed_seconds": round(time.time() - start, 6)})

    # -------------------------------------------------------------------------
    # RSI
    # -------------------------------------------------------------------------
    family = "rsi"
    if FEATURE_CONFIG[family]["enabled"]:
        start = time.time()

        delta = close.diff()
        gain = delta.clip(lower=0)
        loss = -delta.clip(upper=0)

        for w in FEATURE_CONFIG[family]["windows"]:
            avg_gain = gain.ewm(alpha=1 / w, adjust=False, min_periods=w).mean()
            avg_loss = loss.ewm(alpha=1 / w, adjust=False, min_periods=w).mean()
            rs = safe_div(avg_gain, avg_loss.replace(0, np.nan))
            rsi = 100.0 - (100.0 / (1.0 + rs))

            features[f"feat_rsi_{w}"] = rsi
            features[f"feat_rsi_{w}_norm"] = (rsi - 50.0) / 50.0

            add_feature_meta(metadata, f"feat_rsi_{w}", family, f"RSI {w}.", f"RSI({w})", w)
            add_feature_meta(metadata, f"feat_rsi_{w}_norm", family, f"RSI normalizado {w}.", f"(RSI-50)/50", w)

        execution_steps.append({"family": family, "status": "OK", "elapsed_seconds": round(time.time() - start, 6)})

    # -------------------------------------------------------------------------
    # MACD
    # -------------------------------------------------------------------------
    family = "macd"
    if FEATURE_CONFIG[family]["enabled"]:
        start = time.time()

        fast = int(FEATURE_CONFIG[family]["fast"])
        slow = int(FEATURE_CONFIG[family]["slow"])
        signal = int(FEATURE_CONFIG[family]["signal"])

        ema_fast = close.ewm(span=fast, adjust=False, min_periods=fast).mean()
        ema_slow = close.ewm(span=slow, adjust=False, min_periods=slow).mean()
        macd = ema_fast - ema_slow
        macd_signal = macd.ewm(span=signal, adjust=False, min_periods=signal).mean()
        macd_hist = macd - macd_signal

        features["feat_macd"] = macd
        features["feat_macd_signal"] = macd_signal
        features["feat_macd_hist"] = macd_hist
        features["feat_macd_close_ratio"] = safe_div(macd, close)
        features["feat_macd_hist_close_ratio"] = safe_div(macd_hist, close)

        for col in ["feat_macd", "feat_macd_signal", "feat_macd_hist", "feat_macd_close_ratio", "feat_macd_hist_close_ratio"]:
            add_feature_meta(metadata, col, family, col, "MACD family", slow + signal)

        execution_steps.append({"family": family, "status": "OK", "elapsed_seconds": round(time.time() - start, 6)})

    # -------------------------------------------------------------------------
    # RANGE CANDLE
    # -------------------------------------------------------------------------
    family = "range_candle"
    if FEATURE_CONFIG[family]["enabled"]:
        start = time.time()

        candle_range = (high - low).replace(0, np.nan)
        body = close - open_
        body_abs = body.abs()

        oc_max = np.maximum(open_, close)
        oc_min = np.minimum(open_, close)

        upper_wick = high - oc_max
        lower_wick = oc_min - low

        feature_map = {
            "feat_hl_range": safe_div(high, low) - 1.0,
            "feat_oc_return": safe_div(close, open_) - 1.0,
            "feat_body_size_ratio": safe_div(body_abs, candle_range),
            "feat_upper_wick_ratio": safe_div(upper_wick, candle_range),
            "feat_lower_wick_ratio": safe_div(lower_wick, candle_range),
            "feat_close_position_in_range": safe_div(close - low, candle_range),
            "feat_candle_direction": np.sign(body),
            "feat_body_signed_ratio": safe_div(body, candle_range),
        }

        features.update(feature_map)

        for col in feature_map:
            add_feature_meta(metadata, col, family, col, "OHLC candle geometry", 0)

        execution_steps.append({"family": family, "status": "OK", "elapsed_seconds": round(time.time() - start, 6)})

    # -------------------------------------------------------------------------
    # VOLUME
    # -------------------------------------------------------------------------
    family = "volume"
    if FEATURE_CONFIG[family]["enabled"]:
        start = time.time()
        rolling_reports: List[Dict[str, Any]] = []

        vol = volume.replace(0, np.nan)
        log_volume = np.log(vol)
        log1p_volume = np.log1p(volume)

        features["feat_log_volume"] = log_volume
        features["feat_log1p_volume"] = log1p_volume
        features["feat_volume_ret_1"] = log_volume.diff(1)

        add_feature_meta(metadata, "feat_log_volume", family, "Log volume.", "ln(volume)", 0)
        add_feature_meta(metadata, "feat_log1p_volume", family, "Log1p volume.", "ln(1+volume)", 0)
        add_feature_meta(metadata, "feat_volume_ret_1", family, "Retorno log do volume.", "diff(ln(volume))", 1)

        for w in filter_windows_for_timeframe(FEATURE_CONFIG[family]["windows"], timeframe):
            log_stats, log_report = compute_rolling_stats(log_volume, w, backend_plan, need_mean=True, need_std=True)
            vol_stats, vol_report = compute_rolling_stats(vol, w, backend_plan, need_mean=True)
            rolling_reports.extend([log_report, vol_report])
            mean = log_stats["mean"]
            std = log_stats["std"]
            vol_ma = vol_stats["mean"]

            features[f"feat_volume_z_{w}"] = safe_div(log_volume - mean, std)
            features[f"feat_volume_ma_ratio_{w}"] = safe_div(vol, vol_ma) - 1.0

            add_feature_meta(metadata, f"feat_volume_z_{w}", family, f"Z-score volume {w}.", "zscore(logvol)", w)
            add_feature_meta(metadata, f"feat_volume_ma_ratio_{w}", family, f"Volume/SMA volume {w}.", "volume/sma(volume)-1", w)

            if not (turbo and TURBO_DISABLE_VOLUME_ROLLING_CORR):
                corr, corr_report = compute_rolling_corr(abs_ret_1, log_volume, w, backend_plan)
                rolling_reports.append(corr_report)
                features[f"feat_absret_volume_corr_{w}"] = corr
                add_feature_meta(metadata, f"feat_absret_volume_corr_{w}", family, f"Corr absret-volume {w}.", "corr(absret,logvol)", w)

        execution_steps.append({"family": family, "status": "OK", "elapsed_seconds": round(time.time() - start, 6)})
        append_cuda_rolling_report(execution_steps, family, rolling_reports)

    # -------------------------------------------------------------------------
    # DRAWDOWN DISTANCE
    # -------------------------------------------------------------------------
    family = "drawdown_distance"
    if FEATURE_CONFIG[family]["enabled"]:
        start = time.time()

        for w in filter_windows_for_timeframe(FEATURE_CONFIG[family]["windows"], timeframe):
            # Rolling min/max permanece em CPU nesta fase para evitar uso de memoria imprevisivel na GTX 1660 Ti.
            rolling_max = close.rolling(w, min_periods=w).max()
            rolling_min = close.rolling(w, min_periods=w).min()
            rolling_range = (rolling_max - rolling_min).replace(0, np.nan)

            features[f"feat_dist_roll_max_{w}"] = safe_div(close, rolling_max) - 1.0
            features[f"feat_dist_roll_min_{w}"] = safe_div(close, rolling_min) - 1.0
            features[f"feat_position_roll_range_{w}"] = safe_div(close - rolling_min, rolling_range)

            add_feature_meta(metadata, f"feat_dist_roll_max_{w}", family, f"Distância máxima móvel {w}.", "Close/max-1", w)
            add_feature_meta(metadata, f"feat_dist_roll_min_{w}", family, f"Distância mínima móvel {w}.", "Close/min-1", w)
            add_feature_meta(metadata, f"feat_position_roll_range_{w}", family, f"Posição no range {w}.", "(Close-min)/(max-min)", w)

        execution_steps.append({"family": family, "status": "OK", "elapsed_seconds": round(time.time() - start, 6)})

    # -------------------------------------------------------------------------
    # TREND STRENGTH
    # -------------------------------------------------------------------------
    family = "trend_strength"
    if FEATURE_CONFIG[family]["enabled"]:
        start = time.time()
        rolling_reports: List[Dict[str, Any]] = []

        for w in filter_windows_for_timeframe(FEATURE_CONFIG[family]["windows"], timeframe):
            vol = vol_cache.get(w)
            if vol is None:
                vol_stats, vol_report = compute_rolling_stats(ret_1, w, backend_plan, need_std=True)
                rolling_reports.append(vol_report)
                vol = vol_stats["std"]

            col = f"feat_trend_strength_{w}"
            features[col] = safe_div(log_close.diff(w), vol * np.sqrt(w))
            add_feature_meta(metadata, col, family, f"Força tendência {w}.", f"ret_{w}/(std*sqrt(w))", w)

        execution_steps.append({"family": family, "status": "OK", "elapsed_seconds": round(time.time() - start, 6)})
        append_cuda_rolling_report(execution_steps, family, rolling_reports)

    # -------------------------------------------------------------------------
    # AUTOCORRELATION
    # -------------------------------------------------------------------------
    family = "autocorrelation"
    if FEATURE_CONFIG[family]["enabled"] and not (turbo and TURBO_DISABLE_AUTOCORR):
        start = time.time()

        for w in filter_windows_for_timeframe(FEATURE_CONFIG[family]["windows"], timeframe):
            roll = ret_1_sensitive_cpu.rolling(w, min_periods=w)
            for lag in FEATURE_CONFIG[family]["lags"]:
                col = f"feat_ret_autocorr_w{w}_lag{lag}"
                features[col] = roll.corr(ret_1_sensitive_cpu.shift(lag))
                add_feature_meta(metadata, col, family, f"Autocorr ret w{w} lag{lag}.", "rolling corr", w + lag)

        execution_steps.append({"family": family, "status": "OK", "elapsed_seconds": round(time.time() - start, 6)})
    elif FEATURE_CONFIG[family]["enabled"]:
        execution_steps.append({"family": family, "status": "SKIPPED_TURBO", "elapsed_seconds": 0.0})

    # -------------------------------------------------------------------------
    # ATR
    # -------------------------------------------------------------------------
    family = "atr_range_volatility"
    if FEATURE_CONFIG[family]["enabled"]:
        start = time.time()

        tr = true_range(high, low, close)
        features["feat_true_range"] = tr
        features["feat_true_range_close_ratio"] = safe_div(tr, close)

        add_feature_meta(metadata, "feat_true_range", family, "True Range.", "TR", 1)
        add_feature_meta(metadata, "feat_true_range_close_ratio", family, "TR/Close.", "TR/Close", 1)

        rolling_reports: List[Dict[str, Any]] = []
        for w in FEATURE_CONFIG[family]["windows"]:
            atr_stats, atr_report = compute_rolling_stats(tr, w, backend_plan, need_mean=True)
            rolling_reports.append(atr_report)
            atr = atr_stats["mean"]

            features[f"feat_atr_{w}"] = atr
            features[f"feat_atr_{w}_ratio"] = safe_div(atr, close)

            add_feature_meta(metadata, f"feat_atr_{w}", family, f"ATR {w}.", "mean(TR)", w)
            add_feature_meta(metadata, f"feat_atr_{w}_ratio", family, f"ATR ratio {w}.", "ATR/Close", w)

        execution_steps.append({"family": family, "status": "OK", "elapsed_seconds": round(time.time() - start, 6)})
        append_cuda_rolling_report(execution_steps, family, rolling_reports)

    # -------------------------------------------------------------------------
    # RANGE VOL ESTIMATORS
    # -------------------------------------------------------------------------
    family = "range_volatility_estimators"
    if FEATURE_CONFIG[family]["enabled"]:
        start = time.time()

        hl_log = np.log(safe_div(high, low))
        co_log = np.log(safe_div(close, open_))
        hl_log_sq = hl_log ** 2
        co_log_sq = co_log ** 2
        gk_base = 0.5 * hl_log_sq - (2.0 * math.log(2.0) - 1.0) * co_log_sq

        rolling_reports: List[Dict[str, Any]] = []
        for w in filter_windows_for_timeframe(FEATURE_CONFIG[family]["windows"], timeframe):
            hl_stats, hl_report = compute_rolling_stats(hl_log_sq, w, backend_plan, need_mean=True)
            gk_stats, gk_report = compute_rolling_stats(gk_base, w, backend_plan, need_mean=True)
            rolling_reports.extend([hl_report, gk_report])
            features[f"feat_parkinson_vol_{w}"] = np.sqrt(
                safe_div(hl_stats["mean"], 4.0 * math.log(2.0))
            )
            features[f"feat_garman_klass_vol_{w}"] = np.sqrt(
                gk_stats["mean"].clip(lower=0)
            )

            add_feature_meta(metadata, f"feat_parkinson_vol_{w}", family, f"Parkinson vol {w}.", "Parkinson", w)
            add_feature_meta(metadata, f"feat_garman_klass_vol_{w}", family, f"Garman-Klass vol {w}.", "GK", w)

        execution_steps.append({"family": family, "status": "OK", "elapsed_seconds": round(time.time() - start, 6)})
        append_cuda_rolling_report(execution_steps, family, rolling_reports)

    # -------------------------------------------------------------------------
    # DONCHIAN
    # -------------------------------------------------------------------------
    family = "donchian"
    if FEATURE_CONFIG[family]["enabled"]:
        start = time.time()

        for w in filter_windows_for_timeframe(FEATURE_CONFIG[family]["windows"], timeframe):
            don_high = high.rolling(w, min_periods=w).max()
            don_low = low.rolling(w, min_periods=w).min()
            don_mid = (don_high + don_low) / 2.0

            cols = {
                f"feat_donchian_high_dist_{w}": safe_div(close, don_high) - 1.0,
                f"feat_donchian_low_dist_{w}": safe_div(close, don_low) - 1.0,
                f"feat_donchian_mid_dist_{w}": safe_div(close, don_mid) - 1.0,
                f"feat_donchian_width_{w}": safe_div(don_high - don_low, don_mid),
            }

            features.update(cols)

            for col in cols:
                add_feature_meta(metadata, col, family, f"Donchian {w}.", "Donchian", w)

        execution_steps.append({"family": family, "status": "OK", "elapsed_seconds": round(time.time() - start, 6)})

    # -------------------------------------------------------------------------
    # OSCILLATORS
    # -------------------------------------------------------------------------
    family = "oscillators"
    if FEATURE_CONFIG[family]["enabled"]:
        start = time.time()

        for w in FEATURE_CONFIG[family]["stoch_windows"]:
            k = stochastic_k(close, high, low, w)
            d = k.rolling(3, min_periods=3).mean()
            wr = williams_r(close, high, low, w)

            features[f"feat_stoch_k_{w}"] = k
            features[f"feat_stoch_d_{w}"] = d
            features[f"feat_williams_r_{w}"] = wr

            add_feature_meta(metadata, f"feat_stoch_k_{w}", family, f"Stoch K {w}.", "Stoch", w)
            add_feature_meta(metadata, f"feat_stoch_d_{w}", family, f"Stoch D {w}.", "SMA K", w + 3)
            add_feature_meta(metadata, f"feat_williams_r_{w}", family, f"Williams R {w}.", "Williams", w)

        for w in FEATURE_CONFIG[family]["cci_windows"]:
            col = f"feat_cci_{w}"
            features[col] = cci(close, high, low, w)
            add_feature_meta(metadata, col, family, f"CCI {w}.", "CCI", w)

        execution_steps.append({"family": family, "status": "OK", "elapsed_seconds": round(time.time() - start, 6)})

    # -------------------------------------------------------------------------
    # HIGHER MOMENTS
    # -------------------------------------------------------------------------
    family = "higher_moments"
    if FEATURE_CONFIG[family]["enabled"] and not (turbo and TURBO_DISABLE_HIGHER_MOMENTS):
        start = time.time()

        for w in filter_windows_for_timeframe(FEATURE_CONFIG[family]["windows"], timeframe):
            features[f"feat_ret_skew_{w}"] = ret_1_sensitive_cpu.rolling(w, min_periods=w).skew()
            features[f"feat_ret_kurt_{w}"] = ret_1_sensitive_cpu.rolling(w, min_periods=w).kurt()

            add_feature_meta(metadata, f"feat_ret_skew_{w}", family, f"Skew ret {w}.", "skew", w)
            add_feature_meta(metadata, f"feat_ret_kurt_{w}", family, f"Kurt ret {w}.", "kurt", w)

        execution_steps.append({"family": family, "status": "OK", "elapsed_seconds": round(time.time() - start, 6)})
    elif FEATURE_CONFIG[family]["enabled"]:
        execution_steps.append({"family": family, "status": "SKIPPED_TURBO", "elapsed_seconds": 0.0})

    # -------------------------------------------------------------------------
    # REGIME
    # -------------------------------------------------------------------------
    family = "regime"
    if FEATURE_CONFIG[family]["enabled"]:
        start = time.time()
        rolling_reports: List[Dict[str, Any]] = []

        for w in filter_windows_for_timeframe(FEATURE_CONFIG[family]["windows"], timeframe):
            trend_abs = log_close.diff(w).abs()
            noise_stats, noise_report = compute_rolling_stats(abs_ret_1, w, backend_plan, need_sum=True)
            rolling_reports.append(noise_report)
            noise = noise_stats["sum"]
            efficiency = safe_div(trend_abs, noise)

            vol = vol_cache.get(w)
            if vol is None:
                vol_stats, vol_report = compute_rolling_stats(ret_1, w, backend_plan, need_std=True)
                rolling_reports.append(vol_report)
                vol = vol_stats["std"]

            vol_z_stats, vol_z_report = compute_rolling_stats(vol, w, backend_plan, need_mean=True, need_std=True)
            rolling_reports.append(vol_z_report)
            features[f"feat_efficiency_ratio_{w}"] = efficiency
            features[f"feat_vol_regime_z_{w}"] = safe_div(vol - vol_z_stats["mean"], vol_z_stats["std"])

            add_feature_meta(metadata, f"feat_efficiency_ratio_{w}", family, f"Efficiency ratio {w}.", "abs(ret_w)/sum(absret)", w)
            add_feature_meta(metadata, f"feat_vol_regime_z_{w}", family, f"Vol regime z {w}.", "zscore(vol)", 2 * w)

        execution_steps.append({"family": family, "status": "OK", "elapsed_seconds": round(time.time() - start, 6)})
        append_cuda_rolling_report(execution_steps, family, rolling_reports)

    # -------------------------------------------------------------------------
    # RISK PROXY
    # -------------------------------------------------------------------------
    family = "risk_proxy"
    if FEATURE_CONFIG[family]["enabled"]:
        start = time.time()

        downside = ret_1_sensitive_cpu.where(ret_1_sensitive_cpu < 0, 0.0)

        for w in filter_windows_for_timeframe(FEATURE_CONFIG[family]["windows"], timeframe):
            # Mantido em Pandas/CPU: Sharpe/Sortino amplificam pequenas diferenças de std.
            roll = ret_1_sensitive_cpu.rolling(w, min_periods=w)
            mean_ret = roll.mean()
            std_ret = roll.std()
            downside_std = downside.rolling(w, min_periods=w).std()

            features[f"feat_rolling_sharpe_proxy_{w}"] = safe_div(mean_ret, std_ret) * sqrt_ann
            features[f"feat_rolling_sortino_proxy_{w}"] = safe_div(mean_ret, downside_std) * sqrt_ann

            add_feature_meta(metadata, f"feat_rolling_sharpe_proxy_{w}", family, f"Sharpe proxy {w}.", "mean/std*sqrtAF", w)
            add_feature_meta(metadata, f"feat_rolling_sortino_proxy_{w}", family, f"Sortino proxy {w}.", "mean/downside*sqrtAF", w)

            if not (turbo and TURBO_DISABLE_ROLLING_VAR_QUANTILE):
                for q in FEATURE_CONFIG[family]["var_quantiles"]:
                    q_token = str(q).replace(".", "")
                    col = f"feat_var_q{q_token}_{w}"
                    features[col] = roll.quantile(q)
                    add_feature_meta(metadata, col, family, f"VaR histórico q={q} {w}.", "rolling quantile", w)

        execution_steps.append({"family": family, "status": "OK", "elapsed_seconds": round(time.time() - start, 6)})

    # -------------------------------------------------------------------------
    # VOLUME PRESSURE
    # -------------------------------------------------------------------------
    family = "volume_pressure"
    if FEATURE_CONFIG[family]["enabled"]:
        start = time.time()
        rolling_reports: List[Dict[str, Any]] = []

        signed_volume = np.sign(close - open_) * volume
        features["feat_signed_volume"] = signed_volume
        add_feature_meta(metadata, "feat_signed_volume", family, "Volume assinado.", "sign(C-O)*V", 0)

        for w in FEATURE_CONFIG[family]["windows"]:
            signed_stats, signed_report = compute_rolling_stats(signed_volume, w, backend_plan, need_sum=True)
            volume_stats, volume_report = compute_rolling_stats(volume, w, backend_plan, need_sum=True)
            rolling_reports.extend([signed_report, volume_report])
            sum_signed = signed_stats["sum"]
            sum_volume = volume_stats["sum"]

            features[f"feat_signed_volume_sum_{w}"] = sum_signed
            features[f"feat_volume_pressure_{w}"] = safe_div(sum_signed, sum_volume)

            add_feature_meta(metadata, f"feat_signed_volume_sum_{w}", family, f"Soma signed volume {w}.", "sum", w)
            add_feature_meta(metadata, f"feat_volume_pressure_{w}", family, f"Volume pressure {w}.", "signed/volume", w)

        execution_steps.append({"family": family, "status": "OK", "elapsed_seconds": round(time.time() - start, 6)})
        append_cuda_rolling_report(execution_steps, family, rolling_reports)

    # -------------------------------------------------------------------------
    # LIQUIDITY
    # -------------------------------------------------------------------------
    family = "liquidity_proxy"
    if FEATURE_CONFIG[family]["enabled"]:
        start = time.time()

        dollar_volume = close * volume
        features["feat_dollar_volume"] = dollar_volume
        add_feature_meta(metadata, "feat_dollar_volume", family, "Dollar volume.", "Close*Volume", 0)

        for w in FEATURE_CONFIG[family]["windows"]:
            # Mantido em Pandas/CPU: escala de dollar volume exige tolerância relativa formal antes de CUDA.
            dv_mean = dollar_volume.rolling(w, min_periods=w).mean()
            dv_std = dollar_volume.rolling(w, min_periods=w).std()

            features[f"feat_dollar_volume_mean_{w}"] = dv_mean
            features[f"feat_dollar_volume_z_{w}"] = safe_div(dollar_volume - dv_mean, dv_std)
            features[f"feat_amihud_illiq_{w}"] = safe_div(abs_ret_1_sensitive_cpu, dollar_volume.replace(0, np.nan)).rolling(w, min_periods=w).mean()

            add_feature_meta(metadata, f"feat_dollar_volume_mean_{w}", family, f"Dollar volume mean {w}.", "mean(DV)", w)
            add_feature_meta(metadata, f"feat_dollar_volume_z_{w}", family, f"Dollar volume z {w}.", "zscore(DV)", w)
            add_feature_meta(metadata, f"feat_amihud_illiq_{w}", family, f"Amihud illiquidity proxy {w}.", "mean(absret/dollar_volume)", w)

        execution_steps.append({"family": family, "status": "OK", "elapsed_seconds": round(time.time() - start, 6)})

    # -------------------------------------------------------------------------
    # SHOCK FEATURES
    # -------------------------------------------------------------------------
    family = "shock_features"
    if FEATURE_CONFIG[family]["enabled"]:
        start = time.time()
        rolling_reports: List[Dict[str, Any]] = []

        candle_range_ret = safe_div(high, low) - 1.0
        dollar_volume = close * volume
        log_dollar_volume = np.log1p(dollar_volume)

        for w in FEATURE_CONFIG[family]["windows"]:
            ret_stats, ret_report = compute_rolling_stats(ret_1, w, backend_plan, need_mean=True, need_std=True)
            range_stats, range_report = compute_rolling_stats(candle_range_ret, w, backend_plan, need_mean=True, need_std=True)
            dv_stats, dv_report = compute_rolling_stats(log_dollar_volume, w, backend_plan, need_mean=True, need_std=True)
            rolling_reports.extend([ret_report, range_report, dv_report])
            ret_z = safe_div(ret_1 - ret_stats["mean"], ret_stats["std"])
            range_z = safe_div(candle_range_ret - range_stats["mean"], range_stats["std"])
            dv_z = safe_div(log_dollar_volume - dv_stats["mean"], dv_stats["std"])

            features[f"feat_ret_shock_z_{w}"] = ret_z
            features[f"feat_range_shock_z_{w}"] = range_z
            features[f"feat_dollar_volume_shock_z_{w}"] = dv_z
            features[f"feat_large_move_flag_{w}"] = (ret_z.abs() >= 2.5).astype("float32")

            add_feature_meta(metadata, f"feat_ret_shock_z_{w}", family, f"Ret shock z {w}.", "zscore(ret)", w)
            add_feature_meta(metadata, f"feat_range_shock_z_{w}", family, f"Range shock z {w}.", "zscore(range)", w)
            add_feature_meta(metadata, f"feat_dollar_volume_shock_z_{w}", family, f"DV shock z {w}.", "zscore(log DV)", w)
            add_feature_meta(metadata, f"feat_large_move_flag_{w}", family, f"Flag movimento extremo {w}.", "abs(ret_z)>=2.5", w)

        execution_steps.append({"family": family, "status": "OK", "elapsed_seconds": round(time.time() - start, 6)})
        append_cuda_rolling_report(execution_steps, family, rolling_reports)

    # -------------------------------------------------------------------------
    # MARKET STRUCTURE
    # -------------------------------------------------------------------------
    family = "market_structure"
    if FEATURE_CONFIG[family]["enabled"]:
        start = time.time()

        for w in FEATURE_CONFIG[family]["windows"]:
            prev_high = high.rolling(w, min_periods=w).max().shift(1)
            prev_low = low.rolling(w, min_periods=w).min().shift(1)

            features[f"feat_breakout_up_{w}"] = (close > prev_high).astype("float32")
            features[f"feat_breakout_down_{w}"] = (close < prev_low).astype("float32")
            features[f"feat_close_to_prev_high_{w}"] = safe_div(close, prev_high) - 1.0
            features[f"feat_close_to_prev_low_{w}"] = safe_div(close, prev_low) - 1.0

            add_feature_meta(metadata, f"feat_breakout_up_{w}", family, f"Breakout up {w}.", "Close > max(high).shift(1)", w + 1)
            add_feature_meta(metadata, f"feat_breakout_down_{w}", family, f"Breakout down {w}.", "Close < min(low).shift(1)", w + 1)
            add_feature_meta(metadata, f"feat_close_to_prev_high_{w}", family, f"Distância high anterior {w}.", "Close/prev_high-1", w + 1)
            add_feature_meta(metadata, f"feat_close_to_prev_low_{w}", family, f"Distância low anterior {w}.", "Close/prev_low-1", w + 1)

        execution_steps.append({"family": family, "status": "OK", "elapsed_seconds": round(time.time() - start, 6)})

    # -------------------------------------------------------------------------
    # TIME CONTEXT
    # -------------------------------------------------------------------------
    family = "time_context"
    if FEATURE_CONFIG[family]["enabled"]:
        start = time.time()

        dt = pd.to_datetime(df["DateTime"])
        hour = dt.dt.hour.astype(float)
        dow = dt.dt.dayofweek.astype(float)

        features["feat_hour_sin"] = np.sin(2.0 * np.pi * hour / 24.0)
        features["feat_hour_cos"] = np.cos(2.0 * np.pi * hour / 24.0)
        features["feat_dow_sin"] = np.sin(2.0 * np.pi * dow / 7.0)
        features["feat_dow_cos"] = np.cos(2.0 * np.pi * dow / 7.0)
        features["feat_is_weekend"] = dt.dt.dayofweek.isin([5, 6]).astype("float32")

        for col in ["feat_hour_sin", "feat_hour_cos", "feat_dow_sin", "feat_dow_cos", "feat_is_weekend"]:
            add_feature_meta(metadata, col, family, col, "datetime context, no future", 0)

        execution_steps.append({"family": family, "status": "OK", "elapsed_seconds": round(time.time() - start, 6)})

    # -------------------------------------------------------------------------
    # ML SAFETY
    # -------------------------------------------------------------------------
    family = "ml_safety"
    if FEATURE_CONFIG[family]["enabled"]:
        start = time.time()

        features["feat_valid_ohlcv_flag"] = (
            close.notna()
            & open_.notna()
            & high.notna()
            & low.notna()
            & volume.notna()
            & (close > 0)
            & (high >= low)
        ).astype("float32")

        features["feat_zero_volume_flag"] = (volume <= 0).astype("float32")

        add_feature_meta(metadata, "feat_valid_ohlcv_flag", family, "Flag de OHLCV válido.", "valid OHLCV", 0)
        add_feature_meta(metadata, "feat_zero_volume_flag", family, "Flag volume zero.", "Volume<=0", 0)

        execution_steps.append({"family": family, "status": "OK", "elapsed_seconds": round(time.time() - start, 6)})

    # -------------------------------------------------------------------------
    # FINAL
    # -------------------------------------------------------------------------
    feature_df = build_feature_df_from_blocks(features, df.index)
    if not ENABLE_FEATURE_BLOCK_CONSTRUCTION:
        feature_df = feature_df.replace([np.inf, -np.inf], np.nan)

    if DROP_FEATURES_WITH_NULL_RATIO_ABOVE is not None and len(feature_df) > 0:
        null_ratios = feature_df.isna().mean()
        to_drop = null_ratios[null_ratios > DROP_FEATURES_WITH_NULL_RATIO_ABOVE].index.tolist()

        if to_drop:
            drop_set = set(to_drop)
            feature_df = feature_df.drop(columns=to_drop)
            metadata = [m for m in metadata if m.get("feature") not in drop_set]

    if DROP_CONSTANT_FEATURES and not feature_df.empty:
        to_drop_const = []

        for col in feature_df.columns:
            try:
                if feature_df[col].nunique(dropna=True) <= 1:
                    to_drop_const.append(col)
            except Exception:
                pass

        if to_drop_const:
            drop_set = set(to_drop_const)
            feature_df = feature_df.drop(columns=to_drop_const)
            metadata = [m for m in metadata if m.get("feature") not in drop_set]

    df = pd.concat([df, feature_df], axis=1)

    return df, metadata, execution_steps


# =============================================================================
# 7. PÓS-AUDITORIA DE SINCRONIZAÇÃO
# =============================================================================

def post_generation_audit(
    df_input_prepared: pd.DataFrame,
    df_output: pd.DataFrame,
    timeframe: Optional[str],
) -> Dict[str, Any]:

    audit: Dict[str, Any] = {
        "audit_status": "PENDING",
        "checks": {},
        "warnings": [],
        "errors": [],
        "time_policy": TIMEZONE_POLICY,
    }

    try:
        tf_seconds = infer_timeframe_seconds(timeframe)
        tf_ms = None if tf_seconds is None else tf_seconds * 1000

        in_dt = pd.to_datetime(df_input_prepared[DATETIME_COL])
        out_dt = pd.to_datetime(df_output[DATETIME_COL])

        audit["checks"]["input_rows"] = int(len(df_input_prepared))
        audit["checks"]["output_rows"] = int(len(df_output))
        audit["checks"]["row_count_match"] = bool(len(df_input_prepared) == len(df_output))

        audit["checks"]["first_datetime_input"] = None if in_dt.empty else str(in_dt.min())
        audit["checks"]["last_datetime_input"] = None if in_dt.empty else str(in_dt.max())
        audit["checks"]["first_datetime_output"] = None if out_dt.empty else str(out_dt.min())
        audit["checks"]["last_datetime_output"] = None if out_dt.empty else str(out_dt.max())

        audit["checks"]["first_datetime_match"] = bool(
            len(in_dt) > 0 and len(out_dt) > 0 and in_dt.iloc[0] == out_dt.iloc[0]
        )
        audit["checks"]["last_datetime_match"] = bool(
            len(in_dt) > 0 and len(out_dt) > 0 and in_dt.iloc[-1] == out_dt.iloc[-1]
        )

        audit["checks"]["datetime_exact_sequence_match"] = bool(
            len(in_dt) == len(out_dt) and np.array_equal(in_dt.to_numpy(), out_dt.to_numpy())
        )

        audit["checks"]["output_datetime_monotonic"] = bool(out_dt.is_monotonic_increasing)
        audit["checks"]["output_datetime_duplicates"] = int(out_dt.duplicated().sum())

        has_ts_input = TIMESTAMP_UTC_MS_COL in df_input_prepared.columns
        has_ts_output = TIMESTAMP_UTC_MS_COL in df_output.columns

        audit["checks"]["timestamp_utc_ms_input_present"] = bool(has_ts_input)
        audit["checks"]["timestamp_utc_ms_output_present"] = bool(has_ts_output)

        if has_ts_input and has_ts_output:
            in_ts = normalize_timestamp_to_utc_ms(df_input_prepared[TIMESTAMP_UTC_MS_COL])
            out_ts = normalize_timestamp_to_utc_ms(df_output[TIMESTAMP_UTC_MS_COL])


            audit["checks"]["first_timestamp_utc_ms_input"] = None if in_ts.empty else int(in_ts.iloc[0])
            audit["checks"]["last_timestamp_utc_ms_input"] = None if in_ts.empty else int(in_ts.iloc[-1])
            audit["checks"]["first_timestamp_utc_ms_output"] = None if out_ts.empty else int(out_ts.iloc[0])
            audit["checks"]["last_timestamp_utc_ms_output"] = None if out_ts.empty else int(out_ts.iloc[-1])

            audit["checks"]["timestamp_utc_ms_exact_sequence_match"] = bool(
                len(in_ts) == len(out_ts)
                and np.array_equal(in_ts.to_numpy(), out_ts.to_numpy())
            )

            audit["checks"]["output_timestamp_utc_ms_monotonic"] = bool(out_ts.is_monotonic_increasing)
            audit["checks"]["output_timestamp_utc_ms_duplicates"] = int(out_ts.duplicated().sum())

            if tf_ms and len(out_ts) > 1:
                ts_deltas = out_ts.diff().dropna()
                median_delta_ms = float(ts_deltas.median()) if not ts_deltas.empty else None
                max_delta_ms = float(ts_deltas.max()) if not ts_deltas.empty else None

                audit["checks"]["expected_timeframe_ms"] = int(tf_ms)
                audit["checks"]["median_delta_ms"] = median_delta_ms
                audit["checks"]["max_delta_ms"] = max_delta_ms
                audit["checks"]["median_delta_ms_matches_timeframe"] = bool(
                    median_delta_ms is not None and abs(median_delta_ms - tf_ms) <= max(1000.0, tf_ms * 0.01)
                )

                gap_threshold_ms = tf_ms * POST_AUDIT_MAX_GAP_MULTIPLIER
                audit["checks"]["large_gap_count_timestamp_utc_ms"] = int((ts_deltas > gap_threshold_ms).sum())
            else:
                audit["checks"]["expected_timeframe_ms"] = tf_ms
                audit["checks"]["median_delta_ms"] = None
                audit["checks"]["max_delta_ms"] = None
                audit["checks"]["large_gap_count_timestamp_utc_ms"] = None

        else:
            audit["warnings"].append("timestamp_utc_ms ausente no input ou output.")

        if tf_seconds and len(out_dt) > 1:
            deltas = out_dt.diff().dropna().dt.total_seconds()
            median_delta = float(deltas.median()) if not deltas.empty else None
            max_delta = float(deltas.max()) if not deltas.empty else None

            audit["checks"]["expected_timeframe_seconds"] = int(tf_seconds)
            audit["checks"]["median_delta_seconds"] = median_delta
            audit["checks"]["max_delta_seconds"] = max_delta
            audit["checks"]["median_delta_matches_timeframe"] = bool(
                median_delta is not None and abs(median_delta - tf_seconds) <= max(1.0, tf_seconds * 0.01)
            )

            gap_threshold = tf_seconds * POST_AUDIT_MAX_GAP_MULTIPLIER
            audit["checks"]["large_gap_count"] = int((deltas > gap_threshold).sum())
        else:
            audit["checks"]["expected_timeframe_seconds"] = tf_seconds
            audit["checks"]["median_delta_seconds"] = None
            audit["checks"]["max_delta_seconds"] = None
            audit["checks"]["large_gap_count"] = None

        for col in ["Open", "High", "Low", "Close", "Volume"]:
            if col in df_output.columns and col in df_input_prepared.columns:
                try:
                    audit["checks"][f"{col}_preserved"] = bool(
                        np.allclose(
                            df_input_prepared[col].astype(float).to_numpy(),
                            df_output[col].astype(float).to_numpy(),
                            equal_nan=True,
                        )
                    )
                except Exception:
                    audit["checks"][f"{col}_preserved"] = False

        feature_cols = [c for c in df_output.columns if str(c).startswith("feat_")]
        audit["checks"]["feature_columns_count"] = int(len(feature_cols))

        if feature_cols:
            quality_df, quality_sampled = sample_df_for_quality_summary(df_output)
            null_ratio = float(
                quality_df[feature_cols].isna().sum().sum()
                / max(1, len(quality_df) * len(feature_cols))
            )
            audit["checks"]["overall_feature_null_ratio"] = round(null_ratio, 8)
            audit["checks"]["overall_feature_null_ratio_exact"] = not quality_sampled
            audit["checks"]["overall_feature_null_ratio_sample_rows"] = int(len(quality_df))
        else:
            audit["checks"]["overall_feature_null_ratio"] = None
            audit["checks"]["overall_feature_null_ratio_exact"] = True
            audit["checks"]["overall_feature_null_ratio_sample_rows"] = 0

        critical_checks = [
            "row_count_match",
            "first_datetime_match",
            "last_datetime_match",
            "datetime_exact_sequence_match",
            "output_datetime_monotonic",
            "timestamp_utc_ms_output_present",
        ]

        if has_ts_input and has_ts_output:
            critical_checks.extend([
                "timestamp_utc_ms_exact_sequence_match",
                "output_timestamp_utc_ms_monotonic",
            ])

        for check in critical_checks:
            if audit["checks"].get(check) is not True:
                audit["errors"].append(f"Critical sync check failed: {check}")

        if audit["checks"].get("output_datetime_duplicates", 0) > 0:
            audit["errors"].append("Output contains duplicated DateTime.")

        if audit["checks"].get("output_timestamp_utc_ms_duplicates", 0) > 0:
            audit["errors"].append("Output contains duplicated timestamp_utc_ms.")

        if audit["checks"].get("large_gap_count") not in [None, 0]:
            audit["warnings"].append("Large DateTime gaps detected vs expected timeframe.")

        if audit["checks"].get("large_gap_count_timestamp_utc_ms") not in [None, 0]:
            audit["warnings"].append("Large timestamp_utc_ms gaps detected vs expected timeframe.")

        if audit["errors"]:
            audit["audit_status"] = "FAIL"
        elif audit["warnings"]:
            audit["audit_status"] = "WARNING"
        else:
            audit["audit_status"] = "PASS"

        return audit

    except Exception as exc:
        audit["audit_status"] = "ERROR"
        audit["errors"].append(str(exc))
        audit["traceback"] = traceback.format_exc()
        return audit

# =============================================================================
# 8. PÓS-PROCESSAMENTO
# =============================================================================

def replace_inf_with_nan(df: pd.DataFrame) -> pd.DataFrame:
    numeric_cols = df.select_dtypes(include=[np.number]).columns
    if len(numeric_cols) > 0:
        df[numeric_cols] = df[numeric_cols].replace([np.inf, -np.inf], np.nan)
    return df


def downcast_float_columns(df: pd.DataFrame) -> pd.DataFrame:
    if not DOWNCAST_FLOATS_TO_FLOAT32:
        return df

    float_cols = df.select_dtypes(include=["float64"]).columns

    if len(float_cols) > 0:
        df[float_cols] = df[float_cols].astype("float32", copy=False)

    return df











def add_metadata_columns(
    df: pd.DataFrame,
    asset: str,
    symbol: str,
    source: str,
    timeframe: str,
    dataset_kind: str,
    series_id: str,
    timeframe_from_path: Optional[str],
    timeframe_from_map: Optional[str],
    quality_status: str = "UNKNOWN",
    quality_record: Optional[Dict[str, Any]] = None,
    input_contract_report: Optional[Dict[str, Any]] = None,
    time_grid_report: Optional[Dict[str, Any]] = None,
    input_path: Optional[Path] = None,
) -> pd.DataFrame:

    if not INCLUDE_METADATA_COLUMNS:
        return df

    time_meta = get_archangel_time_metadata()

    input_path_str = path_to_str(input_path) if input_path else None
    input_file_mtime = None
    input_file_size = None

    if input_path and input_path.exists():
        try:
            input_file_mtime = datetime.fromtimestamp(input_path.stat().st_mtime).isoformat(timespec="seconds")
            input_file_size = int(input_path.stat().st_size)
        except Exception:
            pass

    normalized_quality = normalize_quality_status(quality_status)
    ml_quality_status = get_ml_quality_status_from_record(quality_record)
    ml_usable_for_broad_training = is_quality_usable_for_ml(quality_record, normalized_quality)
    ml_quality_reason = (
        quality_record.get("ml_quality_reason")
        if isinstance(quality_record, dict)
        else None
    )

    contract_status = (
        input_contract_report.get("status")
        if isinstance(input_contract_report, dict)
        else None
    )

    time_grid_status = (
        time_grid_report.get("status")
        if isinstance(time_grid_report, dict)
        else None
    )

    meta_values = {
        "meta_asset": asset,
        "meta_symbol": symbol,
        "meta_source": source,
        "meta_timeframe": timeframe,
        "meta_dataset_kind": dataset_kind,
        "meta_series_id": series_id,
        "meta_timeframe_from_path": timeframe_from_path,
        "meta_timeframe_from_map": timeframe_from_map,

        "meta_schema_version": SCHEMA_VERSION,
        "meta_input_contract_version": INPUT_CONTRACT_VERSION,
        "meta_quality_gate_version": QUALITY_GATE_VERSION,
        "meta_ml_readiness_version": ML_READINESS_VERSION,

        "meta_generated_at": now_iso(),
        "meta_uses_future_data": False,
        "meta_labels_created_here": False,

        "meta_input_path": input_path_str,
        "meta_input_file_mtime": input_file_mtime,
        "meta_input_file_size": input_file_size,

        "meta_timezone": time_meta["meta_timezone"],
        "meta_datetime_is_naive": time_meta["meta_datetime_is_naive"],
        "meta_timestamp_utc_ms_present": time_meta["meta_timestamp_utc_ms_present"],
        "meta_datetime_column": time_meta["meta_datetime_column"],
        "meta_timestamp_utc_ms_column": time_meta["meta_timestamp_utc_ms_column"],
        "meta_bar_timestamp_policy": time_meta["meta_bar_timestamp_policy"],

        "quality_status": normalized_quality,
        "quality_original_status": quality_status,
        "quality_ml_status": ml_quality_status,
        "quality_ml_reason": ml_quality_reason,
        "quality_data_usable_for_ml": bool(ml_usable_for_broad_training),
        "quality_input_contract_status": contract_status,
        "quality_time_grid_status": time_grid_status,
    }

    if METADATA_OUTPUT_MODE in {"MINIMAL", "LEAN", "FAST"}:
        keep_meta = {
            "meta_asset",
            "meta_symbol",
            "meta_source",
            "meta_timeframe",
            "meta_dataset_kind",
            "meta_series_id",
            "meta_timezone",
            "meta_bar_timestamp_policy",
            "quality_status",
            "quality_original_status",
            "quality_ml_status",
            "quality_ml_reason",
            "quality_data_usable_for_ml",
            "quality_input_contract_status",
            "quality_time_grid_status",
        }
        meta_values = {key: value for key, value in meta_values.items() if key in keep_meta}

    meta = pd.DataFrame(meta_values, index=df.index)

    if ENABLE_FEATURE_READY_TIMESTAMP and TIMESTAMP_UTC_MS_COL in df.columns:
        meta["feature_ready_timestamp_utc_ms"] = pd.to_numeric(
            df[TIMESTAMP_UTC_MS_COL],
            errors="coerce",
        ).astype("int64")

    return pd.concat([df, meta], axis=1)





def infer_max_lookback_from_feature_metadata(feature_metadata: List[Dict[str, Any]]) -> int:
    max_lb = 0

    for item in feature_metadata:
        try:
            lb = item.get("lookback")
            if lb is not None:
                max_lb = max(max_lb, int(lb))
        except Exception:
            pass

    return int(max_lb)


def apply_feature_validity_flags(
    df: pd.DataFrame,
    max_lookback_bars: int,
) -> pd.DataFrame:
    """
    Fase 6:
        Marca warm-up e linhas válidas para ML.

    Não remove linhas.
    O filtro final será feito no 05_MONTA_DATASETS_ML.py.
    """
    if not ENABLE_FEATURE_VALIDITY_FLAGS:
        return df

    out = df

    feature_cols = [c for c in out.columns if str(c).startswith("feat_")]

    out["quality_warmup_bars_required"] = int(max_lookback_bars)

    if len(out) > 0:
        out["quality_is_warmup"] = np.arange(len(out)) < int(max_lookback_bars)
    else:
        out["quality_is_warmup"] = pd.Series(dtype=bool)

    if feature_cols:
        feature_non_null_ratio = out[feature_cols].notna().mean(axis=1)
        feature_notna = feature_non_null_ratio >= MIN_FEATURE_NON_NULL_RATIO_FOR_VALID_ROW
        out["quality_feature_non_null_ratio"] = feature_non_null_ratio.astype("float32")



        if FEATURE_VALIDITY_FINITE_CHECK:
            feature_finite = np.isfinite(out[feature_cols].select_dtypes(include=[np.number])).all(axis=1)

            if len(feature_finite) != len(out):
                feature_finite = pd.Series(True, index=out.index)
            out["quality_is_feature_valid"] = (
                (~out["quality_is_warmup"].astype(bool))
                & feature_notna
                & feature_finite
            )
        else:
            out["quality_is_feature_valid"] = (
                (~out["quality_is_warmup"].astype(bool))
                & feature_notna
            )
    else:
        out["quality_is_feature_valid"] = False

    out["quality_feature_valid_int"] = out["quality_is_feature_valid"].astype("int8")
    out["quality_warmup_int"] = out["quality_is_warmup"].astype("int8")

    if out["quality_is_feature_valid"].any():
        first_valid_ts = int(out.loc[out["quality_is_feature_valid"], TIMESTAMP_UTC_MS_COL].min())
    else:
        first_valid_ts = None

    out["quality_first_valid_timestamp_utc_ms"] = first_valid_ts

    return out


def apply_ml_ready_schema_flags(df: pd.DataFrame) -> pd.DataFrame:
    """
    Fase 8:
        Prepara a ponte para 05_MONTA_DATASETS_ML.py.
    """
    if not ENABLE_ML_READY_SCHEMA_FLAGS:
        return df

    out = df

    feature_cols = [c for c in out.columns if str(c).startswith("feat_")]
    label_cols = [c for c in out.columns if str(c).startswith("label_")]

    out["quality_feature_columns_count"] = int(len(feature_cols))
    out["quality_label_columns_count"] = int(len(label_cols))
    out["quality_has_label_leakage_risk"] = int(len(label_cols) > 0)

    if "quality_data_usable_for_ml" in out.columns and "quality_is_feature_valid" in out.columns:
        out["quality_ml_row_eligible"] = (
            out["quality_data_usable_for_ml"].astype(bool)
            & out["quality_is_feature_valid"].astype(bool)
            & (out["quality_has_label_leakage_risk"] == 0)
        )
    else:
        out["quality_ml_row_eligible"] = False

    out["quality_ml_row_eligible_int"] = out["quality_ml_row_eligible"].astype("int8")

    return out








def select_output_columns(df: pd.DataFrame) -> pd.DataFrame:
    base_cols = [
        DATETIME_COL,
        TIMESTAMP_UTC_MS_COL,
    ]

    if ENABLE_FEATURE_READY_TIMESTAMP:
        base_cols.append("feature_ready_timestamp_utc_ms")

    meta_cols = [
        "meta_asset",
        "meta_symbol",
        "meta_source",
        "meta_timeframe",
        "meta_dataset_kind",
        "meta_series_id",
        "meta_timeframe_from_path",
        "meta_timeframe_from_map",

        "meta_schema_version",
        "meta_input_contract_version",
        "meta_quality_gate_version",
        "meta_ml_readiness_version",

        "meta_generated_at",
        "meta_uses_future_data",
        "meta_labels_created_here",

        "meta_input_path",
        "meta_input_file_mtime",
        "meta_input_file_size",

        "meta_timezone",
        "meta_datetime_is_naive",
        "meta_timestamp_utc_ms_present",
        "meta_datetime_column",
        "meta_timestamp_utc_ms_column",
        "meta_bar_timestamp_policy",
    ] if INCLUDE_METADATA_COLUMNS else []

    quality_cols = [c for c in df.columns if str(c).startswith("quality_")]

    ohlcv_cols = ["Open", "High", "Low", "Close", "Volume"] if INCLUDE_OHLCV_IN_OUTPUT else []

    optional_market_cols = [
        "quote_volume",
        "number_of_trades",
        "taker_buy_base_volume",
        "taker_buy_quote_volume",
    ] if INCLUDE_OPTIONAL_MARKET_COLUMNS_IN_OUTPUT else []

    optional_market_cols = [c for c in optional_market_cols if c in df.columns]

    regime_cols = [c for c in df.columns if str(c).startswith("regime_")]
    feature_cols = [c for c in df.columns if str(c).startswith("feat_")]

    final_cols = (
        base_cols
        + meta_cols
        + quality_cols
        + ohlcv_cols
        + optional_market_cols
        + regime_cols
        + feature_cols
    )

    final_cols = list(dict.fromkeys([c for c in final_cols if c in df.columns]))

    return df.loc[:, final_cols]









def compute_quality_summary(df: pd.DataFrame) -> Dict[str, Any]:
    feature_cols = [c for c in df.columns if str(c).startswith("feat_")]
    summary_df, summary_sampled = sample_df_for_quality_summary(df)

    if feature_cols:
        null_counts = summary_df[feature_cols].isna().sum()
        total_values_basis = int(len(summary_df) * len(feature_cols))
        total_nulls_basis = int(null_counts.sum())
        feature_null_ratio = None if total_values_basis == 0 else total_nulls_basis / total_values_basis
        total_values = int(len(df) * len(feature_cols))
        total_nulls = (
            int(round(feature_null_ratio * total_values))
            if summary_sampled and feature_null_ratio is not None
            else total_nulls_basis
        )
    else:
        total_values = int(len(df))
        total_nulls = 0
        total_values_basis = int(len(summary_df))
        total_nulls_basis = 0
        feature_null_ratio = 0.0 if total_values else None
        null_counts = pd.Series(dtype=float)

    first_ts = None
    last_ts = None

    if TIMESTAMP_UTC_MS_COL in df.columns and not df.empty:
        ts = pd.to_numeric(df[TIMESTAMP_UTC_MS_COL], errors="coerce").dropna()
        if not ts.empty:
            first_ts = int(ts.min())
            last_ts = int(ts.max())

    valid_feature_rows = None
    warmup_rows = None
    ml_eligible_rows = None
    first_valid_ts = None

    if "quality_is_feature_valid" in df.columns:
        valid_feature_rows = int(df["quality_is_feature_valid"].astype(bool).sum())

    if "quality_is_warmup" in df.columns:
        warmup_rows = int(df["quality_is_warmup"].astype(bool).sum())

    if "quality_ml_row_eligible" in df.columns:
        ml_eligible_rows = int(df["quality_ml_row_eligible"].astype(bool).sum())

    if "quality_first_valid_timestamp_utc_ms" in df.columns:
        vals = pd.to_numeric(df["quality_first_valid_timestamp_utc_ms"], errors="coerce").dropna()
        first_valid_ts = None if vals.empty else int(vals.iloc[0])

    quality_status = None
    if "quality_status" in df.columns and not df.empty:
        quality_status = str(df["quality_status"].iloc[0])

    input_contract_status = None
    if "quality_input_contract_status" in df.columns and not df.empty:
        input_contract_status = str(df["quality_input_contract_status"].iloc[0])

    time_grid_status = None
    if "quality_time_grid_status" in df.columns and not df.empty:
        time_grid_status = str(df["quality_time_grid_status"].iloc[0])

    return {
        "rows": int(len(df)),
        "columns": int(len(df.columns)),
        "feature_columns_count": int(len(feature_cols)),

        "first_datetime": None if df.empty else str(df[DATETIME_COL].min()),
        "last_datetime": None if df.empty else str(df[DATETIME_COL].max()),
        "first_timestamp_utc_ms": first_ts,
        "last_timestamp_utc_ms": last_ts,
        "first_valid_timestamp_utc_ms": first_valid_ts,

        "timestamp_utc_ms_present": TIMESTAMP_UTC_MS_COL in df.columns,
        "feature_ready_timestamp_present": "feature_ready_timestamp_utc_ms" in df.columns,
        "bar_timestamp_policy": BAR_TIMESTAMP_POLICY,

        "quality_status": quality_status,
        "input_contract_status": input_contract_status,
        "time_grid_status": time_grid_status,

        "valid_feature_rows": valid_feature_rows,
        "warmup_rows": warmup_rows,
        "ml_eligible_rows": ml_eligible_rows,

        "total_feature_nulls": total_nulls,
        "total_feature_values": total_values,
        "feature_null_ratio": None if feature_null_ratio is None else round(float(feature_null_ratio), 8),
        "feature_null_ratio_exact": not summary_sampled,
        "quality_summary_mode": "SAMPLED_FAST" if summary_sampled else "FULL_EXACT",
        "quality_summary_sample_rows": int(len(summary_df)),
        "quality_summary_total_rows": int(len(df)),
        "total_feature_nulls_basis": total_nulls_basis,
        "total_feature_values_basis": total_values_basis,
        "columns_with_any_nulls": int(sum(1 for v in null_counts.values if v > 0)),

        "sample_hash": make_hash_from_dataframe(df),
    }











# =============================================================================
# 9. MAPA E SELEÇÃO DE SÉRIES
# =============================================================================

def get_series_catalog(mapa: Dict[str, Any]) -> List[Dict[str, Any]]:
    if isinstance(mapa.get("series_catalog"), list):
        return mapa["series_catalog"]

    series_by_id = safe_get_nested(mapa, ["indexes", "series_by_id"])

    if isinstance(series_by_id, dict):
        catalog = []
        for sid, value in series_by_id.items():
            if isinstance(value, dict):
                item = dict(value)
                item.setdefault("series_id", sid)
                catalog.append(item)
        return catalog

    raise KeyError("Não foi possível localizar series_catalog nem indexes.series_by_id.")


def extract_series_basic_info(series: Dict[str, Any]) -> Dict[str, Any]:
    periodicity = series.get("periodicity", {}) if isinstance(series.get("periodicity"), dict) else {}
    quality = series.get("quality", {}) if isinstance(series.get("quality"), dict) else {}
    file_info = series.get("file", {}) if isinstance(series.get("file"), dict) else {}

    absolute_path = (
        file_info.get("absolute_path")
        or series.get("absolute_path")
        or series.get("path")
        or series.get("file_path")
        or series.get("parquet_path")
    )

    relative_path = file_info.get("relative_path") or series.get("relative_path")

    timeframe_from_path = infer_timeframe_from_path(absolute_path) or infer_timeframe_from_path(relative_path)

    timeframe_from_map = (
        series.get("timeframe")
        or periodicity.get("inferred_timeframe")
        or periodicity.get("timeframe")
        or "unknown"
    )

    if PRIORITIZE_TIMEFRAME_FROM_PATH and timeframe_from_path:
        timeframe_final = timeframe_from_path
    else:
        timeframe_final = timeframe_from_map or timeframe_from_path or "unknown"

    return {
        "series_id": series.get("series_id") or series.get("id") or "unknown_series_id",
        "dataset_kind": str(series.get("dataset_kind") or series.get("kind") or "unknown").lower(),
        "asset": str(series.get("asset") or "UNKNOWN").upper(),
        "symbol": str(series.get("symbol") or "UNKNOWN").upper(),
        "source": series.get("source") or "unknown_source",
        "timeframe": timeframe_final,
        "timeframe_from_path": timeframe_from_path,
        "timeframe_from_map": timeframe_from_map,
        "quality_status": quality.get("status") or series.get("quality_status") or "UNKNOWN",
        "absolute_path": absolute_path,
        "relative_path": relative_path,
        "rows_reported": series.get("rows") or quality.get("rows"),
    }


def should_process_series(info: Dict[str, Any]) -> bool:
    if info["dataset_kind"] not in DATASET_KIND_ALLOWED:
        return False

    if ONLY_QUALITY_OK and info["quality_status"] != "OK":
        return False

    if FILTER_ASSETS is not None and info["asset"] not in FILTER_ASSETS:
        return False

    if FILTER_SOURCES is not None and info["source"] not in FILTER_SOURCES:
        return False

    if FILTER_TIMEFRAMES is not None and info["timeframe"] not in FILTER_TIMEFRAMES:
        return False

    if not info["absolute_path"]:
        return False

    return True


def select_series_to_process(mapa: Dict[str, Any]) -> List[Dict[str, Any]]:
    catalog = get_series_catalog(mapa)
    selected = []

    for series in catalog:
        info = extract_series_basic_info(series)

        if should_process_series(info):
            selected.append({"raw": series, "info": info})

    selected = sorted(
        selected,
        key=lambda x: (
            timeframe_sort_key(x["info"].get("timeframe")) if PROCESS_TURBO_TIMEFRAMES_FIRST else (0, 0, ""),
            x["info"]["source"],
            x["info"]["asset"],
            x["info"]["symbol"],
            x["info"]["series_id"],
        ),
    )

    return selected


def split_series_batches(selected: List[Dict[str, Any]]) -> List[Tuple[str, List[Dict[str, Any]]]]:
    if not ENABLE_TIMEFRAME_BATCH_EXECUTION:
        return [("ALL", selected)]

    turbo_batch = [
        item for item in selected
        if is_turbo_timeframe((item.get("info") or {}).get("timeframe"))
    ]
    standard_batch = [
        item for item in selected
        if not is_turbo_timeframe((item.get("info") or {}).get("timeframe"))
    ]

    batches: List[Tuple[str, List[Dict[str, Any]]]] = []
    if PROCESS_TURBO_TIMEFRAMES_FIRST and turbo_batch:
        batches.append(("TURBO_TIMEFRAMES_FIRST", turbo_batch))
    if standard_batch:
        batches.append(("STANDARD_TIMEFRAMES", standard_batch))
    if not PROCESS_TURBO_TIMEFRAMES_FIRST and turbo_batch:
        batches.append(("TURBO_TIMEFRAMES", turbo_batch))
    return batches


def resolve_worker_count(batch: List[Dict[str, Any]], batch_name: str) -> int:
    if not batch:
        return 0

    max_workers = resolve_max_worker_count(batch, batch_name)
    if not ENABLE_ADAPTIVE_WORKERS:
        return max_workers

    if "TURBO" in str(batch_name).upper():
        initial_workers = ADAPTIVE_WORKERS_INITIAL_TURBO
    else:
        initial_workers = ADAPTIVE_WORKERS_INITIAL_STANDARD

    initial_workers = max(1, min(max_workers, int(initial_workers), len(batch)))

    if not ENABLE_MEMORY_AWARE_WORKERS:
        return initial_workers

    available_ram_gb = get_available_ram_gb()
    if available_ram_gb is None:
        return initial_workers

    per_worker = estimate_ram_gb_per_worker_for_batch(batch)
    ram_workers = estimate_ram_safe_worker_count(float(available_ram_gb), per_worker)
    return max(1, min(initial_workers, ram_workers))


def resolve_max_worker_count(batch: List[Dict[str, Any]], batch_name: str) -> int:
    if not batch:
        return 0

    cpu_workers = min(MAX_WORKERS_FEATURES, os.cpu_count() or 1, len(batch))
    if "TURBO" in str(batch_name).upper():
        cpu_workers = min(cpu_workers, MAX_WORKERS_TURBO_BATCH)

    if FEATURE_CUDA_MODE in {"auto", "cuda", "cupy", "gpu"}:
        cpu_workers = min(cpu_workers, max(1, int(FEATURE_CUDA_MAX_WORKERS)))

    return max(1, cpu_workers)


def estimate_ram_gb_per_worker_for_batch(batch: List[Dict[str, Any]]) -> float:
    has_heavy_tf = any(
        is_ram_heavy_timeframe((item.get("info") or {}).get("timeframe"))
        for item in batch
    )
    return (
        ESTIMATED_RAM_GB_PER_WORKER_TURBO
        if has_heavy_tf
        else ESTIMATED_RAM_GB_PER_WORKER_STANDARD
    )


def estimate_ram_safe_worker_count(
    available_ram_gb: Optional[float],
    estimated_ram_gb_per_worker: float,
) -> int:
    if available_ram_gb is None:
        return 1

    per_worker = max(1.0, float(estimated_ram_gb_per_worker))
    ram_budget = max(0.0, float(available_ram_gb) - get_target_free_ram_gb())
    return max(1, int(ram_budget // per_worker))


def classify_result_error(result: Dict[str, Any]) -> str:
    status = str(result.get("status") or "")
    text = " ".join(
        str(result.get(key) or "")
        for key in ("error", "traceback", "reason")
    ).lower()

    if "memoryerror" in text or "out of memory" in text or "memory" in text and "allocate" in text:
        return "ERROR_MEMORY"
    if "process pool" in text or "terminated abruptly" in text or "brokenprocesspool" in text:
        return "ERROR_WORKER_CRASH"
    if "parquet" in text or "to_parquet" in text or "arrow" in text:
        return "ERROR_WRITE_PARQUET"
    if status.startswith("ERROR_INPUT_CONTRACT"):
        return "ERROR_INPUT_CONTRACT"
    return status or "ERROR_UNKNOWN"


def is_retryable_result(result: Dict[str, Any]) -> bool:
    status = str(result.get("status") or "")
    if status == "OK" or status.startswith("SKIPPED"):
        return False
    if status.startswith("ERROR_INPUT_CONTRACT"):
        return False
    return status.startswith("ERROR")


# =============================================================================
# 10. CAMINHOS E PROCESSAMENTO
# =============================================================================

def build_feature_output_path(info: Dict[str, Any]) -> Path:
    source = safe_filename_token(info["source"])
    asset = safe_filename_token(info["asset"])
    symbol = safe_filename_token(info["symbol"])
    timeframe = safe_filename_token(info["timeframe"])

    sid_hash = short_series_hash(
        series_id=str(info.get("series_id", "")),
        input_path=str(info.get("absolute_path", "")),
    )

    out_dir = FEATURES_PARQUET_DIR / source / asset / timeframe
    ensure_dir(out_dir)

    return out_dir / f"{symbol}_{source}_{timeframe}_{sid_hash}_features.parquet"



def read_ohlcv_parquet_fast(input_path: Path) -> pd.DataFrame:
    if READ_PARQUET_ONLY_OHLCV_WHEN_POSSIBLE:
        preferred_cols = [
            DATETIME_COL,
            TIMESTAMP_UTC_MS_COL,
            "timestamp",
            "Open",
            "High",
            "Low",
            "Close",
            "Volume",
            "open",
            "high",
            "low",
            "close",
            "volume",
        ]

        if READ_OPTIONAL_MARKET_COLUMNS:
            preferred_cols.extend([
                "quote_volume",
                "quote_asset_volume",
                "number_of_trades",
                "trades",
                "trade_count",
                "taker_buy_base_volume",
                "taker_buy_volume",
                "taker_buy_quote_volume",
            ])

        try:
            import pyarrow.parquet as pq

            parquet_file = pq.ParquetFile(input_path)
            available_cols = parquet_file.schema.names

            cols_to_read = [c for c in preferred_cols if c in available_cols]

            has_time = (
                DATETIME_COL in cols_to_read
                or TIMESTAMP_UTC_MS_COL in cols_to_read
                or "timestamp" in cols_to_read
            )

            has_ohlcv_pascal = all(c in cols_to_read for c in ["Open", "High", "Low", "Close", "Volume"])
            has_ohlcv_lower = all(c in cols_to_read for c in ["open", "high", "low", "close", "volume"])

            if has_time and (has_ohlcv_pascal or has_ohlcv_lower):
                return pd.read_parquet(
                    input_path,
                    columns=cols_to_read,
                    engine=PARQUET_ENGINE,
                )

        except Exception:
            pass

    return pd.read_parquet(input_path, engine=PARQUET_ENGINE)






def process_one_series(item: Dict[str, Any]) -> Dict[str, Any]:
    info = item["info"]
    quality_report = item.get("quality_report", {})

    series_id = info["series_id"]
    dataset_kind = info["dataset_kind"]
    asset = info["asset"]
    symbol = info["symbol"]
    source = info["source"]
    timeframe = info["timeframe"]
    timeframe_from_path = info.get("timeframe_from_path")
    timeframe_from_map = info.get("timeframe_from_map")

    quality_status_from_map = info["quality_status"]

    input_path = Path(info["absolute_path"])
    output_path = build_feature_output_path(info)


    quality_record = get_series_quality_record(quality_report, series_id)

    quality_report_missing = (
        not isinstance(quality_report, dict)
        or quality_report.get("status") == "MISSING_REPORT"
        or not quality_report.get("_normalized_series")
    )

    if quality_report_missing:
        normalized_quality_status = normalize_quality_status(quality_status_from_map)

        quality_record = {
            "series_id": series_id,
            "status": quality_status_from_map,
            "normalized_status": normalized_quality_status,
            "source": "FALLBACK_FROM_MAPA_ATIVOS_BECAUSE_DATA_QUALITY_REPORT_MISSING",
            "checks": {},
        }
    else:
        normalized_quality_status = quality_record.get("normalized_status", "WARNING")

    

    result = {
        "run_id": RUN_ID,
        "series_id": series_id,
        "asset": asset,
        "symbol": symbol,
        "source": source,
        "timeframe": timeframe,
        "timeframe_from_path": timeframe_from_path,
        "timeframe_from_map": timeframe_from_map,
        "timeframe_seconds": infer_timeframe_seconds(timeframe),
        "turbo_mode_used": is_turbo_timeframe(timeframe),
        "dataset_kind": dataset_kind,

        "quality_status_input_from_map": quality_status_from_map,
        "quality_status_from_report": quality_record.get("status"),
        "quality_status_normalized": normalized_quality_status,
        "quality_ml_status": get_ml_quality_status_from_record(quality_record),
        "quality_ml_usable_for_broad_training": is_quality_usable_for_ml(
            quality_record,
            normalized_quality_status,
        ),
        "quality_ml_blocking_reasons": quality_record.get("ml_blocking_reasons", []),
        "quality_record": quality_record,

        "input_path": path_to_str(input_path),
        "output_path": path_to_str(output_path),
        "status": "PENDING",
    }

    start = time.time()
    result["memory_mb_start"] = get_process_memory_mb()
    phase_timings: List[Dict[str, Any]] = []
    result["phase_timings"] = phase_timings

    def begin_phase(name: str) -> float:
        result["current_phase"] = name
        return time.time()

    def finish_phase(name: str, phase_started_at: float) -> None:
        phase_timings.append({
            "phase": name,
            "elapsed_seconds": round(time.time() - phase_started_at, 6),
            "memory_mb": get_process_memory_mb(),
        })

    try:
        # ---------------------------------------------------------------------
        # Fase 2 e 3 — Quality gate antes de ler/processar tudo
        # ---------------------------------------------------------------------
        if ENABLE_PRE_FEATURE_QUALITY_GATE:
            if not can_process_quality_status(normalized_quality_status):
                result.update({
                    "status": "SKIPPED_DATA_QUALITY_FAIL",
                    "reason": "DATA_QUALITY_FAIL",
                    "elapsed_seconds": round(time.time() - start, 6),
                })
                return result

        if not input_path.exists():
            raise FileNotFoundError(f"Arquivo Parquet de entrada não encontrado: {input_path}")

        if output_path.exists() and not OVERWRITE_EXISTING:
            result["status"] = "SKIPPED_EXISTS"
            result["elapsed_seconds"] = round(time.time() - start, 6)
            return result

        # ---------------------------------------------------------------------
        # Leitura e preparação
        # ---------------------------------------------------------------------
        phase_started_at = begin_phase("read_parquet")
        df_raw = read_ohlcv_parquet_fast(input_path)

        result["raw_rows"] = int(len(df_raw))
        result["raw_columns"] = int(len(df_raw.columns))
        result["memory_mb_after_read"] = get_process_memory_mb()
        finish_phase("read_parquet", phase_started_at)

        phase_started_at = begin_phase("prepare_ohlcv")
        df_prepared = prepare_ohlcv(df_raw)

        result["prepared_rows"] = int(len(df_prepared))
        result["prepared_columns"] = int(len(df_prepared.columns))
        result["memory_mb_after_prepare"] = get_process_memory_mb()
        finish_phase("prepare_ohlcv", phase_started_at)

        del df_raw


        required_min_rows = min_rows_for_timeframe(timeframe)
        result["required_min_rows"] = int(required_min_rows)

        if len(df_prepared) < required_min_rows:
            raise ValueError(
                f"Série com poucas linhas após preparação: "
                f"{len(df_prepared)} < {required_min_rows}"
            ) 



        # ---------------------------------------------------------------------
        # Fase 1 — Contrato de entrada
        # ---------------------------------------------------------------------
        input_contract_report = {}

        if ENABLE_INPUT_CONTRACT_VALIDATION:
            phase_started_at = begin_phase("input_contract_validation")
            input_contract_report = validate_input_contract(df_prepared, info)
            result["input_contract_report"] = input_contract_report
            finish_phase("input_contract_validation", phase_started_at)

            if input_contract_report.get("status") == "FAIL" and BLOCK_ON_INPUT_CONTRACT_FAIL:
                result.update({
                    "status": "ERROR_INPUT_CONTRACT_FAIL",
                    "error": "Input contract failed.",
                    "elapsed_seconds": round(time.time() - start, 6),
                })
                return result

        # ---------------------------------------------------------------------
        # Fase 4 — Validação temporal forte
        # ---------------------------------------------------------------------
        time_grid_report = {}

        if ENABLE_TIME_GRID_VALIDATION:
            phase_started_at = begin_phase("time_grid_validation")
            time_grid_report = validate_time_grid(
                df=df_prepared,
                timeframe=timeframe,
                info=info,
                max_gap_multiplier=POST_AUDIT_MAX_GAP_MULTIPLIER,
            )
            result["time_grid_report"] = time_grid_report
            finish_phase("time_grid_validation", phase_started_at)

            if time_grid_report.get("status") == "FAIL" and BLOCK_ON_TIME_GRID_FAIL:
                result.update({
                    "status": "SKIPPED_TIME_GRID_FAIL",
                    "reason": "TIME_GRID_FAIL",
                    "elapsed_seconds": round(time.time() - start, 6),
                })
                return result

        # ---------------------------------------------------------------------
        # Metadados pré-feature
        # ---------------------------------------------------------------------
        phase_started_at = begin_phase("add_metadata")
        df = add_metadata_columns(
            df=df_prepared,
            asset=asset,
            symbol=symbol,
            source=source,
            timeframe=timeframe,
            dataset_kind=dataset_kind,
            series_id=series_id,
            timeframe_from_path=timeframe_from_path,
            timeframe_from_map=timeframe_from_map,
            quality_status=normalized_quality_status,
            quality_record=quality_record,
            input_contract_report=input_contract_report,
            time_grid_report=time_grid_report,
            input_path=input_path,
        )
        finish_phase("add_metadata", phase_started_at)

        # ---------------------------------------------------------------------
        # Geração de features
        # ---------------------------------------------------------------------
        phase_started_at = begin_phase("generate_all_features")
        df, feature_metadata, execution_steps = generate_all_features(df, timeframe=timeframe)
        compute_backend_step = next(
            (step for step in execution_steps if step.get("family") == "__compute_backend__"),
            {},
        )
        cuda_vector_steps = [
            step for step in execution_steps
            if step.get("family") == "__cuda_vector_core__"
        ]
        cuda_rolling_steps = [
            step for step in execution_steps
            if step.get("family") == "__cuda_rolling__"
        ]
        result["feature_compute_backend"] = compute_backend_step
        result["feature_cuda_vector_steps"] = cuda_vector_steps
        result["feature_cuda_rolling_steps"] = cuda_rolling_steps
        result["memory_mb_after_feature_generation"] = get_process_memory_mb()
        finish_phase("generate_all_features", phase_started_at)

        max_lookback_bars = infer_max_lookback_from_feature_metadata(feature_metadata)

        # ---------------------------------------------------------------------
        # Fase 6 — Warm-up e validade das features
        # ---------------------------------------------------------------------
        phase_started_at = begin_phase("feature_validity_flags")
        df = replace_inf_with_nan(df)

        df = apply_feature_validity_flags(
            df=df,
            max_lookback_bars=max_lookback_bars,
        )
        finish_phase("feature_validity_flags", phase_started_at)

        # ---------------------------------------------------------------------
        # Fase 8 — ML-ready flags
        # ---------------------------------------------------------------------
        phase_started_at = begin_phase("ml_ready_flags")
        df = apply_ml_ready_schema_flags(df)
        finish_phase("ml_ready_flags", phase_started_at)

        # ---------------------------------------------------------------------
        # Seleção, downcast e auditoria pós-geração
        # ---------------------------------------------------------------------
        phase_started_at = begin_phase("select_output_columns")
        df = select_output_columns(df)
        finish_phase("select_output_columns", phase_started_at)

        phase_started_at = begin_phase("downcast_float_columns")
        df = downcast_float_columns(df)
        result["memory_mb_after_downcast"] = get_process_memory_mb()
        finish_phase("downcast_float_columns", phase_started_at)

        phase_started_at = begin_phase("post_generation_audit")
        post_audit = post_generation_audit(df_prepared, df, timeframe) if ENABLE_POST_AUDIT else {}
        finish_phase("post_generation_audit", phase_started_at)

        if ENABLE_POST_AUDIT and post_audit.get("audit_status") in {"FAIL", "ERROR"}:
            raise ValueError(f"Post-generation audit failed: {post_audit}")

        # ---------------------------------------------------------------------
        # Escrita atômica
        # ---------------------------------------------------------------------
        ensure_dir(output_path.parent)

        phase_started_at = begin_phase("write_parquet")
        tmp_path = output_path.with_suffix(".tmp.parquet")
        df.to_parquet(
            tmp_path,
            index=False,
            engine=PARQUET_ENGINE,
            compression=PARQUET_COMPRESSION,
            row_group_size=PARQUET_ROW_GROUP_SIZE,
        )
        tmp_path.replace(output_path)
        result["memory_mb_after_write"] = get_process_memory_mb()
        finish_phase("write_parquet", phase_started_at)

        phase_started_at = begin_phase("quality_summary")
        quality_summary = compute_quality_summary(df)
        finish_phase("quality_summary", phase_started_at)

        feature_cols = [c for c in df.columns if str(c).startswith("feat_")]
        profile_step = next(
            (step for step in execution_steps if step.get("family") == "__profile__"),
            {},
        )
        family_benchmark = [
            step for step in execution_steps
            if step.get("family") != "__profile__"
        ]
        family_benchmark_sorted = sorted(
            family_benchmark,
            key=lambda item: float(item.get("elapsed_seconds") or 0.0),
            reverse=True,
        )
        phase_timings_sorted = sorted(
            phase_timings,
            key=lambda item: float(item.get("elapsed_seconds") or 0.0),
            reverse=True,
        )

        result.update({
            "status": "OK",
            "output_rows": int(len(df)),
            "output_columns": int(len(df.columns)),
            "feature_columns_count": int(len(feature_cols)),
            "feature_metadata_count": len(feature_metadata),
            "feature_families_executed": execution_steps,
            "feature_compute_backend": compute_backend_step,
            "feature_cuda_vector_steps": cuda_vector_steps,
            "feature_cuda_rolling_steps": cuda_rolling_steps,
            "feature_execution_profile": profile_step,
            "feature_family_benchmark": family_benchmark_sorted,
            "feature_family_slowest": family_benchmark_sorted[:5],
            "phase_timings": phase_timings,
            "phase_slowest": phase_timings_sorted[:8],

            "max_lookback_bars": int(max_lookback_bars),
            "valid_feature_rows": quality_summary.get("valid_feature_rows"),
            "warmup_rows": quality_summary.get("warmup_rows"),
            "ml_eligible_rows": quality_summary.get("ml_eligible_rows"),
            "first_valid_timestamp_utc_ms": quality_summary.get("first_valid_timestamp_utc_ms"),

            "quality_summary": quality_summary,
            "post_audit": post_audit,
            "elapsed_seconds": round(time.time() - start, 6),
        })

        del df
        del df_prepared

        if FORCE_GC_EACH_SERIES:
            gc.collect()

        return result

    except Exception as exc:
        result.update({
            "status": "ERROR",
            "error": str(exc),
            "traceback": traceback.format_exc(),
            "elapsed_seconds": round(time.time() - start, 6),
        })

        if FORCE_GC_EACH_SERIES:
            gc.collect()

        return result



















# =============================================================================
# 11. FEATURE CATALOG
# =============================================================================

def build_feature_catalog_from_config() -> List[Dict[str, Any]]:
    dates = pd.date_range("2020-01-01", periods=700, freq="min")

    base = np.linspace(100, 130, len(dates))
    close = base + np.sin(np.arange(len(dates)) / 7) * 2 + np.sin(np.arange(len(dates)) / 31) * 4
    open_ = close * (1 + np.sin(np.arange(len(dates)) / 11) * 0.001)
    high = np.maximum(open_, close) * 1.002
    low = np.minimum(open_, close) * 0.998
    volume = np.full(len(dates), 1000.0) + np.sin(np.arange(len(dates)) / 5) * 100


    df_sample = pd.DataFrame({
        "DateTime": dates,
        "Open": open_,
        "High": high,
        "Low": low,
        "Close": close,
        "Volume": volume,
    })


    df_sample = ensure_archangel_time_columns(df_sample)


    _, metadata, _ = generate_all_features(df_sample, timeframe="5min")

    seen = set()
    clean = []

    for item in metadata:
        feature = item.get("feature")

        if feature in seen:
            continue

        seen.add(feature)
        clean.append(item)

    return clean


# =============================================================================
# 12. REPORTS JSON E EXCEL UX
# =============================================================================

def build_dashboard_df(selected: List[Dict[str, Any]], results: List[Dict[str, Any]], feature_catalog: List[Dict[str, Any]]) -> pd.DataFrame:
    ok = [r for r in results if r.get("status") == "OK"]
    errors = [r for r in results if str(r.get("status", "")).startswith("ERROR")]
    skipped = [r for r in results if str(r.get("status", "")).startswith("SKIPPED")]

    total_elapsed = sum(float(r.get("elapsed_seconds") or 0) for r in results)
    avg_elapsed = total_elapsed / max(1, len(results))

    audit_pass = sum(1 for r in ok if (r.get("post_audit") or {}).get("audit_status") == "PASS")
    audit_warning = sum(1 for r in ok if (r.get("post_audit") or {}).get("audit_status") == "WARNING")
    audit_fail = sum(1 for r in ok if (r.get("post_audit") or {}).get("audit_status") in {"FAIL", "ERROR"})

    rows = [
        {"section": "RUN", "metric": "Run ID", "value": RUN_ID, "interpretation": "Identificador único da execução."},
        {"section": "RUN", "metric": "Generated At", "value": now_iso(), "interpretation": "Horário de geração."},
        {"section": "RUN", "metric": "Schema", "value": SCHEMA_VERSION, "interpretation": "Versão do feature store."},

        {"section": "COVERAGE", "metric": "Series Selected", "value": len(selected), "interpretation": "Séries elegíveis via MAPA_ATIVOS."},
        {"section": "COVERAGE", "metric": "Series OK", "value": len(ok), "interpretation": "Séries processadas com sucesso."},
        {"section": "COVERAGE", "metric": "Series Error", "value": len(errors), "interpretation": "Séries com erro."},
        {"section": "COVERAGE", "metric": "Series Skipped", "value": len(skipped), "interpretation": "Séries puladas."},

        {"section": "FEATURES", "metric": "Feature Catalog Count", "value": len(feature_catalog), "interpretation": "Número de features únicas no catálogo."},
        {"section": "FEATURES", "metric": "Feature Families", "value": len(set(x.get("family") for x in feature_catalog)), "interpretation": "Famílias de features."},

        {"section": "PERFORMANCE", "metric": "Parallel Processing", "value": ENABLE_PARALLEL_PROCESSING, "interpretation": "Processamento paralelo ativo."},
        {"section": "PERFORMANCE", "metric": "Max Workers", "value": MAX_WORKERS_FEATURES, "interpretation": "Número máximo de workers."},
        {"section": "PERFORMANCE", "metric": "Avg Seconds / Series", "value": round(avg_elapsed, 3), "interpretation": "Tempo médio por série."},
        {"section": "PERFORMANCE", "metric": "Turbo Timeframes", "value": ", ".join(sorted(TURBO_TIMEFRAMES)), "interpretation": "Timeframes com modo turbo."},

        {"section": "AUDIT", "metric": "Post Audit PASS", "value": audit_pass, "interpretation": "Séries sincronizadas e aprovadas."},
        {"section": "AUDIT", "metric": "Post Audit WARNING", "value": audit_warning, "interpretation": "Séries com alertas, mas não bloqueadas."},
        {"section": "AUDIT", "metric": "Post Audit FAIL/ERROR", "value": audit_fail, "interpretation": "Séries com falha crítica."},

        {"section": "ANTI-LEAKAGE", "metric": "Uses Future Data", "value": False, "interpretation": "O script não usa dados futuros."},
        {"section": "ANTI-LEAKAGE", "metric": "Creates Labels", "value": False, "interpretation": "Labels devem ser criados em módulo separado."},
    ]

    return pd.DataFrame(rows)



def build_series_excel_df(results: List[Dict[str, Any]]) -> pd.DataFrame:
    rows = []

    for r in results:
        q = r.get("quality_summary", {}) if isinstance(r.get("quality_summary"), dict) else {}
        pa = r.get("post_audit", {}) if isinstance(r.get("post_audit"), dict) else {}
        checks = pa.get("checks", {}) if isinstance(pa.get("checks"), dict) else {}

        input_contract = r.get("input_contract_report", {}) if isinstance(r.get("input_contract_report"), dict) else {}
        time_grid = r.get("time_grid_report", {}) if isinstance(r.get("time_grid_report"), dict) else {}
        time_grid_checks = time_grid.get("checks", {}) if isinstance(time_grid.get("checks"), dict) else {}

        rows.append({
            "status": r.get("status"),
            "audit_status": pa.get("audit_status"),

            "quality_status_normalized": r.get("quality_status_normalized"),
            "quality_status_from_report": r.get("quality_status_from_report"),
            "quality_status_input_from_map": r.get("quality_status_input_from_map"),

            "input_contract_status": input_contract.get("status"),
            "time_grid_status": time_grid.get("status"),

            "asset": r.get("asset"),
            "symbol": r.get("symbol"),
            "source": r.get("source"),
            "timeframe": r.get("timeframe"),
            "turbo_mode": r.get("turbo_mode_used"),
            "batch_name": r.get("batch_name"),

            "raw_rows": r.get("raw_rows"),
            "prepared_rows": r.get("prepared_rows"),
            "required_min_rows": r.get("required_min_rows"),
            "output_rows": r.get("output_rows"),


            "feature_count": r.get("feature_columns_count"),
            "feature_execution_profile": (r.get("feature_execution_profile") or {}).get("feature_execution_profile"),
            "ram_safe_applied": (r.get("feature_execution_profile") or {}).get("ram_safe_applied"),
            "disabled_families": ", ".join((r.get("feature_execution_profile") or {}).get("disabled_families", [])),
            "max_lookback_bars": r.get("max_lookback_bars"),
            "warmup_rows": r.get("warmup_rows"),
            "valid_feature_rows": r.get("valid_feature_rows"),
            "ml_eligible_rows": r.get("ml_eligible_rows"),

            "feature_null_ratio": q.get("feature_null_ratio"),
            "timestamp_utc_ms_present": q.get("timestamp_utc_ms_present"),
            "feature_ready_timestamp_present": q.get("feature_ready_timestamp_present"),

            "first_timestamp_utc_ms": q.get("first_timestamp_utc_ms"),
            "last_timestamp_utc_ms": q.get("last_timestamp_utc_ms"),
            "first_valid_timestamp_utc_ms": q.get("first_valid_timestamp_utc_ms"),

            "datetime_sync": checks.get("datetime_exact_sequence_match"),
            "timestamp_sync": checks.get("timestamp_utc_ms_exact_sequence_match"),

            "pre_time_grid_irregular_count": time_grid_checks.get("irregular_delta_count"),
            "pre_time_grid_large_gap_count": time_grid_checks.get("large_gap_count"),
            "pre_time_grid_irregular_ratio": time_grid_checks.get("irregular_ratio"),
            "pre_time_grid_large_gap_ratio": time_grid_checks.get("large_gap_ratio"),

            "large_gap_count": checks.get("large_gap_count"),
            "large_gap_count_timestamp_utc_ms": checks.get("large_gap_count_timestamp_utc_ms"),
            "median_delta_seconds": checks.get("median_delta_seconds"),
            "median_delta_ms": checks.get("median_delta_ms"),
            "expected_tf_seconds": checks.get("expected_timeframe_seconds"),
            "expected_tf_ms": checks.get("expected_timeframe_ms"),

            "bar_timestamp_policy": q.get("bar_timestamp_policy"),
            "elapsed_seconds": r.get("elapsed_seconds"),
            "memory_mb_start": r.get("memory_mb_start"),
            "memory_mb_after_read": r.get("memory_mb_after_read"),
            "memory_mb_after_prepare": r.get("memory_mb_after_prepare"),
            "memory_mb_after_feature_generation": r.get("memory_mb_after_feature_generation"),
            "memory_mb_after_downcast": r.get("memory_mb_after_downcast"),
            "memory_mb_after_write": r.get("memory_mb_after_write"),

            "series_id": r.get("series_id"),
            "output_path": r.get("output_path"),
            "error": r.get("error"),
        })

    return pd.DataFrame(rows)





def build_feature_families_df(feature_catalog: List[Dict[str, Any]]) -> pd.DataFrame:
    rows = []

    for family in sorted(set(x.get("family") for x in feature_catalog if x.get("family"))):
        fam_items = [x for x in feature_catalog if x.get("family") == family]
        max_lb = max([x.get("lookback") or 0 for x in fam_items]) if fam_items else 0

        rows.append({
            "family": family,
            "enabled": bool(FEATURE_CONFIG.get(family, {}).get("enabled", False)),
            "feature_count": len(fam_items),
            "max_lookback": max_lb,
            "description": FEATURE_FAMILY_DESCRIPTIONS.get(family),
            "ml_value": classify_family_ml_value(family),
            "risk_value": classify_family_risk_value(family),
            "speed_cost": classify_family_speed_cost(family),
        })

    return pd.DataFrame(rows)


def classify_family_ml_value(family: str) -> str:
    high = {"returns", "return_lags", "rolling_return", "volatility", "ema", "rsi", "bollinger", "regime", "time_context", "shock_features"}
    medium = {"macd", "sma", "volume", "trend_strength", "market_structure", "oscillators", "donchian"}
    if family in high:
        return "HIGH"
    if family in medium:
        return "MEDIUM"
    return "SPECIALIZED"


def classify_family_risk_value(family: str) -> str:
    high = {"volatility", "vol_of_vol", "atr_range_volatility", "risk_proxy", "liquidity_proxy", "drawdown_distance"}
    medium = {"shock_features", "volume", "volume_pressure", "regime"}
    if family in high:
        return "HIGH"
    if family in medium:
        return "MEDIUM"
    return "LOW"


def classify_family_speed_cost(family: str) -> str:
    expensive = {"autocorrelation", "higher_moments", "slopes", "risk_proxy"}
    medium = {"bollinger", "donchian", "range_volatility_estimators", "volume"}
    if family in expensive:
        return "HIGH"
    if family in medium:
        return "MEDIUM"
    return "LOW"


def build_feature_catalog_excel_df(feature_catalog: List[Dict[str, Any]]) -> pd.DataFrame:
    df = pd.DataFrame(feature_catalog)

    keep_cols = [
        "feature",
        "family",
        "description",
        "formula",
        "lookback",
        "uses_future_data",
        "risk_relevance",
        "ml_relevance",
    ]

    keep_cols = [c for c in keep_cols if c in df.columns]

    if df.empty:
        return df

    return df.loc[:, keep_cols].copy()


def build_post_audit_df(results: List[Dict[str, Any]]) -> pd.DataFrame:
    rows = []

    for r in results:
        pa = r.get("post_audit", {}) if isinstance(r.get("post_audit"), dict) else {}
        checks = pa.get("checks", {}) if isinstance(pa.get("checks"), dict) else {}

        rows.append({
            "asset": r.get("asset"),
            "symbol": r.get("symbol"),
            "source": r.get("source"),
            "timeframe": r.get("timeframe"),
            "status": r.get("status"),
            "audit_status": pa.get("audit_status"),
            "row_count_match": checks.get("row_count_match"),
            "datetime_exact_sequence_match": checks.get("datetime_exact_sequence_match"),
            "timestamp_utc_ms_input_present": checks.get("timestamp_utc_ms_input_present"),
            "timestamp_utc_ms_output_present": checks.get("timestamp_utc_ms_output_present"),
            "timestamp_utc_ms_exact_sequence_match": checks.get("timestamp_utc_ms_exact_sequence_match"),
            "output_datetime_monotonic": checks.get("output_datetime_monotonic"),
            "output_timestamp_utc_ms_monotonic": checks.get("output_timestamp_utc_ms_monotonic"),
            "output_datetime_duplicates": checks.get("output_datetime_duplicates"),
            "output_timestamp_utc_ms_duplicates": checks.get("output_timestamp_utc_ms_duplicates"),
            "expected_timeframe_seconds": checks.get("expected_timeframe_seconds"),
            "expected_timeframe_ms": checks.get("expected_timeframe_ms"),
            "median_delta_seconds": checks.get("median_delta_seconds"),
            "median_delta_ms": checks.get("median_delta_ms"),
            "max_delta_seconds": checks.get("max_delta_seconds"),
            "max_delta_ms": checks.get("max_delta_ms"),
            "large_gap_count": checks.get("large_gap_count"),
            "large_gap_count_timestamp_utc_ms": checks.get("large_gap_count_timestamp_utc_ms"),
            "overall_feature_null_ratio": checks.get("overall_feature_null_ratio"),
            "warnings": " | ".join(pa.get("warnings", [])) if isinstance(pa.get("warnings"), list) else None,
            "errors": " | ".join(pa.get("errors", [])) if isinstance(pa.get("errors"), list) else None,
            "output_path": r.get("output_path"),
        })

    return pd.DataFrame(rows)

def build_errors_df(results: List[Dict[str, Any]]) -> pd.DataFrame:
    rows = []

    for r in results:
        if not str(r.get("status", "")).startswith("ERROR"):
            continue

        rows.append({
            "status": r.get("status"),
            "asset": r.get("asset"),
            "symbol": r.get("symbol"),
            "source": r.get("source"),
            "timeframe": r.get("timeframe"),
            "series_id": r.get("series_id"),
            "raw_rows": r.get("raw_rows"),
            "prepared_rows": r.get("prepared_rows"),
            "required_min_rows": r.get("required_min_rows"),
            "error": r.get("error"),
            "elapsed_seconds": r.get("elapsed_seconds"),
            "input_path": r.get("input_path"),
            "output_path": r.get("output_path"),
                        
        })

    return pd.DataFrame(rows)


def get_latest_result_by_series(results: List[Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
    latest: Dict[str, Dict[str, Any]] = {}

    for result in results:
        series_id = str(result.get("series_id") or "")
        if not series_id:
            continue
        latest[series_id] = result

    return latest


def build_retry_plan_payload(results: List[Dict[str, Any]], selected: List[Dict[str, Any]]) -> Dict[str, Any]:
    latest_by_series = get_latest_result_by_series(results)
    selected_by_series = {
        str((item.get("info") or {}).get("series_id")): item
        for item in selected
        if (item.get("info") or {}).get("series_id")
    }

    retryable_series = []
    non_retryable_failures = []

    for series_id, result in latest_by_series.items():
        status = str(result.get("status") or "")
        if not status.startswith("ERROR"):
            continue

        entry = {
            "series_id": series_id,
            "status": status,
            "error_classification": classify_result_error(result),
            "retryable": is_retryable_result(result),
            "retry_attempt": result.get("retry_attempt"),
            "retry_scheduled": result.get("retry_scheduled"),
            "asset": result.get("asset"),
            "symbol": result.get("symbol"),
            "source": result.get("source"),
            "timeframe": result.get("timeframe"),
            "input_path": result.get("input_path"),
            "output_path": result.get("output_path"),
            "error": result.get("error"),
            "elapsed_seconds": result.get("elapsed_seconds"),
            "selected_for_retry_mode": series_id in selected_by_series,
        }

        if entry["retryable"]:
            retryable_series.append(entry)
        else:
            non_retryable_failures.append(entry)

    retryable_series = sorted(
        retryable_series,
        key=lambda x: (
            str(x.get("source") or ""),
            str(x.get("asset") or ""),
            str(x.get("timeframe") or ""),
            str(x.get("series_id") or ""),
        ),
    )
    non_retryable_failures = sorted(
        non_retryable_failures,
        key=lambda x: (
            str(x.get("source") or ""),
            str(x.get("asset") or ""),
            str(x.get("timeframe") or ""),
            str(x.get("series_id") or ""),
        ),
    )

    return {
        "schema_version": "ARCHANGEL_FEATURES_RETRY_PLAN_1.0",
        "generated_at": now_iso(),
        "run_id": RUN_ID,
        "script": SCRIPT_NAME,
        "source_run_report_path": path_to_str(RUN_REPORT_PATH),
        "source_run_report_base_json_path": path_to_str(RUN_REPORT_BASE_JSON_PATH),
        "latest_run_report_base_json_path": path_to_str(RUN_REPORT_LATEST_PATH),
        "retry_mode_env": "ARCHANGEL_RETRY_FAILED_ONLY_FROM_REPORT",
        "retry_mode_example": (
            f"$env:ARCHANGEL_RETRY_FAILED_ONLY_FROM_REPORT=\"{path_to_str(RUN_REPORT_PATH)}\"; "
            f"& \"python\" \"{path_to_str(BASE_REGRAS_DIR / '03_GERA_FEATURES.py')}\""
        ),
        "summary": {
            "series_selected": len(selected),
            "series_attempted_including_retries": len(results),
            "latest_series_with_error": len(retryable_series) + len(non_retryable_failures),
            "retryable_series_count": len(retryable_series),
            "non_retryable_failure_count": len(non_retryable_failures),
            "recovered_by_retry_count": sum(1 for r in results if r.get("recovered_by_retry")),
        },
        "retryable_series": retryable_series,
        "non_retryable_failures": non_retryable_failures,
    }


def save_retry_plan(results: List[Dict[str, Any]], selected: List[Dict[str, Any]]) -> Tuple[Path, Path]:
    ensure_dir(FEATURES_LOG_DIR)

    payload = build_retry_plan_payload(results, selected)
    write_json_atomic(payload, RETRY_OUTPUT_JSON_PATH)
    write_json_atomic(payload, RETRY_OUTPUT_BASE_JSON_PATH)
    write_json_atomic(payload, RETRY_OUTPUT_LATEST_PATH)

    rows = payload.get("retryable_series", []) + payload.get("non_retryable_failures", [])
    pd.DataFrame(rows).to_csv(RETRY_OUTPUT_CSV_PATH, index=False, encoding="utf-8-sig")

    return RETRY_OUTPUT_JSON_PATH, RETRY_OUTPUT_CSV_PATH


def load_retry_series_ids(path: Path) -> set[str]:
    payload = load_json(path)
    series_ids: set[str] = set()

    if isinstance(payload.get("retryable_series"), list):
        for item in payload["retryable_series"]:
            if isinstance(item, dict) and item.get("series_id"):
                series_ids.add(str(item["series_id"]))
        return series_ids

    if isinstance(payload.get("results"), list):
        latest = get_latest_result_by_series(payload["results"])
        for series_id, result in latest.items():
            if is_retryable_result(result):
                series_ids.add(series_id)
        return series_ids

    raise ValueError(
        f"Arquivo de retry inválido: {path}. Esperado retryable_series ou results."
    )


def filter_selected_for_retry_only(selected: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    if not RETRY_FAILED_ONLY_FROM_REPORT:
        return selected

    retry_path = Path(RETRY_FAILED_ONLY_FROM_REPORT)
    retry_series_ids = load_retry_series_ids(retry_path)
    filtered = [
        item for item in selected
        if str((item.get("info") or {}).get("series_id")) in retry_series_ids
    ]

    print(
        f"[RETRY-ONLY] origem={retry_path} | series_retry={len(retry_series_ids)} | "
        f"selecionadas={len(filtered)}",
        flush=True,
    )

    return filtered


def aggregate_feature_family_benchmark(results: List[Dict[str, Any]]) -> Dict[str, Any]:
    aggregate: Dict[str, Dict[str, Any]] = {}

    for result in results:
        for step in result.get("feature_family_benchmark", []) or []:
            if not isinstance(step, dict):
                continue
            family = str(step.get("family") or "")
            if not family:
                continue
            elapsed = float(step.get("elapsed_seconds") or 0.0)
            status = str(step.get("status") or "")
            item = aggregate.setdefault(family, {
                "runs": 0,
                "ok_runs": 0,
                "skipped_runs": 0,
                "total_elapsed_seconds": 0.0,
                "max_elapsed_seconds": 0.0,
            })
            item["runs"] += 1
            item["ok_runs"] += int(status == "OK")
            item["skipped_runs"] += int(status.startswith("SKIPPED"))
            item["total_elapsed_seconds"] += elapsed
            item["max_elapsed_seconds"] = max(float(item["max_elapsed_seconds"]), elapsed)

    for item in aggregate.values():
        runs = max(1, int(item["runs"]))
        item["avg_elapsed_seconds"] = round(float(item["total_elapsed_seconds"]) / runs, 6)
        item["total_elapsed_seconds"] = round(float(item["total_elapsed_seconds"]), 6)
        item["max_elapsed_seconds"] = round(float(item["max_elapsed_seconds"]), 6)

    return dict(
        sorted(
            aggregate.items(),
            key=lambda kv: float(kv[1].get("total_elapsed_seconds") or 0.0),
            reverse=True,
        )
    )


def aggregate_feature_compute_backend(results: List[Dict[str, Any]]) -> Dict[str, Any]:
    backend_counts: Dict[str, int] = {}
    vector_backend_counts: Dict[str, int] = {}
    rolling_backend_counts: Dict[str, int] = {}
    rolling_family_counts: Dict[str, int] = {}
    rolling_accelerated_total = 0
    rolling_fallback_total = 0
    rolling_cpu_total = 0
    rolling_cpu_policy_total = 0
    fallback_errors: List[str] = []
    cpu_policy_reasons: List[str] = []

    for result in results:
        if result.get("status") != "OK":
            continue

        backend = (
            (result.get("feature_compute_backend") or {}).get("resolved_backend")
            or "unknown"
        )
        backend_counts[str(backend)] = backend_counts.get(str(backend), 0) + 1

        for step in result.get("feature_cuda_vector_steps") or []:
            if not isinstance(step, dict):
                continue
            step_backend = str(step.get("backend") or "unknown")
            vector_backend_counts[step_backend] = vector_backend_counts.get(step_backend, 0) + 1
            if step.get("error"):
                fallback_errors.append(str(step.get("error")))

        for step in result.get("feature_cuda_rolling_steps") or []:
            if not isinstance(step, dict):
                continue
            step_backend = str(step.get("backend") or "unknown")
            source_family = str(step.get("source_family") or "unknown")
            rolling_backend_counts[step_backend] = rolling_backend_counts.get(step_backend, 0) + 1
            rolling_family_counts[source_family] = rolling_family_counts.get(source_family, 0) + 1
            rolling_accelerated_total += int(step.get("accelerated_count") or 0)
            rolling_fallback_total += int(step.get("fallback_count") or 0)
            rolling_cpu_total += int(step.get("cpu_count") or 0)
            rolling_cpu_policy_total += int(step.get("cpu_policy_count") or 0)
            for error in step.get("errors_first_5") or []:
                fallback_errors.append(str(error))
            for reason in step.get("cpu_policy_reasons_first_5") or []:
                cpu_policy_reasons.append(str(reason))

    return {
        "mode": FEATURE_CUDA_MODE,
        "series_backend_counts": dict(sorted(backend_counts.items())),
        "vector_step_backend_counts": dict(sorted(vector_backend_counts.items())),
        "rolling_step_backend_counts": dict(sorted(rolling_backend_counts.items())),
        "rolling_step_family_counts": dict(sorted(rolling_family_counts.items())),
        "rolling_accelerated_operations": int(rolling_accelerated_total),
        "rolling_fallback_operations": int(rolling_fallback_total),
        "rolling_cpu_operations_reported": int(rolling_cpu_total),
        "rolling_cpu_policy_operations": int(rolling_cpu_policy_total),
        "cuda_series_count": int(backend_counts.get("cupy_cuda", 0)),
        "fallback_errors_first_20": fallback_errors[:20],
        "cpu_policy_reasons_first_20": cpu_policy_reasons[:20],
        "accelerated_blocks": list(FEATURE_CUDA_ACCELERATED_BLOCKS),
        "note": FEATURE_CUDA_NOTE,
    }


def aggregate_phase_timings(results: List[Dict[str, Any]]) -> Dict[str, Any]:
    aggregate: Dict[str, Dict[str, Any]] = {}

    for result in results:
        for step in result.get("phase_timings", []) or []:
            if not isinstance(step, dict):
                continue
            phase = str(step.get("phase") or "")
            if not phase:
                continue
            elapsed = float(step.get("elapsed_seconds") or 0.0)
            item = aggregate.setdefault(phase, {
                "runs": 0,
                "total_elapsed_seconds": 0.0,
                "max_elapsed_seconds": 0.0,
            })
            item["runs"] += 1
            item["total_elapsed_seconds"] += elapsed
            item["max_elapsed_seconds"] = max(float(item["max_elapsed_seconds"]), elapsed)

    for item in aggregate.values():
        runs = max(1, int(item["runs"]))
        item["avg_elapsed_seconds"] = round(float(item["total_elapsed_seconds"]) / runs, 6)
        item["total_elapsed_seconds"] = round(float(item["total_elapsed_seconds"]), 6)
        item["max_elapsed_seconds"] = round(float(item["max_elapsed_seconds"]), 6)

    return dict(
        sorted(
            aggregate.items(),
            key=lambda kv: float(kv[1].get("total_elapsed_seconds") or 0.0),
            reverse=True,
        )
    )






def build_config_df() -> pd.DataFrame:
    rows = [
        {"config": "SCHEMA_VERSION", "value": SCHEMA_VERSION},
        {"config": "INPUT_CONTRACT_VERSION", "value": INPUT_CONTRACT_VERSION},
        {"config": "QUALITY_GATE_VERSION", "value": QUALITY_GATE_VERSION},
        {"config": "ML_READINESS_VERSION", "value": ML_READINESS_VERSION},

        {"config": "TIMEZONE_LOCAL", "value": TIMEZONE_LOCAL},
        {"config": "DATETIME_COL", "value": DATETIME_COL},
        {"config": "TIMESTAMP_UTC_MS_COL", "value": TIMESTAMP_UTC_MS_COL},
        {"config": "BAR_TIMESTAMP_POLICY", "value": BAR_TIMESTAMP_POLICY},

        {"config": "DATA_QUALITY_REPORT_PATH", "value": path_to_str(DATA_QUALITY_REPORT_PATH)},
        {"config": "COST_MODEL_PATH", "value": path_to_str(COST_MODEL_PATH)},

        {"config": "ENABLE_PRE_FEATURE_QUALITY_GATE", "value": ENABLE_PRE_FEATURE_QUALITY_GATE},
        {"config": "ALLOW_WARNING_QUALITY_TO_PROCESS", "value": ALLOW_WARNING_QUALITY_TO_PROCESS},
        {"config": "DEFAULT_QUALITY_STATUS_IF_MISSING", "value": DEFAULT_QUALITY_STATUS_IF_MISSING},
        {"config": "BLOCK_ON_QUALITY_FAIL", "value": BLOCK_ON_QUALITY_FAIL},
        {"config": "BLOCK_ON_INPUT_CONTRACT_FAIL", "value": BLOCK_ON_INPUT_CONTRACT_FAIL},
        {"config": "BLOCK_ON_TIME_GRID_FAIL", "value": BLOCK_ON_TIME_GRID_FAIL},

        {"config": "ENABLE_INPUT_CONTRACT_VALIDATION", "value": ENABLE_INPUT_CONTRACT_VALIDATION},
        {"config": "ENABLE_TIME_GRID_VALIDATION", "value": ENABLE_TIME_GRID_VALIDATION},
        {"config": "ENABLE_TIMEZONE_CONSISTENCY_CHECK", "value": ENABLE_TIMEZONE_CONSISTENCY_CHECK},
        {"config": "BLOCK_ON_TIMEZONE_CONSISTENCY_FAIL", "value": BLOCK_ON_TIMEZONE_CONSISTENCY_FAIL},
        {"config": "ENABLE_FEATURE_READY_TIMESTAMP", "value": ENABLE_FEATURE_READY_TIMESTAMP},
        {"config": "ENABLE_FEATURE_VALIDITY_FLAGS", "value": ENABLE_FEATURE_VALIDITY_FLAGS},
        {"config": "ENABLE_ML_READY_SCHEMA_FLAGS", "value": ENABLE_ML_READY_SCHEMA_FLAGS},


        {"config": "MIN_FEATURE_NON_NULL_RATIO_FOR_VALID_ROW", "value": MIN_FEATURE_NON_NULL_RATIO_FOR_VALID_ROW},


        {"config": "MAX_TIMESTAMP_DATETIME_DRIFT_MS", "value": MAX_TIMESTAMP_DATETIME_DRIFT_MS},
        {"config": "TIME_GRID_FAIL_LARGE_GAP_RATIO_THRESHOLD", "value": TIME_GRID_FAIL_LARGE_GAP_RATIO_THRESHOLD},
        {"config": "TIME_GRID_FAIL_IRREGULAR_RATIO_THRESHOLD", "value": TIME_GRID_FAIL_IRREGULAR_RATIO_THRESHOLD},

        {"config": "METADATA_OUTPUT_MODE", "value": METADATA_OUTPUT_MODE},
        {"config": "READ_OPTIONAL_MARKET_COLUMNS", "value": READ_OPTIONAL_MARKET_COLUMNS},
        {"config": "INCLUDE_OPTIONAL_MARKET_COLUMNS_IN_OUTPUT", "value": INCLUDE_OPTIONAL_MARKET_COLUMNS_IN_OUTPUT},
        {"config": "FEATURE_RESOURCE_PROFILE", "value": FEATURE_RESOURCE_PROFILE},
        {"config": "MAX_WORKERS_FEATURES", "value": MAX_WORKERS_FEATURES},
        {"config": "ENABLE_MEMORY_AWARE_WORKERS", "value": ENABLE_MEMORY_AWARE_WORKERS},
        {"config": "ARCHANGEL_RAM_CAP_GB", "value": ARCHANGEL_RAM_CAP_GB},
        {"config": "TARGET_RAM_USED_GB", "value": TARGET_RAM_USED_GB},
        {"config": "TARGET_CPU_PERCENT", "value": TARGET_CPU_PERCENT},
        {"config": "TARGET_FREE_RAM_GB", "value": get_target_free_ram_gb()},
        {"config": "EFFECTIVE_MIN_FREE_RAM_GB", "value": get_effective_min_free_ram_gb()},
        {"config": "MIN_FREE_RAM_GB_TO_START_BATCH", "value": MIN_FREE_RAM_GB_TO_START_BATCH},
        {"config": "ESTIMATED_RAM_GB_PER_WORKER_STANDARD", "value": ESTIMATED_RAM_GB_PER_WORKER_STANDARD},
        {"config": "ESTIMATED_RAM_GB_PER_WORKER_TURBO", "value": ESTIMATED_RAM_GB_PER_WORKER_TURBO},
        {"config": "PROCESS_TURBO_TIMEFRAMES_FIRST", "value": PROCESS_TURBO_TIMEFRAMES_FIRST},
        {"config": "ENABLE_TIMEFRAME_BATCH_EXECUTION", "value": ENABLE_TIMEFRAME_BATCH_EXECUTION},
        {"config": "MAX_WORKERS_TURBO_BATCH", "value": MAX_WORKERS_TURBO_BATCH},
        {"config": "ENABLE_ADAPTIVE_WORKERS", "value": ENABLE_ADAPTIVE_WORKERS},
        {"config": "ADAPTIVE_WORKERS_INITIAL_TURBO", "value": ADAPTIVE_WORKERS_INITIAL_TURBO},
        {"config": "ADAPTIVE_WORKERS_INITIAL_STANDARD", "value": ADAPTIVE_WORKERS_INITIAL_STANDARD},
        {"config": "ADAPTIVE_SCALE_UP_FREE_RAM_GB", "value": ADAPTIVE_SCALE_UP_FREE_RAM_GB},
        {"config": "ADAPTIVE_HOLD_FREE_RAM_GB", "value": ADAPTIVE_HOLD_FREE_RAM_GB},
        {"config": "ADAPTIVE_SCALE_DOWN_FREE_RAM_GB", "value": ADAPTIVE_SCALE_DOWN_FREE_RAM_GB},
        {"config": "ADAPTIVE_SCALE_UP_STABLE_HEARTBEATS", "value": ADAPTIVE_SCALE_UP_STABLE_HEARTBEATS},
        {"config": "ADAPTIVE_SCALE_UP_WORKER_STEP", "value": ADAPTIVE_SCALE_UP_WORKER_STEP},
        {"config": "ADAPTIVE_SCALE_DOWN_WORKER_STEP", "value": ADAPTIVE_SCALE_DOWN_WORKER_STEP},
        {"config": "ADAPTIVE_FAILURE_COOLDOWN_HEARTBEATS", "value": ADAPTIVE_FAILURE_COOLDOWN_HEARTBEATS},
        {"config": "ADAPTIVE_CPU_LOW_SCALE_UP_THRESHOLD", "value": ADAPTIVE_CPU_LOW_SCALE_UP_THRESHOLD},
        {"config": "ENABLE_RETRY_FAILED_SERIES", "value": ENABLE_RETRY_FAILED_SERIES},
        {"config": "MAX_RETRY_ATTEMPTS_PER_SERIES", "value": MAX_RETRY_ATTEMPTS_PER_SERIES},
        {"config": "RETRY_FAILED_ONLY_FROM_REPORT", "value": RETRY_FAILED_ONLY_FROM_REPORT},
        {"config": "ENABLE_PROGRESS_HEARTBEAT", "value": ENABLE_PROGRESS_HEARTBEAT},
        {"config": "PROGRESS_HEARTBEAT_SECONDS", "value": PROGRESS_HEARTBEAT_SECONDS},
        {"config": "PROGRESS_SHOW_ACTIVE_LIMIT", "value": PROGRESS_SHOW_ACTIVE_LIMIT},
        {"config": "PROGRESS_SHOW_RECENT_LIMIT", "value": PROGRESS_SHOW_RECENT_LIMIT},
        {"config": "FEATURE_EXECUTION_PROFILE", "value": FEATURE_EXECUTION_PROFILE},
        {"config": "FAST_FIRST_ROW_THRESHOLD", "value": FAST_FIRST_ROW_THRESHOLD},
        {"config": "FAST_FIRST_SKIP_FAMILIES", "value": ", ".join(sorted(FAST_FIRST_SKIP_FAMILIES))},
        {"config": "RAM_SAFE_ROW_THRESHOLD", "value": RAM_SAFE_ROW_THRESHOLD},
        {"config": "RAM_SAFE_SKIP_FAMILIES_FOR_TURBO", "value": ", ".join(sorted(RAM_SAFE_SKIP_FAMILIES_FOR_TURBO))},
        {"config": "RAM_SAFE_SKIP_FAMILIES_FOR_LONG_SERIES", "value": ", ".join(sorted(RAM_SAFE_SKIP_FAMILIES_FOR_LONG_SERIES))},
        {"config": "ENABLE_FEATURE_BLOCK_CONSTRUCTION", "value": ENABLE_FEATURE_BLOCK_CONSTRUCTION},
        {"config": "FEATURE_BLOCK_COLUMN_COUNT", "value": FEATURE_BLOCK_COLUMN_COUNT},
        {"config": "FORCE_GC_EACH_FEATURE_BLOCK", "value": FORCE_GC_EACH_FEATURE_BLOCK},
        {"config": "ENABLE_PROCESS_MEMORY_TELEMETRY", "value": ENABLE_PROCESS_MEMORY_TELEMETRY},
        {"config": "ENABLE_TIMEFRAME_TURBO_MODE", "value": ENABLE_TIMEFRAME_TURBO_MODE},
        {"config": "TURBO_TIMEFRAMES", "value": ", ".join(sorted(TURBO_TIMEFRAMES))},
        {"config": "TURBO_DISABLE_AUTOCORR", "value": TURBO_DISABLE_AUTOCORR},
        {"config": "TURBO_DISABLE_HIGHER_MOMENTS", "value": TURBO_DISABLE_HIGHER_MOMENTS},
        {"config": "TURBO_DISABLE_ROLLING_VAR_QUANTILE", "value": TURBO_DISABLE_ROLLING_VAR_QUANTILE},
        {"config": "TURBO_DISABLE_VOLUME_ROLLING_CORR", "value": TURBO_DISABLE_VOLUME_ROLLING_CORR},
        {"config": "FEATURE_VALIDITY_FINITE_CHECK", "value": FEATURE_VALIDITY_FINITE_CHECK},

        {"config": "ENABLE_POST_AUDIT", "value": ENABLE_POST_AUDIT},
        {"config": "DROP_FEATURES_WITH_NULL_RATIO_ABOVE", "value": DROP_FEATURES_WITH_NULL_RATIO_ABOVE},
        {"config": "DROP_CONSTANT_FEATURES", "value": DROP_CONSTANT_FEATURES},
        {"config": "PARQUET_COMPRESSION", "value": PARQUET_COMPRESSION},
        {"config": "PARQUET_ROW_GROUP_SIZE", "value": PARQUET_ROW_GROUP_SIZE},
        {"config": "DOWNCAST_FLOATS_TO_FLOAT32", "value": DOWNCAST_FLOATS_TO_FLOAT32},
        {"config": "QUALITY_SUMMARY_FAST_FOR_LARGE_SERIES", "value": QUALITY_SUMMARY_FAST_FOR_LARGE_SERIES},
        {"config": "QUALITY_SUMMARY_FULL_NULLS_ROW_THRESHOLD", "value": QUALITY_SUMMARY_FULL_NULLS_ROW_THRESHOLD},
        {"config": "QUALITY_SUMMARY_SAMPLE_ROWS", "value": QUALITY_SUMMARY_SAMPLE_ROWS},
    ]

    return pd.DataFrame(rows)






def save_run_report(results: List[Dict[str, Any]], selected: List[Dict[str, Any]]) -> Path:
    ensure_dir(FEATURES_LOG_DIR)
    retry_plan = build_retry_plan_payload(results, selected)

    payload = {
        "schema_version": "ARCHANGEL_FEATURES_RUN_REPORT_3.3_CUDA_CONTROLLED_ROLLING",
        "generated_at": now_iso(),
        "run_id": RUN_ID,
        "script": SCRIPT_NAME,
        "timezone_policy": TIMEZONE_POLICY,

        "governance": {
            "input_contract_version": INPUT_CONTRACT_VERSION,
            "quality_gate_version": QUALITY_GATE_VERSION,
            "ml_readiness_version": ML_READINESS_VERSION,
            "data_quality_report_path": path_to_str(DATA_QUALITY_REPORT_PATH),
            "cost_model_path": path_to_str(COST_MODEL_PATH),
            "pre_feature_quality_gate_enabled": ENABLE_PRE_FEATURE_QUALITY_GATE,
            "input_contract_validation_enabled": ENABLE_INPUT_CONTRACT_VALIDATION,
            "time_grid_validation_enabled": ENABLE_TIME_GRID_VALIDATION,
            "feature_validity_flags_enabled": ENABLE_FEATURE_VALIDITY_FLAGS,
            "ml_ready_schema_flags_enabled": ENABLE_ML_READY_SCHEMA_FLAGS,
            "min_feature_non_null_ratio_for_valid_row": MIN_FEATURE_NON_NULL_RATIO_FOR_VALID_ROW,
        },

        "summary": {
            "series_selected": len(selected),
            "series_attempted": len(results),
            "series_ok": sum(1 for r in results if r.get("status") == "OK"),
            "series_error": sum(
                1 for r in results if str(r.get("status", "")).startswith("ERROR")
            ),



            "series_skipped": sum(1 for r in results if str(r.get("status", "")).startswith("SKIPPED")),

            "series_skipped_quality_fail": sum(
                1 for r in results if r.get("status") == "SKIPPED_DATA_QUALITY_FAIL"
            ),
            "series_input_contract_fail": sum(
                1 for r in results if r.get("status") == "ERROR_INPUT_CONTRACT_FAIL"
            ),
            "series_time_grid_fail": sum(
                1 for r in results if r.get("status") == "SKIPPED_TIME_GRID_FAIL"
            ),
            "series_recovered_by_retry": retry_plan.get("summary", {}).get("recovered_by_retry_count"),
            "series_retryable_after_run": retry_plan.get("summary", {}).get("retryable_series_count"),
            "series_non_retryable_failures_after_run": retry_plan.get("summary", {}).get("non_retryable_failure_count"),

            "total_valid_feature_rows": int(sum(r.get("valid_feature_rows", 0) or 0 for r in results)),
            "total_ml_eligible_rows": int(sum(r.get("ml_eligible_rows", 0) or 0 for r in results)),
            "feature_compute_backend": aggregate_feature_compute_backend(results),
        },

        "paths": {
            "incremental_audit_path": path_to_str(INCREMENTAL_AUDIT_PATH),
            "run_report_path": path_to_str(RUN_REPORT_PATH),
            "run_report_base_json_path": path_to_str(RUN_REPORT_BASE_JSON_PATH),
            "run_report_latest_path": path_to_str(RUN_REPORT_LATEST_PATH),
            "features_json_path": path_to_str(FEATURES_JSON_PATH),
            "features_excel_path": path_to_str(FEATURES_EXCEL_PATH),
            "features_parquet_dir": path_to_str(FEATURES_PARQUET_DIR),
            "python_environment_path": path_to_str(PYTHON_ENVIRONMENT_PATH),
            "retry_plan_json_path": path_to_str(RETRY_OUTPUT_JSON_PATH),
            "retry_plan_base_json_path": path_to_str(RETRY_OUTPUT_BASE_JSON_PATH),
            "retry_plan_latest_path": path_to_str(RETRY_OUTPUT_LATEST_PATH),
            "retry_plan_csv_path": path_to_str(RETRY_OUTPUT_CSV_PATH),
        },

        "performance_summary": {
            "feature_family_benchmark": aggregate_feature_family_benchmark(results),
            "phase_benchmark": aggregate_phase_timings(results),
            "feature_compute_backend": aggregate_feature_compute_backend(results),
            "max_memory_mb_after_feature_generation": max(
                [
                    float(r.get("memory_mb_after_feature_generation") or 0.0)
                    for r in results
                ] or [0.0]
            ),
            "max_memory_mb_after_write": max(
                [float(r.get("memory_mb_after_write") or 0.0) for r in results] or [0.0]
            ),
        },

        "config": {
            "feature_config": FEATURE_CONFIG,
            "performance": {
                "enable_parallel_processing": ENABLE_PARALLEL_PROCESSING,
                "feature_cuda_mode": FEATURE_CUDA_MODE,
                "feature_cuda_auto_min_gpu_memory_mb": FEATURE_CUDA_AUTO_MIN_GPU_MEMORY_MB,
                "feature_cuda_auto_min_rows": FEATURE_CUDA_AUTO_MIN_ROWS,
                "feature_cuda_max_workers": FEATURE_CUDA_MAX_WORKERS,
                "feature_cuda_accelerated_blocks": list(FEATURE_CUDA_ACCELERATED_BLOCKS),
                "feature_cuda_note": FEATURE_CUDA_NOTE,
                "metadata_output_mode": METADATA_OUTPUT_MODE,
                "read_optional_market_columns": READ_OPTIONAL_MARKET_COLUMNS,
                "include_optional_market_columns_in_output": INCLUDE_OPTIONAL_MARKET_COLUMNS_IN_OUTPUT,
                "feature_resource_profile": FEATURE_RESOURCE_PROFILE,
                "max_workers_features": MAX_WORKERS_FEATURES,
                "enable_memory_aware_workers": ENABLE_MEMORY_AWARE_WORKERS,
                "archangel_ram_cap_gb": ARCHANGEL_RAM_CAP_GB,
                "target_ram_used_gb": TARGET_RAM_USED_GB,
                "target_cpu_percent": TARGET_CPU_PERCENT,
                "target_free_ram_gb": get_target_free_ram_gb(),
                "effective_min_free_ram_gb": get_effective_min_free_ram_gb(),
                "min_free_ram_gb_to_start_batch": MIN_FREE_RAM_GB_TO_START_BATCH,
                "estimated_ram_gb_per_worker_standard": ESTIMATED_RAM_GB_PER_WORKER_STANDARD,
                "estimated_ram_gb_per_worker_turbo": ESTIMATED_RAM_GB_PER_WORKER_TURBO,
                "process_turbo_timeframes_first": PROCESS_TURBO_TIMEFRAMES_FIRST,
                "enable_timeframe_batch_execution": ENABLE_TIMEFRAME_BATCH_EXECUTION,
                "max_workers_turbo_batch": MAX_WORKERS_TURBO_BATCH,
                "enable_adaptive_workers": ENABLE_ADAPTIVE_WORKERS,
                "adaptive_workers_initial_turbo": ADAPTIVE_WORKERS_INITIAL_TURBO,
                "adaptive_workers_initial_standard": ADAPTIVE_WORKERS_INITIAL_STANDARD,
                "adaptive_scale_up_free_ram_gb": ADAPTIVE_SCALE_UP_FREE_RAM_GB,
                "adaptive_hold_free_ram_gb": ADAPTIVE_HOLD_FREE_RAM_GB,
                "adaptive_scale_down_free_ram_gb": ADAPTIVE_SCALE_DOWN_FREE_RAM_GB,
                "adaptive_scale_up_stable_heartbeats": ADAPTIVE_SCALE_UP_STABLE_HEARTBEATS,
                "adaptive_scale_up_worker_step": ADAPTIVE_SCALE_UP_WORKER_STEP,
                "adaptive_scale_down_worker_step": ADAPTIVE_SCALE_DOWN_WORKER_STEP,
                "adaptive_failure_cooldown_heartbeats": ADAPTIVE_FAILURE_COOLDOWN_HEARTBEATS,
                "adaptive_cpu_low_scale_up_threshold": ADAPTIVE_CPU_LOW_SCALE_UP_THRESHOLD,
                "enable_retry_failed_series": ENABLE_RETRY_FAILED_SERIES,
                "max_retry_attempts_per_series": MAX_RETRY_ATTEMPTS_PER_SERIES,
                "retry_failed_only_from_report": RETRY_FAILED_ONLY_FROM_REPORT,
                "enable_progress_heartbeat": ENABLE_PROGRESS_HEARTBEAT,
                "progress_heartbeat_seconds": PROGRESS_HEARTBEAT_SECONDS,
                "progress_show_active_limit": PROGRESS_SHOW_ACTIVE_LIMIT,
                "progress_show_recent_limit": PROGRESS_SHOW_RECENT_LIMIT,
                "feature_execution_profile": FEATURE_EXECUTION_PROFILE,
                "fast_first_row_threshold": FAST_FIRST_ROW_THRESHOLD,
                "fast_first_skip_families": sorted(FAST_FIRST_SKIP_FAMILIES),
                "ram_safe_row_threshold": RAM_SAFE_ROW_THRESHOLD,
                "ram_safe_skip_families_for_turbo": sorted(RAM_SAFE_SKIP_FAMILIES_FOR_TURBO),
                "ram_safe_skip_families_for_long_series": sorted(RAM_SAFE_SKIP_FAMILIES_FOR_LONG_SERIES),
                "enable_feature_block_construction": ENABLE_FEATURE_BLOCK_CONSTRUCTION,
                "feature_block_column_count": FEATURE_BLOCK_COLUMN_COUNT,
                "force_gc_each_feature_block": FORCE_GC_EACH_FEATURE_BLOCK,
                "enable_process_memory_telemetry": ENABLE_PROCESS_MEMORY_TELEMETRY,
                "quality_summary_fast_for_large_series": QUALITY_SUMMARY_FAST_FOR_LARGE_SERIES,
                "quality_summary_full_nulls_row_threshold": QUALITY_SUMMARY_FULL_NULLS_ROW_THRESHOLD,
                "quality_summary_sample_rows": QUALITY_SUMMARY_SAMPLE_ROWS,
                "parquet_compression": PARQUET_COMPRESSION,
                "parquet_row_group_size": PARQUET_ROW_GROUP_SIZE,
                "enable_timeframe_turbo_mode": ENABLE_TIMEFRAME_TURBO_MODE,
                "turbo_timeframes": sorted(TURBO_TIMEFRAMES),
                "turbo_disable_volume_rolling_corr": TURBO_DISABLE_VOLUME_ROLLING_CORR,
                "feature_validity_finite_check": FEATURE_VALIDITY_FINITE_CHECK,
            },
            "post_audit": {
                "enabled": ENABLE_POST_AUDIT,
                "max_gap_multiplier": POST_AUDIT_MAX_GAP_MULTIPLIER,
            },
        },

        "retry_plan": retry_plan,
        "results": results,
    }

    write_json_atomic(payload, RUN_REPORT_PATH)
    write_json_atomic(payload, RUN_REPORT_BASE_JSON_PATH)
    write_json_atomic(payload, RUN_REPORT_LATEST_PATH)
    return RUN_REPORT_PATH













def build_features_json_payload(
    mapa: Dict[str, Any],
    selected: List[Dict[str, Any]],
    results: List[Dict[str, Any]],
    run_report_path: Path,
    feature_catalog: List[Dict[str, Any]],
) -> Dict[str, Any]:

    ok_results = [r for r in results if r.get("status") == "OK"]
    error_results = [r for r in results if str(r.get("status", "")).startswith("ERROR")]

    skipped_results = [r for r in results if str(r.get("status", "")).startswith("SKIPPED")]
    discovered_universe = mapa.get("discovered_universe", {})

    families_summary = {}
    for item in feature_catalog:
        fam = item.get("family")
        if not fam:
            continue

        families_summary.setdefault(fam, {"feature_count": 0, "max_lookback": 0, "features": []})
        families_summary[fam]["feature_count"] += 1
        families_summary[fam]["features"].append(item.get("feature"))

        lb = item.get("lookback")
        if isinstance(lb, int):
            families_summary[fam]["max_lookback"] = max(families_summary[fam]["max_lookback"], lb)

    return {
        "schema_version": SCHEMA_VERSION,
        "system": {
            "name": SYSTEM_NAME,
            "version": SYSTEM_VERSION,
            "layer": "3_FEATURES",
            "script": SCRIPT_NAME,
            "run_id": RUN_ID,
            "generated_at": now_iso(),
        },
        "paths": {
            "root_dir": path_to_str(ROOT_DIR),
            "features_dir": path_to_str(FEATURES_DIR),
            "features_parquet_dir": path_to_str(FEATURES_PARQUET_DIR),
            "features_json_path": path_to_str(FEATURES_JSON_PATH),
            "features_excel_path": path_to_str(FEATURES_EXCEL_PATH),
            "mapa_ativos_path": path_to_str(MAPA_ATIVOS_PATH),
            "python_environment_path": path_to_str(PYTHON_ENVIRONMENT_PATH),
            "run_report_path": path_to_str(run_report_path),
            "incremental_audit_path": path_to_str(INCREMENTAL_AUDIT_PATH),
            "retry_plan_json_path": path_to_str(RETRY_OUTPUT_JSON_PATH),
            "retry_plan_csv_path": path_to_str(RETRY_OUTPUT_CSV_PATH),
        },
        "timezone_policy": TIMEZONE_POLICY,


        "governance_policy": {
            "input_contract_version": INPUT_CONTRACT_VERSION,
            "quality_gate_version": QUALITY_GATE_VERSION,
            "ml_readiness_version": ML_READINESS_VERSION,
            "data_quality_report_path": path_to_str(DATA_QUALITY_REPORT_PATH),
            "cost_model_path": path_to_str(COST_MODEL_PATH),
            "enable_pre_feature_quality_gate": ENABLE_PRE_FEATURE_QUALITY_GATE,
            "allow_warning_quality_to_process": ALLOW_WARNING_QUALITY_TO_PROCESS,
            "default_quality_status_if_missing": DEFAULT_QUALITY_STATUS_IF_MISSING,
            "block_on_quality_fail": BLOCK_ON_QUALITY_FAIL,
            "block_on_input_contract_fail": BLOCK_ON_INPUT_CONTRACT_FAIL,
            "block_on_time_grid_fail": BLOCK_ON_TIME_GRID_FAIL,
            "enable_input_contract_validation": ENABLE_INPUT_CONTRACT_VALIDATION,
            "enable_time_grid_validation": ENABLE_TIME_GRID_VALIDATION,
            "enable_timezone_consistency_check": ENABLE_TIMEZONE_CONSISTENCY_CHECK,
            "enable_feature_ready_timestamp": ENABLE_FEATURE_READY_TIMESTAMP,
            "enable_feature_validity_flags": ENABLE_FEATURE_VALIDITY_FLAGS,
            "enable_ml_ready_schema_flags": ENABLE_ML_READY_SCHEMA_FLAGS,
            "min_feature_non_null_ratio_for_valid_row": MIN_FEATURE_NON_NULL_RATIO_FOR_VALID_ROW,

        },



        "input_data_policy": {
            "dataset_kinds_allowed": sorted(DATASET_KIND_ALLOWED),
            "only_quality_ok": ONLY_QUALITY_OK,
            "min_rows": MIN_ROWS,
            "filters": {
                "assets": None if FILTER_ASSETS is None else sorted(FILTER_ASSETS),
                "sources": None if FILTER_SOURCES is None else sorted(FILTER_SOURCES),
                "timeframes": None if FILTER_TIMEFRAMES is None else sorted(FILTER_TIMEFRAMES),
            },
        },
        "performance_policy": {
            "enable_parallel_processing": ENABLE_PARALLEL_PROCESSING,
            "feature_cuda_mode": FEATURE_CUDA_MODE,
            "feature_cuda_auto_min_gpu_memory_mb": FEATURE_CUDA_AUTO_MIN_GPU_MEMORY_MB,
            "feature_cuda_auto_min_rows": FEATURE_CUDA_AUTO_MIN_ROWS,
            "feature_cuda_max_workers": FEATURE_CUDA_MAX_WORKERS,
            "feature_cuda_accelerated_blocks": list(FEATURE_CUDA_ACCELERATED_BLOCKS),
            "feature_cuda_note": FEATURE_CUDA_NOTE,
            "metadata_output_mode": METADATA_OUTPUT_MODE,
            "read_optional_market_columns": READ_OPTIONAL_MARKET_COLUMNS,
            "include_optional_market_columns_in_output": INCLUDE_OPTIONAL_MARKET_COLUMNS_IN_OUTPUT,
            "feature_resource_profile": FEATURE_RESOURCE_PROFILE,
            "max_workers_features": MAX_WORKERS_FEATURES,
            "enable_memory_aware_workers": ENABLE_MEMORY_AWARE_WORKERS,
            "archangel_ram_cap_gb": ARCHANGEL_RAM_CAP_GB,
            "target_ram_used_gb": TARGET_RAM_USED_GB,
            "target_cpu_percent": TARGET_CPU_PERCENT,
            "target_free_ram_gb": get_target_free_ram_gb(),
            "effective_min_free_ram_gb": get_effective_min_free_ram_gb(),
            "min_free_ram_gb_to_start_batch": MIN_FREE_RAM_GB_TO_START_BATCH,
            "estimated_ram_gb_per_worker_standard": ESTIMATED_RAM_GB_PER_WORKER_STANDARD,
            "estimated_ram_gb_per_worker_turbo": ESTIMATED_RAM_GB_PER_WORKER_TURBO,
            "process_turbo_timeframes_first": PROCESS_TURBO_TIMEFRAMES_FIRST,
            "enable_timeframe_batch_execution": ENABLE_TIMEFRAME_BATCH_EXECUTION,
            "max_workers_turbo_batch": MAX_WORKERS_TURBO_BATCH,
            "enable_adaptive_workers": ENABLE_ADAPTIVE_WORKERS,
            "adaptive_workers_initial_turbo": ADAPTIVE_WORKERS_INITIAL_TURBO,
            "adaptive_workers_initial_standard": ADAPTIVE_WORKERS_INITIAL_STANDARD,
            "adaptive_scale_up_free_ram_gb": ADAPTIVE_SCALE_UP_FREE_RAM_GB,
            "adaptive_hold_free_ram_gb": ADAPTIVE_HOLD_FREE_RAM_GB,
            "adaptive_scale_down_free_ram_gb": ADAPTIVE_SCALE_DOWN_FREE_RAM_GB,
            "adaptive_scale_up_stable_heartbeats": ADAPTIVE_SCALE_UP_STABLE_HEARTBEATS,
            "adaptive_scale_up_worker_step": ADAPTIVE_SCALE_UP_WORKER_STEP,
            "adaptive_scale_down_worker_step": ADAPTIVE_SCALE_DOWN_WORKER_STEP,
            "adaptive_failure_cooldown_heartbeats": ADAPTIVE_FAILURE_COOLDOWN_HEARTBEATS,
            "adaptive_cpu_low_scale_up_threshold": ADAPTIVE_CPU_LOW_SCALE_UP_THRESHOLD,
            "enable_retry_failed_series": ENABLE_RETRY_FAILED_SERIES,
            "max_retry_attempts_per_series": MAX_RETRY_ATTEMPTS_PER_SERIES,
            "retry_failed_only_from_report": RETRY_FAILED_ONLY_FROM_REPORT,
            "enable_progress_heartbeat": ENABLE_PROGRESS_HEARTBEAT,
            "progress_heartbeat_seconds": PROGRESS_HEARTBEAT_SECONDS,
            "progress_show_active_limit": PROGRESS_SHOW_ACTIVE_LIMIT,
            "progress_show_recent_limit": PROGRESS_SHOW_RECENT_LIMIT,
            "feature_execution_profile": FEATURE_EXECUTION_PROFILE,
            "ram_safe_row_threshold": RAM_SAFE_ROW_THRESHOLD,
            "ram_safe_skip_families_for_turbo": sorted(RAM_SAFE_SKIP_FAMILIES_FOR_TURBO),
            "ram_safe_skip_families_for_long_series": sorted(RAM_SAFE_SKIP_FAMILIES_FOR_LONG_SERIES),
            "enable_feature_block_construction": ENABLE_FEATURE_BLOCK_CONSTRUCTION,
            "feature_block_column_count": FEATURE_BLOCK_COLUMN_COUNT,
            "enable_process_memory_telemetry": ENABLE_PROCESS_MEMORY_TELEMETRY,
            "quality_summary_fast_for_large_series": QUALITY_SUMMARY_FAST_FOR_LARGE_SERIES,
            "quality_summary_full_nulls_row_threshold": QUALITY_SUMMARY_FULL_NULLS_ROW_THRESHOLD,
            "quality_summary_sample_rows": QUALITY_SUMMARY_SAMPLE_ROWS,
            "enable_timeframe_turbo_mode": ENABLE_TIMEFRAME_TURBO_MODE,
            "turbo_timeframes": sorted(TURBO_TIMEFRAMES),
            "turbo_disable_volume_rolling_corr": TURBO_DISABLE_VOLUME_ROLLING_CORR,
            "feature_validity_finite_check": FEATURE_VALIDITY_FINITE_CHECK,
            "write_full_feature_details_to_excel": WRITE_FULL_FEATURE_DETAILS_TO_EXCEL,
            "save_feature_quality_details": SAVE_FEATURE_QUALITY_DETAILS,
        },
        "post_audit_policy": {
            "enabled": ENABLE_POST_AUDIT,
            "max_gap_multiplier": POST_AUDIT_MAX_GAP_MULTIPLIER,
            "purpose": "Garantir sincronização entre OHLCV e features no mesmo grid temporal.",
        },
        "performance_summary": {
            "feature_family_benchmark": aggregate_feature_family_benchmark(ok_results),
            "phase_benchmark": aggregate_phase_timings(ok_results),
            "feature_compute_backend": aggregate_feature_compute_backend(ok_results),
        },
        "discovered_universe_reference": {
            "assets_discovered": discovered_universe.get("assets_discovered"),
            "symbols_discovered": discovered_universe.get("symbols_discovered"),
            "sources_discovered": discovered_universe.get("sources_discovered"),
            "timeframes_discovered": discovered_universe.get("timeframes_discovered"),
            "dataset_kinds_discovered": discovered_universe.get("dataset_kinds_discovered"),
        },
        "feature_config": FEATURE_CONFIG,
        "feature_families_summary": families_summary,
        "feature_catalog": feature_catalog,

        "recommended_next_layer": {
            "datasets_ml": (
                "Próxima etapa prioritária: criar 05_MONTA_DATASETS_ML.py para juntar "
                "features, labels, custos, filtros de qualidade e cross-asset regressors."
            ),
            "cross_asset_features": (
                "Cross-asset regressors devem ser montados no 05_MONTA_DATASETS_ML.py "
                "via timestamp_utc_ms, não dentro da feature store individual."
            ),
            "cost_model": (
                "Usar 00_COST_MODEL.json em 05_MONTA_DATASETS_ML.py e 07_BACKTEST_PORTFOLIO.py "
                "para aplicar fees, slippage, funding e impacto."
            ),
            "labels": (
                "Criar módulo separado de labels. Labels devem usar apenas futuro posterior "
                "ao timestamp_utc_ms da barra de decisão e exigir purging/embargo no walk-forward."
            ),
            "ml_feature_selection_rule": (
                "Modelos só podem consumir colunas feat_*, xasset_* e regime_*. "
                "Nunca usar label_*, meta_*, DateTime ou timestamp_utc_ms como features diretas."
            ),
        },





        "summary": {
            "series_selected": len(selected),
            "series_attempted": len(results),
            "series_ok": len(ok_results),
            "series_error": len(error_results),
            "series_skipped": len(skipped_results),
            "total_output_rows": int(sum(r.get("output_rows", 0) or 0 for r in ok_results)),
            "total_feature_files": len(ok_results),
            "feature_catalog_count": len(feature_catalog),
            "feature_families_count": len(families_summary),
            "series_skipped_quality_fail": sum(
                1 for r in results if r.get("status") == "SKIPPED_DATA_QUALITY_FAIL"
            ),
            "series_input_contract_fail": sum(
                1 for r in results if r.get("status") == "ERROR_INPUT_CONTRACT_FAIL"
            ),
            "series_time_grid_fail": sum(
                1 for r in results if r.get("status") == "SKIPPED_TIME_GRID_FAIL"
            ),
            "total_valid_feature_rows": int(sum(r.get("valid_feature_rows", 0) or 0 for r in ok_results)),
            "total_ml_eligible_rows": int(sum(r.get("ml_eligible_rows", 0) or 0 for r in ok_results)),
            "max_memory_mb_after_feature_generation": max(
                [
                    float(r.get("memory_mb_after_feature_generation") or 0.0)
                    for r in ok_results
                ] or [0.0]
            ),
            "max_memory_mb_after_write": max(
                [float(r.get("memory_mb_after_write") or 0.0) for r in ok_results] or [0.0]
            ),
        },
        "series_outputs": [
            {
                "status": r.get("status"),
                "series_id": r.get("series_id"),
                "asset": r.get("asset"),
                "symbol": r.get("symbol"),
                "source": r.get("source"),
                "timeframe": r.get("timeframe"),
                "turbo_mode_used": r.get("turbo_mode_used"),
                "timeframe_seconds": r.get("timeframe_seconds"),
                "input_path": r.get("input_path"),
                "output_path": r.get("output_path"),
                "output_rows": r.get("output_rows"),
                "output_columns": r.get("output_columns"),
                "required_min_rows": r.get("required_min_rows"),
                "feature_columns_count": r.get("feature_columns_count"),
                "feature_execution_profile": r.get("feature_execution_profile"),
                "feature_family_slowest": r.get("feature_family_slowest"),
                "feature_family_benchmark": r.get("feature_family_benchmark"),
                "phase_slowest": r.get("phase_slowest"),
                "phase_timings": r.get("phase_timings"),
                "quality_summary": r.get("quality_summary"),
                "quality_summary_mode": safe_get_nested(r, ["quality_summary", "quality_summary_mode"]),
                "post_audit": r.get("post_audit"),
                "elapsed_seconds": r.get("elapsed_seconds"),
                "batch_name": r.get("batch_name"),
                "memory_mb_start": r.get("memory_mb_start"),
                "memory_mb_after_read": r.get("memory_mb_after_read"),
                "memory_mb_after_prepare": r.get("memory_mb_after_prepare"),
                "memory_mb_after_feature_generation": r.get("memory_mb_after_feature_generation"),
                "memory_mb_after_downcast": r.get("memory_mb_after_downcast"),
                "memory_mb_after_write": r.get("memory_mb_after_write"),
                "error": r.get("error"),
                "quality_status_normalized": r.get("quality_status_normalized"),
                "input_contract_report": r.get("input_contract_report"),
                "time_grid_report": r.get("time_grid_report"),
                "max_lookback_bars": r.get("max_lookback_bars"),
                "valid_feature_rows": r.get("valid_feature_rows"),
                "warmup_rows": r.get("warmup_rows"),
                "ml_eligible_rows": r.get("ml_eligible_rows"),
                "first_valid_timestamp_utc_ms": r.get("first_valid_timestamp_utc_ms"),
            }

            for r in results
        ],
        "anti_leakage_policy": {
            "uses_future_data": False,
            "labels_created_here": False,
            "warning": "Este script não cria labels e não usa dados futuros.",
        },
    }






# =============================================================================
# 13. EXCEL UX
# =============================================================================

def style_excel_workbook(writer: pd.ExcelWriter) -> None:
    try:
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.formatting.rule import CellIsRule
        from openpyxl.utils import get_column_letter

        wb = writer.book

        dark_blue = "17365D"
        medium_blue = "1F4E78"
        light_blue = "D9EAF7"
        light_green = "E2F0D9"
        light_red = "F4CCCC"
        light_orange = "FCE4D6"
        light_yellow = "FFF2CC"
        white = "FFFFFF"
        gray = "F3F6F9"

        thin = Side(style="thin", color="D9E2F3")

        for ws in wb.worksheets:
            ws.sheet_view.showGridLines = False

            if ws.max_row >= 1:
                for cell in ws[1]:
                    cell.fill = PatternFill("solid", fgColor=dark_blue)
                    cell.font = Font(color=white, bold=True)
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    cell.border = Border(bottom=thin)

                ws.freeze_panes = "A2"
                ws.auto_filter.ref = ws.dimensions

            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(vertical="top", wrap_text=False)
                    cell.border = Border(bottom=thin)

            for idx, col_cells in enumerate(ws.columns, start=1):
                max_len = 10
                for cell in list(col_cells)[:300]:
                    value = "" if cell.value is None else str(cell.value)
                    max_len = max(max_len, len(value))
                ws.column_dimensions[get_column_letter(idx)].width = min(max_len + 2, 55)

            if ws.title == "DASHBOARD":
                ws.column_dimensions["A"].width = 18
                ws.column_dimensions["B"].width = 28
                ws.column_dimensions["C"].width = 28
                ws.column_dimensions["D"].width = 65

            if ws.title in {"SERIES", "POST_AUDIT"}:
                header = [c.value for c in ws[1]]
                for col_idx, name in enumerate(header, start=1):
                    col_letter = get_column_letter(col_idx)

                    if name in {"status", "audit_status"}:
                        ws.conditional_formatting.add(
                            f"{col_letter}2:{col_letter}{ws.max_row}",
                            CellIsRule(operator="equal", formula=['"OK"'], fill=PatternFill("solid", fgColor=light_green)),
                        )
                        ws.conditional_formatting.add(
                            f"{col_letter}2:{col_letter}{ws.max_row}",
                            CellIsRule(operator="equal", formula=['"PASS"'], fill=PatternFill("solid", fgColor=light_green)),
                        )
                        ws.conditional_formatting.add(
                            f"{col_letter}2:{col_letter}{ws.max_row}",
                            CellIsRule(operator="equal", formula=['"WARNING"'], fill=PatternFill("solid", fgColor=light_yellow)),
                        )
                        ws.conditional_formatting.add(
                            f"{col_letter}2:{col_letter}{ws.max_row}",
                            CellIsRule(operator="equal", formula=['"ERROR"'], fill=PatternFill("solid", fgColor=light_red)),
                        )
                        ws.conditional_formatting.add(
                            f"{col_letter}2:{col_letter}{ws.max_row}",
                            CellIsRule(operator="equal", formula=['"FAIL"'], fill=PatternFill("solid", fgColor=light_red)),
                        )

    except Exception:
        pass


def save_features_excel_report(
    selected: List[Dict[str, Any]],
    results: List[Dict[str, Any]],
    feature_catalog: List[Dict[str, Any]],
) -> Path:

    ensure_dir(BASE_EXCEL_DIR)

    dashboard_df = build_dashboard_df(selected, results, feature_catalog)
    series_df = build_series_excel_df(results)
    families_df = build_feature_families_df(feature_catalog)
    catalog_df = build_feature_catalog_excel_df(feature_catalog)
    post_audit_df = build_post_audit_df(results)
    errors_df = build_errors_df(results)
    config_df = build_config_df()
    retry_plan = build_retry_plan_payload(results, selected)
    retry_df = pd.DataFrame(
        retry_plan.get("retryable_series", [])
        + retry_plan.get("non_retryable_failures", [])
    )

    with pd.ExcelWriter(FEATURES_EXCEL_PATH, engine="openpyxl") as writer:
        dashboard_df.to_excel(writer, sheet_name="DASHBOARD", index=False)
        series_df.to_excel(writer, sheet_name="SERIES", index=False)
        families_df.to_excel(writer, sheet_name="FEATURE_FAMILIES", index=False)
        catalog_df.to_excel(writer, sheet_name="FEATURE_CATALOG", index=False)
        post_audit_df.to_excel(writer, sheet_name="POST_AUDIT", index=False)
        errors_df.to_excel(writer, sheet_name="ERRORS", index=False)
        retry_df.to_excel(writer, sheet_name="RETRY_PLAN", index=False)
        config_df.to_excel(writer, sheet_name="CONFIG", index=False)

        style_excel_workbook(writer)

    return FEATURES_EXCEL_PATH


def configure_console_output() -> None:
    try:
        sys.stdout.reconfigure(line_buffering=True)
        sys.stderr.reconfigure(line_buffering=True)
    except Exception:
        pass


def format_duration(seconds: Optional[float]) -> str:
    if seconds is None:
        return "?"

    try:
        total = max(0, int(seconds))
    except Exception:
        return "?"

    hours, remainder = divmod(total, 3600)
    minutes, secs = divmod(remainder, 60)

    if hours:
        return f"{hours:02d}:{minutes:02d}:{secs:02d}"

    return f"{minutes:02d}:{secs:02d}"


def item_progress_label(item: Dict[str, Any]) -> str:
    info = item.get("info", {})
    if not isinstance(info, dict):
        info = {}

    source = info.get("source") or item.get("source") or "?"
    asset = info.get("asset") or item.get("asset") or "?"
    timeframe = info.get("timeframe") or item.get("timeframe") or "?"
    rows = info.get("rows") or item.get("rows")

    suffix = f" | rows={rows}" if rows else ""
    return f"{source} | {asset} | {timeframe}{suffix}"


def build_worker_error_result(
    item: Dict[str, Any],
    exc: Exception,
    tb: str,
) -> Dict[str, Any]:
    info = item.get("info", {}) if isinstance(item.get("info"), dict) else {}
    return {
        "run_id": RUN_ID,
        "series_id": info.get("series_id"),
        "asset": info.get("asset"),
        "symbol": info.get("symbol"),
        "source": info.get("source"),
        "timeframe": info.get("timeframe"),
        "input_path": info.get("absolute_path"),
        "status": "ERROR",
        "error": str(exc),
        "traceback": tb,
    }


def print_batch_heartbeat(
    batch_name: str,
    batch_total: int,
    results: List[Dict[str, Any]],
    total_selected: int,
    batch_started_at: float,
    active_futures: Dict[Any, Dict[str, Any]],
    completed_in_batch: int,
    target_workers: Optional[int] = None,
    max_workers: Optional[int] = None,
    pending_count: Optional[int] = None,
) -> None:
    if not ENABLE_PROGRESS_HEARTBEAT:
        return

    elapsed = time.time() - batch_started_at
    remaining_in_batch = max(0, batch_total - completed_in_batch)
    avg_seconds = elapsed / completed_in_batch if completed_in_batch else None
    eta_seconds = avg_seconds * remaining_in_batch if avg_seconds is not None else None
    available_ram_gb = get_available_ram_gb()
    used_ram_gb = get_used_ram_gb()
    cpu_percent = get_system_cpu_percent()

    ram_label = "?" if available_ram_gb is None else f"{available_ram_gb:.1f}"
    used_ram_label = "?" if used_ram_gb is None else f"{used_ram_gb:.1f}"
    cpu_label = "?" if cpu_percent is None else f"{cpu_percent:.1f}"
    avg_label = "?" if avg_seconds is None else format_duration(avg_seconds)

    print(
        f"[MONITOR] batch={batch_name} | "
        f"batch={completed_in_batch}/{batch_total} | "
        f"global={len(results)}/{total_selected} | "
        f"ativos={len(active_futures)} | "
        f"alvo_workers={target_workers if target_workers is not None else '?'} | "
        f"max_workers={max_workers if max_workers is not None else '?'} | "
        f"pendentes={pending_count if pending_count is not None else '?'} | "
        f"elapsed={format_duration(elapsed)} | "
        f"media={avg_label} | "
        f"eta_batch={format_duration(eta_seconds)} | "
        f"cpu={cpu_label}%/{TARGET_CPU_PERCENT:.0f}% | "
        f"ram_usada_gb={used_ram_label}/{TARGET_RAM_USED_GB:.0f} | "
        f"ram_livre_gb={ram_label} | "
        f"ram_livre_alvo_gb={get_target_free_ram_gb()}",
        flush=True,
    )

    if active_futures:
        print("[MONITOR] Em processamento:", flush=True)
        for active in list(active_futures.values())[:PROGRESS_SHOW_ACTIVE_LIMIT]:
            active_elapsed = time.time() - float(active.get("started_at") or time.time())
            print(
                f"  - idx={active.get('idx')} | {item_progress_label(active.get('item', {}))} | "
                f"rodando_ha={format_duration(active_elapsed)}",
                flush=True,
            )

        hidden = len(active_futures) - PROGRESS_SHOW_ACTIVE_LIMIT
        if hidden > 0:
            print(f"  - ... mais {hidden} worker(s) ativo(s)", flush=True)

    recent_results = results[-PROGRESS_SHOW_RECENT_LIMIT:]
    if recent_results:
        print("[MONITOR] Ultimas concluídas:", flush=True)
        for result in recent_results:
            elapsed_result = result.get("elapsed_seconds")
            elapsed_label = (
                format_duration(float(elapsed_result))
                if elapsed_result is not None
                else "?"
            )
            print(
                f"  - idx={result.get('progress_index')} | status={result.get('status')} | "
                f"{result.get('source')} | {result.get('asset')} | {result.get('timeframe')} | "
                f"features={result.get('feature_columns_count')} | tempo={elapsed_label}",
                flush=True,
            )


def process_batch(
    batch_name: str,
    batch: List[Dict[str, Any]],
    results: List[Dict[str, Any]],
    total_selected: int,
    progress_offset: int,
) -> int:
    if not batch:
        return progress_offset

    workers = resolve_worker_count(batch, batch_name)
    max_workers = resolve_max_worker_count(batch, batch_name)
    effective_min_free_ram_gb = get_effective_min_free_ram_gb()
    target_free_ram_gb = get_target_free_ram_gb()
    estimated_ram_gb_per_worker = estimate_ram_gb_per_worker_for_batch(batch)
    available_ram_gb = get_available_ram_gb()
    print(
        f"[BATCH] {batch_name} | séries={len(batch)} | workers_iniciais={workers} | "
        f"workers_max={max_workers} | ram_cap_gb={ARCHANGEL_RAM_CAP_GB} | "
        f"ram_reserva_gb={effective_min_free_ram_gb} | "
        f"ram_livre_alvo_gb={target_free_ram_gb} | "
        f"ram_worker_estimado_gb={estimated_ram_gb_per_worker} | "
        f"ram_disponivel_gb={available_ram_gb}",
        flush=True,
    )

    append_incremental_audit({
        "event": "BATCH_STARTED",
        "batch_name": batch_name,
        "batch_series_count": len(batch),
        "workers": workers,
        "max_workers": max_workers,
        "adaptive_workers_enabled": ENABLE_ADAPTIVE_WORKERS,
        "ram_cap_gb": ARCHANGEL_RAM_CAP_GB,
        "effective_min_free_ram_gb": effective_min_free_ram_gb,
        "target_free_ram_gb": target_free_ram_gb,
        "target_ram_used_gb": TARGET_RAM_USED_GB,
        "target_cpu_percent": TARGET_CPU_PERCENT,
        "estimated_ram_gb_per_worker": estimated_ram_gb_per_worker,
        "available_ram_gb": available_ram_gb,
        "progress_offset": progress_offset,
        "selected_series_count": total_selected,
    })

    batch_started_at = time.time()

    if ENABLE_PARALLEL_PROCESSING and len(batch) > 1 and workers > 1:
        with ProcessPoolExecutor(max_workers=max_workers) as executor:
            pending_items: List[Dict[str, Any]] = [
                {
                    "local_idx": local_idx,
                    "idx": progress_offset + local_idx,
                    "item": item,
                    "attempt": int(item.get("_retry_attempt", 0) or 0),
                }
                for local_idx, item in enumerate(batch, start=1)
            ]
            pending_pos = 0
            active_futures: Dict[Any, Dict[str, Any]] = {}
            completed_in_batch = 0
            target_workers = max(1, min(workers, max_workers))
            last_heartbeat_at = time.time()
            batch_start_available_ram_gb = available_ram_gb
            high_ram_heartbeats = 0
            failure_cooldown = 0

            def pending_count() -> int:
                return max(0, len(pending_items) - pending_pos)

            def can_submit_more() -> bool:
                if pending_count() <= 0:
                    return False
                if len(active_futures) >= target_workers:
                    return False

                current_ram_gb = get_available_ram_gb()
                if current_ram_gb is None:
                    return True
                if not active_futures:
                    return True
                return float(current_ram_gb) >= ADAPTIVE_HOLD_FREE_RAM_GB

            def submit_next_item() -> bool:
                nonlocal pending_pos

                if not can_submit_more():
                    return False

                try:
                    pending = pending_items[pending_pos]
                except IndexError:
                    return False

                pending_pos += 1
                idx = int(pending["idx"])
                item = pending["item"]
                attempt = int(pending.get("attempt") or 0)

                try:
                    future = executor.submit(process_one_series, item)
                except Exception as exc:
                    result = build_worker_error_result(item, exc, traceback.format_exc())
                    result["batch_name"] = batch_name
                    result["progress_index"] = idx
                    result["progress_total"] = total_selected
                    result["retry_attempt"] = attempt
                    result["error_classification"] = classify_result_error(result)
                    results.append(result)
                    append_incremental_audit(result)
                    return False

                active_futures[future] = {
                    "idx": idx,
                    "local_idx": pending.get("local_idx"),
                    "item": item,
                    "attempt": attempt,
                    "started_at": time.time(),
                }
                print(
                    f"[START] {idx}/{total_selected} | batch={batch_name} | "
                    f"retry={attempt} | ativos={len(active_futures)}/{target_workers} | "
                    f"{item_progress_label(item)}",
                    flush=True,
                )
                return True

            def fill_worker_slots() -> None:
                while can_submit_more():
                    if not submit_next_item():
                        break

            def update_target_workers(force_print: bool = False) -> None:
                nonlocal target_workers, high_ram_heartbeats, failure_cooldown

                if not ENABLE_ADAPTIVE_WORKERS:
                    return

                current_ram_gb = get_available_ram_gb()
                current_cpu_percent = get_system_cpu_percent()
                previous_target = target_workers

                observed_worker_ram_gb = estimated_ram_gb_per_worker
                if (
                    batch_start_available_ram_gb is not None
                    and current_ram_gb is not None
                    and active_futures
                ):
                    batch_ram_delta = max(
                        0.0,
                        float(batch_start_available_ram_gb) - float(current_ram_gb),
                    )
                    if batch_ram_delta > 1.0:
                        observed_worker_ram_gb = max(
                            1.0,
                            batch_ram_delta / max(1, len(active_futures)),
                        )

                ram_safe_workers = estimate_ram_safe_worker_count(
                    current_ram_gb,
                    min(estimated_ram_gb_per_worker, observed_worker_ram_gb),
                )
                desired_by_ram = max(1, min(max_workers, ram_safe_workers))

                if failure_cooldown > 0:
                    failure_cooldown -= 1
                    high_ram_heartbeats = 0
                elif current_ram_gb is not None and current_ram_gb <= ADAPTIVE_SCALE_DOWN_FREE_RAM_GB:
                    target_workers = max(1, target_workers - ADAPTIVE_SCALE_DOWN_WORKER_STEP)
                    high_ram_heartbeats = 0
                elif current_ram_gb is not None and current_ram_gb < ADAPTIVE_HOLD_FREE_RAM_GB:
                    target_workers = max(1, min(target_workers, desired_by_ram))
                    high_ram_heartbeats = 0
                elif (
                    current_ram_gb is not None
                    and current_ram_gb > target_free_ram_gb
                    and (
                        current_ram_gb >= ADAPTIVE_SCALE_UP_FREE_RAM_GB
                        or (
                            current_cpu_percent is not None
                            and current_cpu_percent < ADAPTIVE_CPU_LOW_SCALE_UP_THRESHOLD
                        )
                    )
                ):
                    high_ram_heartbeats += 1
                    if high_ram_heartbeats >= ADAPTIVE_SCALE_UP_STABLE_HEARTBEATS:
                        target_workers = min(
                            max_workers,
                            desired_by_ram,
                            target_workers + max(1, ADAPTIVE_SCALE_UP_WORKER_STEP),
                        )
                        high_ram_heartbeats = 0
                else:
                    high_ram_heartbeats = 0

                if force_print or previous_target != target_workers:
                    append_incremental_audit({
                        "event": "ADAPTIVE_WORKERS_STATUS",
                        "batch_name": batch_name,
                        "target_workers_previous": previous_target,
                        "target_workers": target_workers,
                        "max_workers": max_workers,
                        "active_workers": len(active_futures),
                        "pending_count": pending_count(),
                        "available_ram_gb": current_ram_gb,
                        "cpu_percent": current_cpu_percent,
                        "desired_workers_by_ram": desired_by_ram,
                        "estimated_ram_gb_per_worker": estimated_ram_gb_per_worker,
                        "observed_worker_ram_gb": round(observed_worker_ram_gb, 3),
                        "target_free_ram_gb": target_free_ram_gb,
                        "target_ram_used_gb": TARGET_RAM_USED_GB,
                        "target_cpu_percent": TARGET_CPU_PERCENT,
                        "scale_up_stable_heartbeats": high_ram_heartbeats,
                        "failure_cooldown": failure_cooldown,
                    })
                    if previous_target != target_workers:
                        print(
                            f"[ADAPTIVE] batch={batch_name} | workers_alvo "
                            f"{previous_target} -> {target_workers} | "
                            f"cpu={current_cpu_percent}% | "
                            f"ram_livre_gb={current_ram_gb} | "
                            f"ram_worker_obs_gb={round(observed_worker_ram_gb, 2)} | "
                            f"pendentes={pending_count()}",
                            flush=True,
                        )

            fill_worker_slots()

            print_batch_heartbeat(
                batch_name=batch_name,
                batch_total=len(pending_items),
                results=results,
                total_selected=total_selected,
                batch_started_at=batch_started_at,
                active_futures=active_futures,
                completed_in_batch=completed_in_batch,
                target_workers=target_workers,
                max_workers=max_workers,
                pending_count=pending_count(),
            )

            while active_futures or pending_count() > 0:
                fill_worker_slots()

                if not active_futures:
                    update_target_workers()
                    time.sleep(min(5.0, PROGRESS_HEARTBEAT_SECONDS))
                    fill_worker_slots()
                    continue

                done_futures, _ = wait(
                    list(active_futures.keys()),
                    timeout=PROGRESS_HEARTBEAT_SECONDS,
                    return_when=FIRST_COMPLETED,
                )

                if not done_futures:
                    update_target_workers()
                    print_batch_heartbeat(
                        batch_name=batch_name,
                        batch_total=len(pending_items),
                        results=results,
                        total_selected=total_selected,
                        batch_started_at=batch_started_at,
                        active_futures=active_futures,
                        completed_in_batch=completed_in_batch,
                        target_workers=target_workers,
                        max_workers=max_workers,
                        pending_count=pending_count(),
                    )
                    last_heartbeat_at = time.time()
                    continue

                for future in done_futures:
                    active = active_futures.pop(future)
                    idx = int(active["idx"])
                    item = active.get("item", {})
                    attempt = int(active.get("attempt") or 0)

                    try:
                        result = future.result()
                    except Exception as exc:
                        result = build_worker_error_result(item, exc, traceback.format_exc())

                    result["batch_name"] = batch_name
                    result["progress_index"] = idx
                    result["progress_total"] = total_selected
                    result["retry_attempt"] = attempt
                    result["error_classification"] = classify_result_error(result)

                    results.append(result)
                    append_incremental_audit(result)
                    completed_in_batch += 1

                    if result.get("status") == "OK":
                        for previous in results:
                            if (
                                previous is not result
                                and previous.get("series_id") == result.get("series_id")
                                and str(previous.get("status", "")).startswith("ERROR")
                            ):
                                previous["recovered_by_retry"] = True
                                previous["retry_final_status"] = "OK"

                    if (
                        ENABLE_RETRY_FAILED_SERIES
                        and is_retryable_result(result)
                        and attempt < MAX_RETRY_ATTEMPTS_PER_SERIES
                    ):
                        next_attempt = attempt + 1
                        retry_item = dict(item)
                        retry_item["_retry_attempt"] = next_attempt
                        pending_items.append({
                            "local_idx": active.get("local_idx"),
                            "idx": idx,
                            "item": retry_item,
                            "attempt": next_attempt,
                        })
                        result["retry_scheduled"] = True
                        result["retry_next_attempt"] = next_attempt
                        failure_cooldown = ADAPTIVE_FAILURE_COOLDOWN_HEARTBEATS
                        target_workers = max(1, target_workers - 1)
                        append_incremental_audit({
                            "event": "SERIES_RETRY_SCHEDULED",
                            "batch_name": batch_name,
                            "series_id": result.get("series_id"),
                            "asset": result.get("asset"),
                            "symbol": result.get("symbol"),
                            "source": result.get("source"),
                            "timeframe": result.get("timeframe"),
                            "retry_attempt": next_attempt,
                            "target_workers": target_workers,
                            "error_classification": result.get("error_classification"),
                        })
                        print(
                            f"[RETRY] agendado | idx={idx} | tentativa={next_attempt} | "
                            f"workers_alvo={target_workers} | {result.get('source')} | "
                            f"{result.get('asset')} | {result.get('timeframe')} | "
                            f"erro={result.get('error_classification')}",
                            flush=True,
                        )
                    elif is_retryable_result(result):
                        result["retry_scheduled"] = False

                    err_preview = str(result.get("error") or "")
                    if len(err_preview) > 180:
                        err_preview = err_preview[:180] + "..."

                    elapsed_label = (
                        format_duration(float(result.get("elapsed_seconds")))
                        if result.get("elapsed_seconds") is not None
                        else "?"
                    )

                    print(
                        f"[PROGRESSO] {len(results)}/{total_selected} | "
                        f"batch={batch_name} | idx={idx} | status={result.get('status')} | "
                        f"audit={(result.get('post_audit') or {}).get('audit_status')} | "
                        f"q={result.get('quality_status_normalized')} | "
                        f"mlq={result.get('quality_ml_status')} | "
                        f"{result.get('source')} | {result.get('asset')} | {result.get('timeframe')} | "
                        f"retry={attempt} | "
                        f"turbo={result.get('turbo_mode_used')} | "
                        f"features={result.get('feature_columns_count')} | "
                        f"mem={result.get('memory_mb_after_write')}MB | "
                        f"tempo={elapsed_label}"
                        + (f" | error={err_preview}" if err_preview else ""),
                        flush=True,
                    )

                update_target_workers()
                fill_worker_slots()

                if time.time() - last_heartbeat_at >= PROGRESS_HEARTBEAT_SECONDS:
                    print_batch_heartbeat(
                        batch_name=batch_name,
                        batch_total=len(pending_items),
                        results=results,
                        total_selected=total_selected,
                        batch_started_at=batch_started_at,
                        active_futures=active_futures,
                        completed_in_batch=completed_in_batch,
                        target_workers=target_workers,
                        max_workers=max_workers,
                        pending_count=pending_count(),
                    )
                    last_heartbeat_at = time.time()
    else:
        pending_items: List[Dict[str, Any]] = [
            {
                "local_idx": local_idx,
                "idx": progress_offset + local_idx,
                "item": item,
                "attempt": int(item.get("_retry_attempt", 0) or 0),
            }
            for local_idx, item in enumerate(batch, start=1)
        ]
        pending_pos = 0

        while pending_pos < len(pending_items):
            pending = pending_items[pending_pos]
            pending_pos += 1

            idx = int(pending["idx"])
            item = pending["item"]
            attempt = int(pending.get("attempt") or 0)
            print(
                f"\n[START] {idx}/{total_selected} | batch={batch_name} | "
                f"retry={attempt} | {item_progress_label(item)}",
                flush=True,
            )
            result = process_one_series(item)
            result["batch_name"] = batch_name
            result["progress_index"] = idx
            result["progress_total"] = total_selected
            result["retry_attempt"] = attempt
            result["error_classification"] = classify_result_error(result)

            results.append(result)
            append_incremental_audit(result)

            if result.get("status") == "OK":
                for previous in results:
                    if (
                        previous is not result
                        and previous.get("series_id") == result.get("series_id")
                        and str(previous.get("status", "")).startswith("ERROR")
                    ):
                        previous["recovered_by_retry"] = True
                        previous["retry_final_status"] = "OK"

            if (
                ENABLE_RETRY_FAILED_SERIES
                and is_retryable_result(result)
                and attempt < MAX_RETRY_ATTEMPTS_PER_SERIES
            ):
                next_attempt = attempt + 1
                retry_item = dict(item)
                retry_item["_retry_attempt"] = next_attempt
                pending_items.append({
                    "local_idx": pending.get("local_idx"),
                    "idx": idx,
                    "item": retry_item,
                    "attempt": next_attempt,
                })
                result["retry_scheduled"] = True
                result["retry_next_attempt"] = next_attempt
                print(
                    f"[RETRY] agendado | idx={idx} | tentativa={next_attempt} | "
                    f"{result.get('source')} | {result.get('asset')} | {result.get('timeframe')} | "
                    f"erro={result.get('error_classification')}",
                    flush=True,
                )
            elif is_retryable_result(result):
                result["retry_scheduled"] = False

            print(
                f"[PROGRESSO] {len(results)}/{total_selected} | "
                f"batch={batch_name} | idx={idx} | status={result.get('status')} | "
                f"{result.get('source')} | {result.get('asset')} | {result.get('timeframe')} | "
                f"retry={attempt} | "
                f"features={result.get('feature_columns_count')} | "
                f"tempo={format_duration(float(result.get('elapsed_seconds') or 0.0))}",
                flush=True,
            )

    append_incremental_audit({
        "event": "BATCH_FINISHED",
        "batch_name": batch_name,
        "batch_series_count": len(batch),
        "processed_total_so_far": len(results),
        "selected_series_count": total_selected,
    })

    print(
        f"[BATCH FIM] {batch_name} | series_originais={len(batch)} | "
        f"elapsed={format_duration(time.time() - batch_started_at)} | "
        f"global={len(results)}/{total_selected}",
        flush=True,
    )

    if FORCE_GC_EACH_SERIES:
        gc.collect()

    return progress_offset + len(batch)


# =============================================================================
# 14. MAIN
# =============================================================================

def main() -> None:
    configure_console_output()
    total_start = time.time()

    ensure_dir(FEATURES_DIR)
    ensure_dir(FEATURES_PARQUET_DIR)
    ensure_dir(FEATURES_LOG_DIR)
    ensure_dir(BASE_JSON_DIR)
    ensure_dir(BASE_EXCEL_DIR)

    print("=" * 120)
    print("ARCHANGEL v1 | 3_FEATURES | GERADOR DE FEATURES V3 FAST AUDITED")
    print("=" * 120)
    print(f"[SCRIPT] {SCRIPT_NAME}")
    print(f"[SCHEMA] {SCHEMA_VERSION}")
    print(f"[RUN_ID] {RUN_ID}")
    print(f"[ROOT_DIR] {ROOT_DIR}")
    print(f"[MAPA_ATIVOS] {MAPA_ATIVOS_PATH}")
    print(f"[FEATURES_PARQUET_DIR] {FEATURES_PARQUET_DIR}")
    print(f"[FEATURES_JSON_PATH] {FEATURES_JSON_PATH}")
    print(f"[FEATURES_EXCEL_PATH] {FEATURES_EXCEL_PATH}")
    print(f"[INCREMENTAL_AUDIT] {INCREMENTAL_AUDIT_PATH}")
    print(f"[INÍCIO] {now_iso()}")
    print("=" * 120)

    if not MAPA_ATIVOS_PATH.exists():
        raise FileNotFoundError(
            f"01_MAPA_ATIVOS_LATEST.json não encontrado em: {MAPA_ATIVOS_PATH}. "
            "Rode antes o script de geração do mapa de ativos."
        )

    mapa = load_json(MAPA_ATIVOS_PATH)
    quality_report = load_quality_report(DATA_QUALITY_REPORT_PATH)

    selected = select_series_to_process(mapa)
    selected = filter_selected_for_retry_only(selected)

    # Injeta relatório de qualidade em cada item.
    # Isso funciona com ProcessPoolExecutor porque dict é serializável.
    for item in selected:
        item["quality_report"] = quality_report

    feature_catalog = build_feature_catalog_from_config()



    print(f"[SÉRIES SELECIONADAS] {len(selected)}")
    print(f"[FEATURES NO CATÁLOGO] {len(feature_catalog)}")
    print(f"[FAMÍLIAS] {len(set(x.get('family') for x in feature_catalog))}")
    print(f"[ONLY_QUALITY_OK] {ONLY_QUALITY_OK}")
    print(f"[FILTER_ASSETS] {FILTER_ASSETS}")
    print(f"[FILTER_SOURCES] {FILTER_SOURCES}")
    print(f"[FILTER_TIMEFRAMES] {FILTER_TIMEFRAMES}")
    print(f"[PARALELISMO] {ENABLE_PARALLEL_PROCESSING}")
    print(f"[FEATURE_CUDA_MODE] {FEATURE_CUDA_MODE}")
    print(f"[FEATURE_CUDA_AUTO_MIN_GPU_MEMORY_MB] {FEATURE_CUDA_AUTO_MIN_GPU_MEMORY_MB}")
    print(f"[FEATURE_CUDA_AUTO_MIN_ROWS] {FEATURE_CUDA_AUTO_MIN_ROWS}")
    print(f"[FEATURE_CUDA_MAX_WORKERS] {FEATURE_CUDA_MAX_WORKERS}")
    print(f"[FEATURE_CUDA_BLOCKS] {list(FEATURE_CUDA_ACCELERATED_BLOCKS)}")
    print(f"[RESOURCE_PROFILE] {FEATURE_RESOURCE_PROFILE}")
    print(f"[MAX_WORKERS_FEATURES] {MAX_WORKERS_FEATURES}")
    print(f"[MEMORY_AWARE_WORKERS] {ENABLE_MEMORY_AWARE_WORKERS}")
    print(f"[RAM_CAP_GB] {ARCHANGEL_RAM_CAP_GB}")
    print(f"[TARGET_RAM_USED_GB] {TARGET_RAM_USED_GB}")
    print(f"[TARGET_CPU_PERCENT] {TARGET_CPU_PERCENT}")
    print(f"[TARGET_FREE_RAM_GB] {get_target_free_ram_gb()}")
    print(f"[RAM_RESERVA_EFETIVA_GB] {get_effective_min_free_ram_gb()}")
    print(f"[TURBO_BATCH_WORKERS] {MAX_WORKERS_TURBO_BATCH}")
    print(f"[ADAPTIVE_WORKERS] {ENABLE_ADAPTIVE_WORKERS}")
    print(
        f"[ADAPTIVE_LIMITS] turbo_ini={ADAPTIVE_WORKERS_INITIAL_TURBO} | "
        f"std_ini={ADAPTIVE_WORKERS_INITIAL_STANDARD} | "
        f"scale_up_ram={ADAPTIVE_SCALE_UP_FREE_RAM_GB}GB | "
        f"hold_ram={ADAPTIVE_HOLD_FREE_RAM_GB}GB | "
        f"down_ram={ADAPTIVE_SCALE_DOWN_FREE_RAM_GB}GB | "
        f"step_up={ADAPTIVE_SCALE_UP_WORKER_STEP}"
    )
    print(f"[RETRY_FAILED_SERIES] {ENABLE_RETRY_FAILED_SERIES} | max={MAX_RETRY_ATTEMPTS_PER_SERIES}")
    print(f"[RETRY_ONLY_FROM_REPORT] {RETRY_FAILED_ONLY_FROM_REPORT}")
    print(f"[FEATURE_EXECUTION_PROFILE] {FEATURE_EXECUTION_PROFILE}")
    print(f"[FEATURE_BLOCK_CONSTRUCTION] {ENABLE_FEATURE_BLOCK_CONSTRUCTION}")
    print(f"[PROGRESS_HEARTBEAT] {ENABLE_PROGRESS_HEARTBEAT} | {PROGRESS_HEARTBEAT_SECONDS}s")
    print(f"[TURBO_MODE] {ENABLE_TIMEFRAME_TURBO_MODE}")
    print(f"[TURBO_TIMEFRAMES] {sorted(TURBO_TIMEFRAMES)}")
    print(f"[POST_AUDIT] {ENABLE_POST_AUDIT}")
    print(f"[TIMEZONE_POLICY] {TIMEZONE_POLICY}")
    print(f"[DATETIME_COL] {DATETIME_COL}")
    print(f"[TIMESTAMP_UTC_MS_COL] {TIMESTAMP_UTC_MS_COL}")
    print(f"[BAR_TIMESTAMP_POLICY] {BAR_TIMESTAMP_POLICY}")
    print(f"[DATA_QUALITY_REPORT] {DATA_QUALITY_REPORT_PATH}")
    print(f"[DATA_QUALITY_REPORT_EXISTS] {DATA_QUALITY_REPORT_PATH.exists()}")
    print(f"[COST_MODEL_PATH] {COST_MODEL_PATH}")
    print(f"[PRE_FEATURE_QUALITY_GATE] {ENABLE_PRE_FEATURE_QUALITY_GATE}")
    print(f"[INPUT_CONTRACT_VALIDATION] {ENABLE_INPUT_CONTRACT_VALIDATION}")
    print(f"[TIME_GRID_VALIDATION] {ENABLE_TIME_GRID_VALIDATION}")
    print(f"[FEATURE_VALIDITY_FLAGS] {ENABLE_FEATURE_VALIDITY_FLAGS}")
    print(f"[ML_READY_SCHEMA_FLAGS] {ENABLE_ML_READY_SCHEMA_FLAGS}")




    if not selected:
        raise RuntimeError("Nenhuma série selecionada. Verifique filtros e 01_MAPA_ATIVOS_LATEST.json.")

    results: List[Dict[str, Any]] = []

    append_incremental_audit({
        "event": "RUN_STARTED",
        "selected_series_count": len(selected),
        "feature_catalog_count": len(feature_catalog),
        "schema_version": SCHEMA_VERSION,
        "config": {
            "only_quality_ok": ONLY_QUALITY_OK,
            "filter_assets": None if FILTER_ASSETS is None else sorted(FILTER_ASSETS),
            "filter_sources": None if FILTER_SOURCES is None else sorted(FILTER_SOURCES),
            "filter_timeframes": None if FILTER_TIMEFRAMES is None else sorted(FILTER_TIMEFRAMES),
            "enable_parallel_processing": ENABLE_PARALLEL_PROCESSING,
            "feature_cuda_mode": FEATURE_CUDA_MODE,
            "feature_cuda_auto_min_gpu_memory_mb": FEATURE_CUDA_AUTO_MIN_GPU_MEMORY_MB,
            "feature_cuda_auto_min_rows": FEATURE_CUDA_AUTO_MIN_ROWS,
            "feature_cuda_max_workers": FEATURE_CUDA_MAX_WORKERS,
            "feature_cuda_accelerated_blocks": list(FEATURE_CUDA_ACCELERATED_BLOCKS),
            "feature_cuda_note": FEATURE_CUDA_NOTE,
            "feature_resource_profile": FEATURE_RESOURCE_PROFILE,
            "max_workers_features": MAX_WORKERS_FEATURES,
            "enable_memory_aware_workers": ENABLE_MEMORY_AWARE_WORKERS,
            "archangel_ram_cap_gb": ARCHANGEL_RAM_CAP_GB,
            "target_ram_used_gb": TARGET_RAM_USED_GB,
            "target_cpu_percent": TARGET_CPU_PERCENT,
            "target_free_ram_gb": get_target_free_ram_gb(),
            "effective_min_free_ram_gb": get_effective_min_free_ram_gb(),
            "min_free_ram_gb_to_start_batch": MIN_FREE_RAM_GB_TO_START_BATCH,
            "estimated_ram_gb_per_worker_standard": ESTIMATED_RAM_GB_PER_WORKER_STANDARD,
            "estimated_ram_gb_per_worker_turbo": ESTIMATED_RAM_GB_PER_WORKER_TURBO,
            "process_turbo_timeframes_first": PROCESS_TURBO_TIMEFRAMES_FIRST,
            "enable_timeframe_batch_execution": ENABLE_TIMEFRAME_BATCH_EXECUTION,
            "max_workers_turbo_batch": MAX_WORKERS_TURBO_BATCH,
            "enable_adaptive_workers": ENABLE_ADAPTIVE_WORKERS,
            "adaptive_workers_initial_turbo": ADAPTIVE_WORKERS_INITIAL_TURBO,
            "adaptive_workers_initial_standard": ADAPTIVE_WORKERS_INITIAL_STANDARD,
            "adaptive_scale_up_free_ram_gb": ADAPTIVE_SCALE_UP_FREE_RAM_GB,
            "adaptive_hold_free_ram_gb": ADAPTIVE_HOLD_FREE_RAM_GB,
            "adaptive_scale_down_free_ram_gb": ADAPTIVE_SCALE_DOWN_FREE_RAM_GB,
            "adaptive_scale_up_stable_heartbeats": ADAPTIVE_SCALE_UP_STABLE_HEARTBEATS,
            "adaptive_scale_up_worker_step": ADAPTIVE_SCALE_UP_WORKER_STEP,
            "adaptive_scale_down_worker_step": ADAPTIVE_SCALE_DOWN_WORKER_STEP,
            "adaptive_failure_cooldown_heartbeats": ADAPTIVE_FAILURE_COOLDOWN_HEARTBEATS,
            "adaptive_cpu_low_scale_up_threshold": ADAPTIVE_CPU_LOW_SCALE_UP_THRESHOLD,
            "enable_retry_failed_series": ENABLE_RETRY_FAILED_SERIES,
            "max_retry_attempts_per_series": MAX_RETRY_ATTEMPTS_PER_SERIES,
            "retry_failed_only_from_report": RETRY_FAILED_ONLY_FROM_REPORT,
            "enable_progress_heartbeat": ENABLE_PROGRESS_HEARTBEAT,
            "progress_heartbeat_seconds": PROGRESS_HEARTBEAT_SECONDS,
            "progress_show_active_limit": PROGRESS_SHOW_ACTIVE_LIMIT,
            "progress_show_recent_limit": PROGRESS_SHOW_RECENT_LIMIT,
            "feature_execution_profile": FEATURE_EXECUTION_PROFILE,
            "enable_feature_block_construction": ENABLE_FEATURE_BLOCK_CONSTRUCTION,
            "feature_block_column_count": FEATURE_BLOCK_COLUMN_COUNT,
            "enable_timeframe_turbo_mode": ENABLE_TIMEFRAME_TURBO_MODE,
            "turbo_timeframes": sorted(TURBO_TIMEFRAMES),
            "enable_post_audit": ENABLE_POST_AUDIT,
            "timezone_policy": TIMEZONE_POLICY,
            "datetime_col": DATETIME_COL,
            "timestamp_utc_ms_col": TIMESTAMP_UTC_MS_COL,
            "bar_timestamp_policy": BAR_TIMESTAMP_POLICY,
            "data_quality_report_path": path_to_str(DATA_QUALITY_REPORT_PATH),
            "data_quality_report_exists": DATA_QUALITY_REPORT_PATH.exists(),
            "cost_model_path": path_to_str(COST_MODEL_PATH),
            "enable_pre_feature_quality_gate": ENABLE_PRE_FEATURE_QUALITY_GATE,
            "enable_input_contract_validation": ENABLE_INPUT_CONTRACT_VALIDATION,
            "enable_time_grid_validation": ENABLE_TIME_GRID_VALIDATION,
            "enable_feature_validity_flags": ENABLE_FEATURE_VALIDITY_FLAGS,
            "enable_ml_ready_schema_flags": ENABLE_ML_READY_SCHEMA_FLAGS,

        },
    })

    try:
        progress_offset = 0
        for batch_name, batch in split_series_batches(selected):
            progress_offset = process_batch(
                batch_name=batch_name,
                batch=batch,
                results=results,
                total_selected=len(selected),
                progress_offset=progress_offset,
            )

    except KeyboardInterrupt:
        print("\n[INTERRUPÇÃO] Execução interrompida manualmente. Salvando relatórios parciais...")

        append_incremental_audit({
            "event": "RUN_INTERRUPTED_KEYBOARD",
            "processed_so_far": len(results),
            "selected_series_count": len(selected),
        })

    except Exception as exc:
        print("\n[ERRO FATAL] Salvando relatórios parciais...")
        print(str(exc))

        append_incremental_audit({
            "event": "RUN_FATAL_ERROR",
            "processed_so_far": len(results),
            "selected_series_count": len(selected),
            "error": str(exc),
            "traceback": traceback.format_exc(),
        })

    finally:
        results = sorted(
            results,
            key=lambda r: (
                str(r.get("source", "")),
                str(r.get("asset", "")),
                str(r.get("symbol", "")),
                str(r.get("timeframe", "")),
                str(r.get("series_id", "")),
            ),
        )

        run_report_path = save_run_report(results, selected)
        retry_plan_json_path, retry_plan_csv_path = save_retry_plan(results, selected)

        features_json_payload = build_features_json_payload(
            mapa=mapa,
            selected=selected,
            results=results,
            run_report_path=run_report_path,
            feature_catalog=feature_catalog,
        )

        write_json_atomic(features_json_payload, FEATURES_JSON_PATH)

        features_excel_path = save_features_excel_report(
            selected=selected,
            results=results,
            feature_catalog=feature_catalog,
        )

        total_elapsed = round(time.time() - total_start, 6)

        ok_count = sum(1 for r in results if r.get("status") == "OK")

        error_count = sum(1 for r in results if str(r.get("status", "")).startswith("ERROR"))

        skipped_count = sum(1 for r in results if str(r.get("status", "")).startswith("SKIPPED"))
        skipped_quality_fail = sum(
                1 for r in results if r.get("status") == "SKIPPED_DATA_QUALITY_FAIL"
            )
        input_contract_fail = sum(
                1 for r in results if r.get("status") == "ERROR_INPUT_CONTRACT_FAIL"
            )
        time_grid_fail = sum(
            1 for r in results if r.get("status") == "SKIPPED_TIME_GRID_FAIL"
            )
        total_valid_feature_rows = int(sum(r.get("valid_feature_rows", 0) or 0 for r in results))
        total_ml_eligible_rows = int(sum(r.get("ml_eligible_rows", 0) or 0 for r in results))




        audit_pass = sum(1 for r in results if (r.get("post_audit") or {}).get("audit_status") == "PASS")
        audit_warning = sum(1 for r in results if (r.get("post_audit") or {}).get("audit_status") == "WARNING")
        audit_fail = sum(1 for r in results if (r.get("post_audit") or {}).get("audit_status") in {"FAIL", "ERROR"})

        append_incremental_audit({
            "event": "RUN_FINISHED_OR_PARTIAL_SAVED",
            "processed_series": len(results),
            "selected_series_count": len(selected),
            "ok": ok_count,
            "error": error_count,
            "skipped": skipped_count,
            "audit_pass": audit_pass,
            "audit_warning": audit_warning,
            "audit_fail": audit_fail,
            "elapsed_seconds": total_elapsed,
            "features_json_path": path_to_str(FEATURES_JSON_PATH),
            "features_excel_path": path_to_str(features_excel_path),
            "run_report_path": path_to_str(run_report_path),
            "retry_plan_json_path": path_to_str(retry_plan_json_path),
            "retry_plan_csv_path": path_to_str(retry_plan_csv_path),
            "skipped_quality_fail": skipped_quality_fail,
            "input_contract_fail": input_contract_fail,
            "time_grid_fail": time_grid_fail,
            "total_valid_feature_rows": total_valid_feature_rows,
            "total_ml_eligible_rows": total_ml_eligible_rows,
        })

        print("=" * 120)
        print("[FINALIZADO / RELATÓRIOS SALVOS]")
        print(f"[PROCESSADAS] {len(results)} / {len(selected)}")
        print(f"[OK] {ok_count}")
        print(f"[ERROS] {error_count}")
        print(f"[SKIPPED] {skipped_count}")
        print(f"[AUDIT PASS] {audit_pass}")
        print(f"[AUDIT WARNING] {audit_warning}")
        print(f"[AUDIT FAIL/ERROR] {audit_fail}")
        print(f"[TEMPO TOTAL] {total_elapsed}s")
        print(f"[JSON FEATURES] {FEATURES_JSON_PATH}")
        print(f"[EXCEL FEATURES] {features_excel_path}")
        print(f"[RUN REPORT] {run_report_path}")
        print(f"[RETRY PLAN JSON] {retry_plan_json_path}")
        print(f"[RETRY PLAN CSV] {retry_plan_csv_path}")
        print(f"[AUDITORIA INCREMENTAL] {INCREMENTAL_AUDIT_PATH}")
        print(f"[SKIPPED QUALITY FAIL] {skipped_quality_fail}")
        print(f"[INPUT CONTRACT FAIL] {input_contract_fail}")
        print(f"[TIME GRID FAIL] {time_grid_fail}")
        print(f"[VALID FEATURE ROWS] {total_valid_feature_rows}")
        print(f"[ML ELIGIBLE ROWS] {total_ml_eligible_rows}")

        print(f"[FIM] {now_iso()}")
        print("=" * 120)


if __name__ == "__main__":
    main()
