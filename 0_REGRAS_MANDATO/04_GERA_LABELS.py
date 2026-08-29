# -*- coding: utf-8 -*-
"""
ARCHANGEL v1 - 04_GERA_LABELS.py
Extended, Audited, Expandable Label Factory

Objetivo:
    Criar uma fábrica extensível de labels para ML/DL, backtesting, meta-labeling,
    risk management e pesquisa quantitativa.

Principais famílias de labels:
    1) Forward returns
    2) Directional threshold labels
    3) Quantile forward return labels
    4) Future path extremes: MFE, MAE, future drawdown, future run-up
    5) Future realized volatility labels
    6) Future trend regime labels
    7) Triple Barrier long/short
    8) Meta-labels derivados de Triple Barrier
    9) Risk/reward future labels
    10) Tail-event labels
    11) Cross-sectional rank labels, quando houver painel compatível
    12) Relative-to-benchmark labels, quando houver benchmarks compatíveis

Governança:
    - Labels usam futuro por definição.
    - Nunca usar colunas label_* como features.
    - Usar purging e embargo no walk-forward.
    - DateTime preservado como naive, conforme política Asia/Dubai.
"""

from __future__ import annotations

import os
import gc
import re
import json
import time
import math
import hashlib
import shutil
import subprocess
import traceback
import warnings
from pathlib import Path
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple
from concurrent.futures import ProcessPoolExecutor, as_completed
from threading import Event, Lock, Thread

import numpy as np
import pandas as pd
from pandas.errors import PerformanceWarning


def env_bool(name: str, default: bool) -> bool:
    raw = os.environ.get(name)
    if raw is None:
        return default
    return str(raw).strip().lower() in {"1", "true", "yes", "y", "sim", "on"}


# =============================================================================
# 0. WARNINGS
# =============================================================================

warnings.simplefilter("ignore", PerformanceWarning)
warnings.filterwarnings("ignore", category=RuntimeWarning)


# =============================================================================
# 1. PATHS GERAIS DO SISTEMA
# =============================================================================

# A raiz é inferida do local deste script, sem depender da letra da unidade.
ROOT_DIR = Path(__file__).resolve().parent.parent

RULES_DIR = ROOT_DIR / "0_REGRAS_MANDATO"
BASE_JSON_DIR = RULES_DIR / "BASE_JSON"
BASE_EXCEL_DIR = RULES_DIR / "BASE_EXCEL"

FEATURES_DIR = ROOT_DIR / "3_FEATURES"
FEATURES_PARQUET_DIR = FEATURES_DIR / "FEATURES_PARQUET"

LABELS_DIR = ROOT_DIR / "4_LABELS"
LABELS_PARQUET_DIR = LABELS_DIR / "LABELS_PARQUET"
LABELS_LOG_DIR = LABELS_DIR / "_logs"
LABELS_AUDIT_DIR = LABELS_DIR / "_audit"
LABELS_EXCEL_DIR = LABELS_DIR / "_excel"

MAPA_ATIVOS_PATH = BASE_JSON_DIR / "01_MAPA_ATIVOS_LATEST.json"
FEATURES_JSON_PATH = BASE_JSON_DIR / "03_FEATURES_CATALOG_LATEST.json"
BASE_ARQUIVOS_PATH = BASE_JSON_DIR / "00_03_BASE_ARQUIVOS_LATEST.json"
MACHINE_PROFILE_PATH = BASE_JSON_DIR / "00_01_MACHINE_PROFILE_LATEST.json"

LABELS_JSON_PATH = BASE_JSON_DIR / "04_LABELS_CATALOG_LATEST.json"
LABELS_EXCEL_PATH = BASE_EXCEL_DIR / "4_LABELS.xlsx"
COST_MODEL_PATH = BASE_JSON_DIR / "00_COST_MODEL.json"

SCRIPT_NAME = "04_GERA_LABELS.py"
SYSTEM_NAME = "ARCHANGEL"
SYSTEM_VERSION = "v1"
SCHEMA_VERSION = "ARCHANGEL_LABEL_STORE_2.0_EXTENDED_LABEL_FACTORY"

RUN_ID = datetime.now().strftime("%Y%m%d_%H%M%S")

RUN_REPORT_PATH = LABELS_LOG_DIR / f"4_LABELS_RUN_REPORT_{RUN_ID}.json"
RUN_REPORT_LATEST_PATH = BASE_JSON_DIR / "04_LABELS_RUN_REPORT_LATEST.json"
RUN_REPORT_BASE_JSON_PATH = RUN_REPORT_LATEST_PATH
INCREMENTAL_AUDIT_PATH = LABELS_LOG_DIR / f"4_LABELS_RUN_AUDIT_INCREMENTAL_{RUN_ID}.jsonl"

TIMEZONE_POLICY = {
    "reference_timezone": "Asia/Dubai",
    "datetime_assumption": "DateTime salvo como Dubai local time naive.",
    "warning": "Este script preserva DateTime como naive. Não converte timezone.",
}


# =============================================================================
# 2. CONFIGURAÇÕES OPERACIONAIS
# =============================================================================

OVERWRITE_EXISTING = True

ENABLE_PARALLEL_PROCESSING = True
MAX_WORKERS_LABELS = int(os.environ.get("ARCHANGEL_MAX_WORKERS_LABELS", "8"))
LABEL_RESOURCE_PROFILE = os.environ.get("ARCHANGEL_LABEL_RESOURCE_PROFILE", "RAM_SAFE").strip().upper()
RESOURCE_TELEMETRY_ENABLED = env_bool("ARCHANGEL_LABEL_RESOURCE_TELEMETRY", True)
RESOURCE_TELEMETRY_INTERVAL_SECONDS = max(1.0, float(os.environ.get("ARCHANGEL_LABEL_TELEMETRY_INTERVAL_SECONDS", "10")))
RESOURCE_TELEMETRY_MAX_SAMPLES = max(10, int(os.environ.get("ARCHANGEL_LABEL_TELEMETRY_MAX_SAMPLES", "2000")))

MIN_ROWS = 300
FORCE_GC_EACH_SERIES = True
VERBOSE_SERIES_LOG = False

PARQUET_ENGINE = "pyarrow"
PARQUET_COMPRESSION = "zstd"

READ_ONLY_NEEDED_COLUMNS = True
DOWNCAST_FLOATS_TO_FLOAT32 = True

SAVE_RAW_LABEL_COMPONENTS = True
SAVE_AUDIT_COLUMNS_IN_PARQUET = True

ENABLE_EXCEL_REPORT = True
ENABLE_JSON_MASTER_REPORT = True
ENABLE_INCREMENTAL_AUDIT = True

# Descoberta de features
DISCOVER_FEATURE_FILES_FROM_FEATURES_JSON = True
DISCOVER_FEATURE_FILES_FROM_DIRECTORY_FALLBACK = False
SKIP_UNKNOWN_METADATA_FEATURE_FILES = True

# Filtros opcionais. None = processa tudo descoberto dinamicamente.
FILTER_ASSETS: Optional[set[str]] = None
FILTER_SOURCES: Optional[set[str]] = None
FILTER_TIMEFRAMES: Optional[set[str]] = None

# Labels cross-sectional / benchmark precisam de painel carregado.
ENABLE_PANEL_LABELS = True
PANEL_MAX_ROWS_PER_FILE = None

BENCHMARK_ASSETS = ["BTC", "ETH"]
BASKET_EXCLUDE_SELF = True


# =============================================================================
# 3. CUSTOS E SLIPPAGE
# =============================================================================

DEFAULT_COST_CONFIG = {
    "binance_spot": {
        "fee_bps_round_trip": 20.0,
        "slippage_bps_round_trip": 10.0,
        "funding_bps_per_day": 0.0,
        "min_net_edge_bps": 5.0,
    },
    "bybit_linear": {
        "fee_bps_round_trip": 12.0,
        "slippage_bps_round_trip": 12.0,
        "funding_bps_per_day": 2.0,
        "min_net_edge_bps": 5.0,
    },
    "default": {
        "fee_bps_round_trip": 20.0,
        "slippage_bps_round_trip": 15.0,
        "funding_bps_per_day": 0.0,
        "min_net_edge_bps": 5.0,
    },
}


def normalize_cost_model_payload(payload: Dict[str, Any]) -> Dict[str, Dict[str, float]]:
    """Aceita o schema legado por source e o schema novo com default/symbols."""
    if not isinstance(payload, dict):
        return dict(DEFAULT_COST_CONFIG)

    if any(key in payload for key in ("binance_spot", "bybit_linear")):
        cost_source_keys = set(DEFAULT_COST_CONFIG) | {"binance_spot", "bybit_linear"}
        out = dict(DEFAULT_COST_CONFIG)
        for key, cfg in payload.items():
            if isinstance(cfg, dict) and key in cost_source_keys:
                out[key] = {
                    "fee_bps_round_trip": float(cfg.get("fee_bps_round_trip", out.get(key, {}).get("fee_bps_round_trip", 20.0))),
                    "slippage_bps_round_trip": float(cfg.get("slippage_bps_round_trip", out.get(key, {}).get("slippage_bps_round_trip", 15.0))),
                    "funding_bps_per_day": float(cfg.get("funding_bps_per_day", out.get(key, {}).get("funding_bps_per_day", 0.0))),
                    "min_net_edge_bps": float(cfg.get("min_net_edge_bps", out.get(key, {}).get("min_net_edge_bps", 5.0))),
                }
        return out

    default = payload.get("default", {}) if isinstance(payload.get("default"), dict) else {}
    taker_fee = float(default.get("taker_fee_bps", 5.5))
    slippage = float(default.get("slippage_bps", 2.0))
    funding = float(default.get("funding_bps_per_day", 0.0))
    maker_fee = float(default.get("maker_fee_bps", taker_fee))
    min_edge = float(default.get("min_net_edge_bps", 5.0))

    round_trip_fee = 2.0 * taker_fee
    round_trip_slippage = 2.0 * slippage

    return {
        "default": {
            "fee_bps_round_trip": round_trip_fee,
            "slippage_bps_round_trip": round_trip_slippage,
            "funding_bps_per_day": funding,
            "min_net_edge_bps": min_edge,
            "maker_fee_bps_one_way": maker_fee,
            "taker_fee_bps_one_way": taker_fee,
        },
        str(payload.get("source") or payload.get("exchange") or "default"): {
            "fee_bps_round_trip": round_trip_fee,
            "slippage_bps_round_trip": round_trip_slippage,
            "funding_bps_per_day": funding,
            "min_net_edge_bps": min_edge,
            "maker_fee_bps_one_way": maker_fee,
            "taker_fee_bps_one_way": taker_fee,
        },
    }


def load_cost_model(path: Path = COST_MODEL_PATH) -> Dict[str, Dict[str, float]]:
    if not path.exists():
        return dict(DEFAULT_COST_CONFIG)

    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
        return normalize_cost_model_payload(payload)
    except Exception:
        return dict(DEFAULT_COST_CONFIG)


COST_CONFIG = load_cost_model()


# =============================================================================
# 4. LABEL CONFIG REGISTRY - EXPANSÍVEL
# =============================================================================

LABEL_CONFIGS: List[Dict[str, Any]] = [
    # -------------------------------------------------------------------------
    # Forward returns
    # -------------------------------------------------------------------------
    {
        "label_config_id": "FWD_RETURNS_EXPANDED",
        "label_type": "forward_returns",
        "enabled": True,
        "horizons": [1, 2, 3, 5, 8, 10, 13, 16, 20, 30, 40, 60, 90, 120, 180, 240, 360],
        "description": "Forward returns brutos e líquidos para múltiplos horizontes.",
    },

    # -------------------------------------------------------------------------
    # Directional thresholds
    # -------------------------------------------------------------------------
    {
        "label_config_id": "DIR_BPS_MULTI_THRESHOLD",
        "label_type": "directional_threshold",
        "enabled": True,
        "horizons": [3, 5, 10, 20, 30, 60, 120],
        "threshold_bps_list": [5.0, 10.0, 15.0, 25.0, 40.0, 75.0, 100.0],
        "description": "Labels direcionais com múltiplos thresholds em bps.",
    },

    # -------------------------------------------------------------------------
    # Quantile labels
    # -------------------------------------------------------------------------
    {
        "label_config_id": "FWD_RETURN_QUANTILES",
        "label_type": "quantile_forward_return",
        "enabled": True,
        "horizons": [10, 20, 30, 60, 120],
        "rolling_window": 1000,
        "quantiles": [0.10, 0.25, 0.50, 0.75, 0.90],
        "description": "Classifica retorno futuro em buckets por quantis rolling.",
    },

    # -------------------------------------------------------------------------
    # Future path extremes
    # -------------------------------------------------------------------------
    {
        "label_config_id": "FUTURE_PATH_EXTREMES",
        "label_type": "future_path_extremes",
        "enabled": True,
        "horizons": [5, 10, 20, 30, 60, 120],
        "description": "MFE, MAE, drawdown futuro e run-up futuro dentro do horizonte.",
    },

    # -------------------------------------------------------------------------
    # Future realized volatility
    # -------------------------------------------------------------------------
    {
        "label_config_id": "FUTURE_REALIZED_VOL",
        "label_type": "future_realized_vol",
        "enabled": True,
        "horizons": [10, 20, 30, 60, 120],
        "vol_threshold_quantiles": [0.25, 0.50, 0.75, 0.90],
        "rolling_window": 1000,
        "description": "Volatilidade realizada futura e regime de vol futura.",
    },

    # -------------------------------------------------------------------------
    # Future trend regime
    # -------------------------------------------------------------------------
    {
        "label_config_id": "FUTURE_TREND_REGIME",
        "label_type": "trend_regime_forward",
        "enabled": True,
        "horizons": [10, 20, 30, 60, 120],
        "trend_bps_thresholds": [10.0, 25.0, 50.0, 100.0],
        "description": "Regime futuro de tendência, range ou reversão.",
    },

    # -------------------------------------------------------------------------
    # Risk/reward labels
    # -------------------------------------------------------------------------
    {
        "label_config_id": "FUTURE_RISK_REWARD",
        "label_type": "risk_reward_forward",
        "enabled": True,
        "horizons": [10, 20, 30, 60, 120],
        "rr_thresholds": [0.75, 1.0, 1.5, 2.0, 3.0],
        "description": "Labels baseados na relação MFE/MAE futura.",
    },

    # -------------------------------------------------------------------------
    # Tail event labels
    # -------------------------------------------------------------------------
    {
        "label_config_id": "TAIL_EVENTS_FORWARD",
        "label_type": "tail_event_forward",
        "enabled": True,
        "horizons": [10, 20, 30, 60, 120],
        "tail_bps_list": [50.0, 100.0, 150.0, 250.0, 500.0],
        "description": "Eventos extremos positivos/negativos futuros.",
    },

    # -------------------------------------------------------------------------
    # Triple Barrier - volatility
    # -------------------------------------------------------------------------
    {
        "label_config_id": "TB_VOL_GRID",
        "label_type": "triple_barrier_grid",
        "enabled": True,
        "barrier_type": "volatility",
        "vol_windows": [20, 50, 100],
        "horizon_bars_list": [8, 16, 32, 64, 128],
        "tp_mult_list": [1.0, 1.5, 2.0, 3.0],
        "sl_mult_list": [0.75, 1.0, 1.5],
        "min_barrier_bps": 10.0,
        "description": "Grade expansiva de Triple Barrier baseado em realized volatility.",
    },

    # -------------------------------------------------------------------------
    # Triple Barrier - ATR
    # -------------------------------------------------------------------------
    {
        "label_config_id": "TB_ATR_GRID",
        "label_type": "triple_barrier_grid",
        "enabled": True,
        "barrier_type": "atr",
        "atr_windows": [14, 21, 50],
        "horizon_bars_list": [8, 16, 32, 64],
        "tp_mult_list": [1.0, 1.5, 2.0],
        "sl_mult_list": [0.75, 1.0, 1.5],
        "min_barrier_bps": 10.0,
        "description": "Grade expansiva de Triple Barrier baseado em ATR.",
    },

    # -------------------------------------------------------------------------
    # Triple Barrier - fixed bps
    # -------------------------------------------------------------------------
    {
        "label_config_id": "TB_FIXED_BPS_GRID",
        "label_type": "triple_barrier_grid",
        "enabled": True,
        "barrier_type": "fixed_bps",
        "horizon_bars_list": [5, 10, 20, 30, 60, 120],
        "tp_bps_list": [20.0, 40.0, 75.0, 100.0, 150.0, 250.0],
        "sl_bps_list": [15.0, 25.0, 50.0, 75.0, 100.0, 150.0],
        "min_barrier_bps": 10.0,
        "description": "Grade expansiva de Triple Barrier fixo em bps.",
    },

    # -------------------------------------------------------------------------
    # Cross-sectional / benchmark labels
    # -------------------------------------------------------------------------
    {
        "label_config_id": "RELATIVE_TO_BENCHMARKS",
        "label_type": "relative_to_benchmark",
        "enabled": True,
        "horizons": [10, 20, 30, 60, 120],
        "benchmarks": ["BTC", "ETH"],
        "outperformance_threshold_bps": [10.0, 25.0, 50.0],
        "description": "Label de retorno futuro relativo a BTC/ETH.",
    },
    {
        "label_config_id": "CROSS_SECTIONAL_RANK",
        "label_type": "cross_sectional_rank",
        "enabled": True,
        "horizons": [10, 20, 30, 60, 120],
        "top_quantile": 0.25,
        "bottom_quantile": 0.25,
        "description": "Label de ranking futuro cross-sectional por DateTime/source/timeframe.",
    },
]


# =============================================================================
# 5. UTILITÁRIOS
# =============================================================================

def now_iso() -> str:
    return datetime.now().isoformat(timespec="seconds")


def ensure_dir(path: Path) -> Path:
    path.mkdir(parents=True, exist_ok=True)
    return path


def path_to_str(path: Path) -> str:
    return str(path)


def load_json(path: Path, required: bool = True) -> Dict[str, Any]:
    if not path.exists():
        if required:
            raise FileNotFoundError(f"JSON obrigatório não encontrado: {path}")
        return {}
    return json.loads(path.read_text(encoding="utf-8"))


def write_json_atomic(payload: Dict[str, Any], path: Path) -> None:
    ensure_dir(path.parent)
    tmp = path.with_suffix(path.suffix + ".tmp")
    tmp.write_text(json.dumps(payload, ensure_ascii=False, indent=4, default=str), encoding="utf-8")
    tmp.replace(path)


def append_jsonl(path: Path, payload: Dict[str, Any]) -> None:
    ensure_dir(path.parent)
    with path.open("a", encoding="utf-8") as f:
        f.write(json.dumps(payload, ensure_ascii=False, default=str) + "\n")


def append_incremental_audit(record: Dict[str, Any]) -> None:
    if not ENABLE_INCREMENTAL_AUDIT:
        return
    payload = {
        "event_time": now_iso(),
        "run_id": RUN_ID,
        "script": SCRIPT_NAME,
        **record,
    }
    append_jsonl(INCREMENTAL_AUDIT_PATH, payload)


RESOURCE_TELEMETRY_SAMPLES: List[Dict[str, Any]] = []
RESOURCE_TELEMETRY_LOCK = Lock()
RESOURCE_TELEMETRY_STOP_EVENT = Event()
_CPU_TIMES_PREVIOUS: Optional[Tuple[int, int, int]] = None


def get_process_memory_mb() -> Optional[float]:
    try:
        import psutil  # type: ignore
        return round(psutil.Process(os.getpid()).memory_info().rss / (1024 ** 2), 3)
    except Exception:
        pass

    if os.name == "nt":
        try:
            import ctypes
            from ctypes import wintypes

            class PROCESS_MEMORY_COUNTERS_EX(ctypes.Structure):
                _fields_ = [
                    ("cb", wintypes.DWORD),
                    ("PageFaultCount", wintypes.DWORD),
                    ("PeakWorkingSetSize", ctypes.c_size_t),
                    ("WorkingSetSize", ctypes.c_size_t),
                    ("QuotaPeakPagedPoolUsage", ctypes.c_size_t),
                    ("QuotaPagedPoolUsage", ctypes.c_size_t),
                    ("QuotaPeakNonPagedPoolUsage", ctypes.c_size_t),
                    ("QuotaNonPagedPoolUsage", ctypes.c_size_t),
                    ("PagefileUsage", ctypes.c_size_t),
                    ("PeakPagefileUsage", ctypes.c_size_t),
                    ("PrivateUsage", ctypes.c_size_t),
                ]

            psapi = ctypes.WinDLL("psapi.dll")
            counters = PROCESS_MEMORY_COUNTERS_EX()
            counters.cb = ctypes.sizeof(PROCESS_MEMORY_COUNTERS_EX)
            handle = ctypes.windll.kernel32.GetCurrentProcess()
            psapi.GetProcessMemoryInfo.argtypes = [
                wintypes.HANDLE,
                ctypes.POINTER(PROCESS_MEMORY_COUNTERS_EX),
                wintypes.DWORD,
            ]
            psapi.GetProcessMemoryInfo.restype = wintypes.BOOL
            ok = psapi.GetProcessMemoryInfo(handle, ctypes.byref(counters), counters.cb)
            if ok:
                return round(float(counters.WorkingSetSize) / (1024 ** 2), 3)
        except Exception:
            return None

    return None


def get_windows_memory_snapshot() -> Dict[str, Optional[float]]:
    if os.name != "nt":
        return {
            "ram_total_gb": None,
            "ram_available_gb": None,
            "ram_used_gb": None,
            "ram_percent": None,
        }

    try:
        import ctypes

        class MEMORYSTATUSEX(ctypes.Structure):
            _fields_ = [
                ("dwLength", ctypes.c_ulong),
                ("dwMemoryLoad", ctypes.c_ulong),
                ("ullTotalPhys", ctypes.c_ulonglong),
                ("ullAvailPhys", ctypes.c_ulonglong),
                ("ullTotalPageFile", ctypes.c_ulonglong),
                ("ullAvailPageFile", ctypes.c_ulonglong),
                ("ullTotalVirtual", ctypes.c_ulonglong),
                ("ullAvailVirtual", ctypes.c_ulonglong),
                ("sullAvailExtendedVirtual", ctypes.c_ulonglong),
            ]

        status = MEMORYSTATUSEX()
        status.dwLength = ctypes.sizeof(MEMORYSTATUSEX)
        if not ctypes.windll.kernel32.GlobalMemoryStatusEx(ctypes.byref(status)):
            raise RuntimeError("GlobalMemoryStatusEx failed")

        total_gb = float(status.ullTotalPhys) / (1024 ** 3)
        available_gb = float(status.ullAvailPhys) / (1024 ** 3)
        used_gb = max(0.0, total_gb - available_gb)
        return {
            "ram_total_gb": round(total_gb, 3),
            "ram_available_gb": round(available_gb, 3),
            "ram_used_gb": round(used_gb, 3),
            "ram_percent": round((used_gb / total_gb) * 100.0, 2) if total_gb else None,
        }
    except Exception:
        return {
            "ram_total_gb": None,
            "ram_available_gb": None,
            "ram_used_gb": None,
            "ram_percent": None,
        }


def get_system_cpu_percent() -> Optional[float]:
    try:
        import psutil  # type: ignore
        return round(float(psutil.cpu_percent(interval=None)), 2)
    except Exception:
        pass

    if os.name == "nt":
        try:
            import ctypes
            from ctypes import wintypes

            class FILETIME(ctypes.Structure):
                _fields_ = [
                    ("dwLowDateTime", wintypes.DWORD),
                    ("dwHighDateTime", wintypes.DWORD),
                ]

            def filetime_to_int(value: FILETIME) -> int:
                return (int(value.dwHighDateTime) << 32) + int(value.dwLowDateTime)

            idle = FILETIME()
            kernel = FILETIME()
            user = FILETIME()
            ok = ctypes.windll.kernel32.GetSystemTimes(
                ctypes.byref(idle),
                ctypes.byref(kernel),
                ctypes.byref(user),
            )
            if not ok:
                return None

            current = (
                filetime_to_int(idle),
                filetime_to_int(kernel),
                filetime_to_int(user),
            )

            global _CPU_TIMES_PREVIOUS
            previous = _CPU_TIMES_PREVIOUS
            _CPU_TIMES_PREVIOUS = current
            if previous is None:
                return None

            idle_delta = max(0, current[0] - previous[0])
            kernel_delta = max(0, current[1] - previous[1])
            user_delta = max(0, current[2] - previous[2])
            total_delta = kernel_delta + user_delta
            if total_delta <= 0:
                return None

            busy_ratio = 1.0 - (idle_delta / total_delta)
            return round(max(0.0, min(100.0, busy_ratio * 100.0)), 2)
        except Exception:
            return None

    return None


def get_cpu_memory_snapshot() -> Dict[str, Any]:
    snapshot: Dict[str, Any] = {
        "psutil_available": False,
        "cpu_percent": None,
        "ram_total_gb": None,
        "ram_available_gb": None,
        "ram_used_gb": None,
        "ram_percent": None,
        "process_memory_mb": get_process_memory_mb(),
    }

    try:
        import psutil  # type: ignore
        vm = psutil.virtual_memory()
        snapshot.update({
            "psutil_available": True,
            "cpu_percent": round(float(psutil.cpu_percent(interval=None)), 2),
            "ram_total_gb": round(float(vm.total) / (1024 ** 3), 3),
            "ram_available_gb": round(float(vm.available) / (1024 ** 3), 3),
            "ram_used_gb": round(float(vm.used) / (1024 ** 3), 3),
            "ram_percent": round(float(vm.percent), 2),
            "process_memory_mb": get_process_memory_mb(),
        })
    except Exception as exc:
        fallback = get_windows_memory_snapshot()
        snapshot.update(fallback)
        snapshot["cpu_percent"] = get_system_cpu_percent()
        snapshot["process_memory_mb"] = get_process_memory_mb()
        snapshot["fallback"] = "windows_ctypes" if os.name == "nt" else None
        snapshot["error"] = str(exc)

    return snapshot


def parse_float_or_none(value: str) -> Optional[float]:
    text = str(value).strip()
    if not text or text.upper() in {"N/A", "[N/A]", "NOT SUPPORTED"}:
        return None
    try:
        return float(text)
    except Exception:
        return None


def get_gpu_snapshot() -> Dict[str, Any]:
    exe = shutil.which("nvidia-smi")
    if not exe:
        return {
            "nvidia_smi_available": False,
            "gpu_count": 0,
            "gpus": [],
        }

    query = (
        "index,name,utilization.gpu,utilization.memory,"
        "memory.total,memory.used,memory.free,temperature.gpu,power.draw"
    )

    try:
        completed = subprocess.run(
            [
                exe,
                f"--query-gpu={query}",
                "--format=csv,noheader,nounits",
            ],
            capture_output=True,
            text=True,
            timeout=5,
            check=False,
        )

        if completed.returncode != 0:
            return {
                "nvidia_smi_available": True,
                "gpu_count": 0,
                "gpus": [],
                "error": completed.stderr.strip(),
            }

        gpus = []
        for line in completed.stdout.splitlines():
            parts = [part.strip() for part in line.split(",")]
            if len(parts) < 9:
                continue
            gpus.append({
                "index": int(parse_float_or_none(parts[0]) or 0),
                "name": parts[1],
                "utilization_gpu_pct": parse_float_or_none(parts[2]),
                "utilization_memory_pct": parse_float_or_none(parts[3]),
                "memory_total_mb": parse_float_or_none(parts[4]),
                "memory_used_mb": parse_float_or_none(parts[5]),
                "memory_free_mb": parse_float_or_none(parts[6]),
                "temperature_gpu_c": parse_float_or_none(parts[7]),
                "power_draw_w": parse_float_or_none(parts[8]),
            })

        return {
            "nvidia_smi_available": True,
            "gpu_count": len(gpus),
            "gpus": gpus,
        }

    except Exception as exc:
        return {
            "nvidia_smi_available": True,
            "gpu_count": 0,
            "gpus": [],
            "error": str(exc),
        }


def collect_resource_telemetry_sample(event: str) -> Dict[str, Any]:
    sample = {
        "event_time": now_iso(),
        "event": event,
        "cpu_memory": get_cpu_memory_snapshot(),
        "gpu": get_gpu_snapshot(),
    }
    return sample


def append_resource_telemetry_sample(event: str) -> None:
    if not RESOURCE_TELEMETRY_ENABLED:
        return

    sample = collect_resource_telemetry_sample(event)
    with RESOURCE_TELEMETRY_LOCK:
        RESOURCE_TELEMETRY_SAMPLES.append(sample)
        if len(RESOURCE_TELEMETRY_SAMPLES) > RESOURCE_TELEMETRY_MAX_SAMPLES:
            del RESOURCE_TELEMETRY_SAMPLES[:len(RESOURCE_TELEMETRY_SAMPLES) - RESOURCE_TELEMETRY_MAX_SAMPLES]


def resource_telemetry_loop() -> None:
    append_resource_telemetry_sample("TELEMETRY_STARTED")
    while not RESOURCE_TELEMETRY_STOP_EVENT.wait(RESOURCE_TELEMETRY_INTERVAL_SECONDS):
        append_resource_telemetry_sample("TELEMETRY_SAMPLE")
    append_resource_telemetry_sample("TELEMETRY_STOPPED")


def start_resource_telemetry() -> Optional[Thread]:
    if not RESOURCE_TELEMETRY_ENABLED:
        return None
    RESOURCE_TELEMETRY_STOP_EVENT.clear()
    thread = Thread(target=resource_telemetry_loop, name="labels-resource-telemetry", daemon=True)
    thread.start()
    return thread


def stop_resource_telemetry(thread: Optional[Thread]) -> None:
    if thread is None:
        return
    RESOURCE_TELEMETRY_STOP_EVENT.set()
    thread.join(timeout=max(2.0, RESOURCE_TELEMETRY_INTERVAL_SECONDS + 1.0))


def summarize_resource_telemetry(samples: List[Dict[str, Any]]) -> Dict[str, Any]:
    cpu_values = []
    ram_used_values = []
    ram_percent_values = []
    process_mem_values = []
    gpu_util_by_index: Dict[int, List[float]] = {}
    gpu_mem_by_index: Dict[int, List[float]] = {}

    for sample in samples:
        cpu_mem = sample.get("cpu_memory") or {}
        for key, bucket in [
            ("cpu_percent", cpu_values),
            ("ram_used_gb", ram_used_values),
            ("ram_percent", ram_percent_values),
            ("process_memory_mb", process_mem_values),
        ]:
            value = cpu_mem.get(key)
            if value is not None:
                bucket.append(float(value))

        for gpu in ((sample.get("gpu") or {}).get("gpus") or []):
            index = int(gpu.get("index") or 0)
            util = gpu.get("utilization_gpu_pct")
            mem = gpu.get("memory_used_mb")
            if util is not None:
                gpu_util_by_index.setdefault(index, []).append(float(util))
            if mem is not None:
                gpu_mem_by_index.setdefault(index, []).append(float(mem))

    def stats(values: List[float]) -> Dict[str, Optional[float]]:
        if not values:
            return {"avg": None, "max": None, "min": None}
        return {
            "avg": round(float(np.mean(values)), 3),
            "max": round(float(np.max(values)), 3),
            "min": round(float(np.min(values)), 3),
        }

    gpu_summary = {}
    for index, values in gpu_util_by_index.items():
        gpu_summary[str(index)] = {
            "utilization_gpu_pct": stats(values),
            "memory_used_mb": stats(gpu_mem_by_index.get(index, [])),
        }

    return {
        "sample_count": len(samples),
        "cpu_percent": stats(cpu_values),
        "ram_used_gb": stats(ram_used_values),
        "ram_percent": stats(ram_percent_values),
        "process_memory_mb": stats(process_mem_values),
        "gpu": gpu_summary,
    }


def build_resource_telemetry_payload() -> Dict[str, Any]:
    with RESOURCE_TELEMETRY_LOCK:
        samples = list(RESOURCE_TELEMETRY_SAMPLES)

    return {
        "enabled": RESOURCE_TELEMETRY_ENABLED,
        "interval_seconds": RESOURCE_TELEMETRY_INTERVAL_SECONDS,
        "max_samples": RESOURCE_TELEMETRY_MAX_SAMPLES,
        "summary": summarize_resource_telemetry(samples),
        "samples": samples,
    }


def safe_filename_token(value: Any) -> str:
    text = str(value)
    text = text.replace("\\", "_").replace("/", "_").replace(":", "_")
    text = text.replace(" ", "_").replace(".", "_")
    return re.sub(r"[^A-Za-z0-9_\-]+", "_", text)


def make_hash_from_dataframe(df: pd.DataFrame, rows: int = 80) -> str:
    try:
        if df.empty:
            return "empty"
        sample = pd.concat([df.head(rows), df.tail(rows)], axis=0)
        return hashlib.sha256(sample.to_csv(index=False).encode("utf-8", errors="ignore")).hexdigest()
    except Exception:
        return "hash_error"


def infer_timeframe_seconds(timeframe: str | None) -> Optional[int]:
    if not timeframe:
        return None

    mapping = {
        "1min": 60, "2min": 120, "3min": 180, "5min": 300, "7min": 420,
        "13min": 780, "15min": 900, "23min": 1380, "37min": 2220, "47min": 2820,
        "1h": 3600, "4h": 14400, "8h": 28800, "10h": 36000, "11h": 39600,
        "12h": 43200, "1D": 86400, "3D": 259200, "7D": 604800,
    }

    tf = str(timeframe).strip()
    if tf in mapping:
        return mapping[tf]

    lower = tf.lower()
    try:
        if lower.endswith("s"):
            return int(lower[:-1])
        if lower.endswith("min"):
            return int(lower.replace("min", "")) * 60
        if lower.endswith("h"):
            return int(lower.replace("h", "")) * 3600
        if lower.endswith("d"):
            return int(lower.replace("d", "")) * 86400
    except Exception:
        return None

    return None


def annualization_factor_from_timeframe(timeframe: Optional[str]) -> float:
    seconds = infer_timeframe_seconds(timeframe)
    if seconds is None or seconds <= 0:
        return 365.0
    return float(365 * 24 * 3600) / float(seconds)


def safe_div(a: Any, b: Any) -> Any:
    with np.errstate(divide="ignore", invalid="ignore"):
        return a / b


def bps_to_decimal(bps: float) -> float:
    return float(bps) / 10000.0


def decimal_to_bps(x: Any) -> Any:
    return x * 10000.0


def normalize_datetime_column(df: pd.DataFrame) -> pd.DataFrame:
    if "DateTime" not in df.columns:
        raise ValueError("Coluna DateTime ausente.")

    df = df.copy()
    df["DateTime"] = pd.to_datetime(df["DateTime"], errors="coerce")

    try:
        if getattr(df["DateTime"].dt, "tz", None) is not None:
            df["DateTime"] = df["DateTime"].dt.tz_localize(None)
    except Exception:
        pass

    df = df[df["DateTime"].notna()].copy()
    df = df.sort_values("DateTime", kind="mergesort")
    df = df.drop_duplicates(subset=["DateTime"], keep="last")
    df = df.reset_index(drop=True)
    return df


def downcast_float_columns(df: pd.DataFrame) -> pd.DataFrame:
    if not DOWNCAST_FLOATS_TO_FLOAT32:
        return df
    float_cols = df.select_dtypes(include=["float64"]).columns
    if len(float_cols) > 0:
        df[float_cols] = df[float_cols].astype("float32", copy=False)
    return df


def get_cost_config(source: str) -> Dict[str, float]:
    source_key = str(source or "").lower()
    return dict(COST_CONFIG.get(source_key, COST_CONFIG.get("default", DEFAULT_COST_CONFIG["default"])))


def rolling_quantile_label(
    series: pd.Series,
    window: int,
    quantiles: List[float],
    prefix: str,
) -> pd.DataFrame:
    out = pd.DataFrame(index=series.index)

    for q in quantiles:
        q_col = f"{prefix}_rolling_q{int(q * 100)}"
        out[q_col] = series.rolling(window, min_periods=max(50, int(window * 0.2))).quantile(q)

    return out


# =============================================================================
# 6. DESCOBERTA DE FEATURES
# =============================================================================

def extract_feature_files_from_features_json(features_json: Dict[str, Any]) -> List[Dict[str, Any]]:
    files: List[Dict[str, Any]] = []

    series_outputs = features_json.get("series_outputs", [])
    if not isinstance(series_outputs, list):
        return files

    for item in series_outputs:
        if not isinstance(item, dict):
            continue

        if item.get("status") != "OK":
            continue

        output_path = item.get("output_path")
        if not output_path:
            continue

        path = Path(str(output_path))
        if not path.exists():
            continue

        files.append({
            "feature_path": path_to_str(path),
            "asset": item.get("asset"),
            "symbol": item.get("symbol"),
            "source": item.get("source"),
            "timeframe": item.get("timeframe"),
            "timeframe_seconds": item.get("timeframe_seconds"),
            "series_id": item.get("series_id"),
            "discovery_source": "3_JSON_FEATURES.series_outputs",
        })

    return files


def parse_feature_path_fallback(path: Path) -> Dict[str, Any]:
    source = None
    asset = None
    timeframe = None

    try:
        rel = path.relative_to(FEATURES_PARQUET_DIR)
        rel_parts = rel.parts
        if len(rel_parts) >= 4:
            source = rel_parts[0]
            asset = rel_parts[1]
            timeframe = rel_parts[2]
    except Exception:
        pass

    name = path.name
    symbol = None
    match = re.match(r"([A-Z0-9]+)_", name)
    if match:
        symbol = match.group(1)

    return {
        "feature_path": path_to_str(path),
        "asset": asset,
        "symbol": symbol,
        "source": source,
        "timeframe": timeframe,
        "timeframe_seconds": infer_timeframe_seconds(timeframe),
        "series_id": None,
        "discovery_source": "directory_fallback",
    }


def discover_feature_files(features_json: Dict[str, Any]) -> List[Dict[str, Any]]:
    discovered: List[Dict[str, Any]] = []

    if DISCOVER_FEATURE_FILES_FROM_FEATURES_JSON:
        discovered.extend(extract_feature_files_from_features_json(features_json))

    existing_paths = {x["feature_path"] for x in discovered}

    if DISCOVER_FEATURE_FILES_FROM_DIRECTORY_FALLBACK and FEATURES_PARQUET_DIR.exists():
        for path in FEATURES_PARQUET_DIR.rglob("*_features.parquet"):
            p = path_to_str(path)
            if p in existing_paths:
                continue
            discovered.append(parse_feature_path_fallback(path))
            existing_paths.add(p)

    clean: List[Dict[str, Any]] = []

    for item in discovered:
        asset_raw = item.get("asset")
        source_raw = item.get("source")
        timeframe_raw = item.get("timeframe")

        asset = str(asset_raw or "").upper()
        source = str(source_raw or "")
        timeframe = str(timeframe_raw or "")

        if SKIP_UNKNOWN_METADATA_FEATURE_FILES:
            if not asset or asset in {"UNKNOWN", "NONE", "NAN"}:
                continue
            if not source or source in {"unknown_source", "UNKNOWN", "None"}:
                continue
            if not timeframe or timeframe in {"unknown_timeframe", "UNKNOWN", "None"}:
                continue

        if FILTER_ASSETS is not None and asset not in FILTER_ASSETS:
            continue
        if FILTER_SOURCES is not None and source not in FILTER_SOURCES:
            continue
        if FILTER_TIMEFRAMES is not None and timeframe not in FILTER_TIMEFRAMES:
            continue

        clean.append(item)

    clean = sorted(
        clean,
        key=lambda x: (
            str(x.get("source", "")),
            str(x.get("asset", "")),
            str(x.get("symbol", "")),
            str(x.get("timeframe", "")),
            str(x.get("feature_path", "")),
        ),
    )

    return clean


# =============================================================================
# 7. LEITURA E OUTPUT
# =============================================================================

def build_label_output_path(item: Dict[str, Any]) -> Path:
    source = safe_filename_token(item.get("source") or "unknown_source")
    asset = safe_filename_token(item.get("asset") or "UNKNOWN")
    timeframe = safe_filename_token(item.get("timeframe") or "unknown_timeframe")
    symbol = safe_filename_token(item.get("symbol") or asset)

    feature_path = str(item.get("feature_path", ""))
    short_hash = hashlib.sha256(feature_path.encode("utf-8", errors="ignore")).hexdigest()[:10]

    out_dir = LABELS_PARQUET_DIR / source / asset / timeframe
    ensure_dir(out_dir)

    return out_dir / f"{symbol}_{source}_{timeframe}_{short_hash}_labels.parquet"


def required_columns_for_labeling() -> List[str]:
    base = [
        "DateTime", "timestamp_utc_ms", "Open", "High", "Low", "Close", "Volume",
        "meta_asset", "meta_symbol", "meta_source", "meta_timeframe", "meta_series_id",
    ]

    optional = [
        "feat_rv_std_20", "feat_rv_std_50", "feat_rv_std_100",
        "feat_atr_14_ratio", "feat_atr_21_ratio", "feat_atr_50_ratio",
        "feat_dollar_volume", "feat_dollar_volume_mean_20", "feat_dollar_volume_mean_50",
        "feat_amihud_illiq_20", "feat_amihud_illiq_50",
        "feat_rv_std_20_ann", "feat_rv_std_50_ann", "feat_rv_std_100_ann",
    ]

    return base + optional


def read_feature_file(path: Path) -> pd.DataFrame:
    if READ_ONLY_NEEDED_COLUMNS:
        try:
            import pyarrow.parquet as pq
            schema_cols = pq.read_schema(path).names
            cols = [c for c in required_columns_for_labeling() if c in schema_cols]

            for c in ["DateTime", "Open", "High", "Low", "Close", "Volume"]:
                if c not in cols and c in schema_cols:
                    cols.append(c)

            return pd.read_parquet(path, columns=cols, engine=PARQUET_ENGINE)
        except Exception:
            pass

    return pd.read_parquet(path, engine=PARQUET_ENGINE)


def validate_feature_df(df: pd.DataFrame) -> None:
    required = ["DateTime", "Open", "High", "Low", "Close", "Volume"]
    missing = [c for c in required if c not in df.columns]

    if missing:
        raise ValueError(f"Arquivo de features sem colunas obrigatórias: {missing}")

    if len(df) < MIN_ROWS:
        raise ValueError(f"Arquivo com poucas linhas: {len(df)} < {MIN_ROWS}")


# =============================================================================
# 8. LABELS: FORWARD RETURNS
# =============================================================================

def compute_forward_returns(
    df: pd.DataFrame,
    config: Dict[str, Any],
    source: str,
    timeframe: str,
) -> Tuple[pd.DataFrame, List[Dict[str, Any]]]:

    close = df["Close"].astype(float)
    cost_cfg = get_cost_config(source)

    fee = bps_to_decimal(cost_cfg["fee_bps_round_trip"])
    slippage = bps_to_decimal(cost_cfg["slippage_bps_round_trip"])

    tf_seconds = infer_timeframe_seconds(timeframe)
    seconds_per_day = 24 * 3600

    out = pd.DataFrame(index=df.index)
    meta: List[Dict[str, Any]] = []

    for h in config.get("horizons", []):
        h = int(h)
        future_close = close.shift(-h)

        gross_long = safe_div(future_close, close) - 1.0
        gross_short = 1.0 - safe_div(future_close, close)

        funding_cost = 0.0
        if tf_seconds is not None:
            holding_days = (h * tf_seconds) / seconds_per_day
            funding_cost = bps_to_decimal(cost_cfg["funding_bps_per_day"]) * holding_days

        total_cost = fee + slippage + funding_cost

        out[f"label_fwd_ret_gross_long_h{h}"] = gross_long
        out[f"label_fwd_ret_net_long_h{h}"] = gross_long - total_cost
        out[f"label_fwd_ret_gross_short_h{h}"] = gross_short
        out[f"label_fwd_ret_net_short_h{h}"] = gross_short - total_cost

        out[f"label_fwd_ret_gross_long_bps_h{h}"] = decimal_to_bps(gross_long)
        out[f"label_fwd_ret_net_long_bps_h{h}"] = decimal_to_bps(gross_long - total_cost)
        out[f"label_fwd_ret_gross_short_bps_h{h}"] = decimal_to_bps(gross_short)
        out[f"label_fwd_ret_net_short_bps_h{h}"] = decimal_to_bps(gross_short - total_cost)

        out[f"label_fwd_logret_h{h}"] = np.log(future_close / close)

        for side in ["long", "short"]:
            meta.append({
                "label_name": f"label_fwd_ret_net_{side}_h{h}",
                "label_config_id": config["label_config_id"],
                "label_type": "forward_return",
                "side": side,
                "horizon_bars": h,
                "uses_future_data": True,
            })

    return out, meta


# =============================================================================
# 9. LABELS: DIRECTIONAL THRESHOLD
# =============================================================================

def compute_directional_labels(
    df: pd.DataFrame,
    config: Dict[str, Any],
    source: str,
    timeframe: str,
) -> Tuple[pd.DataFrame, List[Dict[str, Any]]]:

    close = df["Close"].astype(float)
    cost_cfg = get_cost_config(source)

    fee = bps_to_decimal(cost_cfg["fee_bps_round_trip"])
    slippage = bps_to_decimal(cost_cfg["slippage_bps_round_trip"])

    out = pd.DataFrame(index=df.index)
    meta: List[Dict[str, Any]] = []

    thresholds = config.get("threshold_bps_list")
    if thresholds is None:
        thresholds = [float(config.get("threshold_bps", 15.0))]

    for h in config.get("horizons", []):
        h = int(h)
        future_close = close.shift(-h)
        gross_ret = safe_div(future_close, close) - 1.0
        net_ret = gross_ret - fee - slippage

        out[f"label_dir_net_ret_bps_h{h}"] = decimal_to_bps(net_ret)

        for thr_bps in thresholds:
            thr_bps = float(thr_bps)
            thr = bps_to_decimal(thr_bps)
            thr_token = int(thr_bps)

            label = pd.Series(0, index=df.index, dtype="float64")
            label[net_ret > thr] = 1.0
            label[net_ret < -thr] = -1.0

            col = f"label_dir_h{h}_thr{thr_token}bps"
            out[col] = label

            meta.append({
                "label_name": col,
                "label_config_id": config["label_config_id"],
                "label_type": "directional_threshold",
                "side": "both",
                "horizon_bars": h,
                "threshold_bps": thr_bps,
                "uses_future_data": True,
            })

    return out, meta


# =============================================================================
# 10. LABELS: QUANTILE FORWARD RETURN
# =============================================================================

def compute_quantile_forward_return_labels(
    df: pd.DataFrame,
    config: Dict[str, Any],
    source: str,
    timeframe: str,
) -> Tuple[pd.DataFrame, List[Dict[str, Any]]]:

    close = df["Close"].astype(float)
    out = pd.DataFrame(index=df.index)
    meta: List[Dict[str, Any]] = []

    window = int(config.get("rolling_window", 1000))
    quantiles = config.get("quantiles", [0.10, 0.25, 0.50, 0.75, 0.90])

    for h in config.get("horizons", []):
        h = int(h)
        fwd = close.shift(-h) / close - 1.0
        out[f"label_fwd_ret_for_quantile_h{h}"] = fwd

        q_values = {}
        for q in quantiles:
            q_col = f"label_fwd_ret_h{h}_rolling_q{int(q * 100)}"
            out[q_col] = fwd.rolling(window, min_periods=max(50, int(window * 0.2))).quantile(q)
            q_values[q] = out[q_col]

        bucket_col = f"label_fwd_ret_quantile_bucket_h{h}"
        bucket = pd.Series(np.nan, index=df.index, dtype="float64")

        if 0.10 in q_values:
            bucket[fwd <= q_values[0.10]] = -2.0
        if 0.25 in q_values:
            bucket[(fwd > q_values.get(0.10, -np.inf)) & (fwd <= q_values[0.25])] = -1.0
        if 0.75 in q_values:
            bucket[(fwd >= q_values.get(0.75)) & (fwd < q_values.get(0.90, np.inf))] = 1.0
        if 0.90 in q_values:
            bucket[fwd >= q_values[0.90]] = 2.0

        bucket[bucket.isna() & fwd.notna()] = 0.0
        out[bucket_col] = bucket

        meta.append({
            "label_name": bucket_col,
            "label_config_id": config["label_config_id"],
            "label_type": "quantile_forward_return",
            "horizon_bars": h,
            "rolling_window": window,
            "uses_future_data": True,
        })

    return out, meta


# =============================================================================
# 11. LABELS: FUTURE PATH EXTREMES
# =============================================================================

def rolling_future_max(series: pd.Series, horizon: int) -> pd.Series:
    return series[::-1].rolling(horizon, min_periods=1).max()[::-1].shift(-1)


def rolling_future_min(series: pd.Series, horizon: int) -> pd.Series:
    return series[::-1].rolling(horizon, min_periods=1).min()[::-1].shift(-1)


def compute_future_path_extremes(
    df: pd.DataFrame,
    config: Dict[str, Any],
    source: str,
    timeframe: str,
) -> Tuple[pd.DataFrame, List[Dict[str, Any]]]:

    close = df["Close"].astype(float)
    high = df["High"].astype(float)
    low = df["Low"].astype(float)

    out = pd.DataFrame(index=df.index)
    meta: List[Dict[str, Any]] = []

    for h in config.get("horizons", []):
        h = int(h)

        future_high = rolling_future_max(high, h)
        future_low = rolling_future_min(low, h)
        future_close = close.shift(-h)

        mfe_long = future_high / close - 1.0
        mae_long = future_low / close - 1.0

        mfe_short = 1.0 - safe_div(future_low, close)
        mae_short = 1.0 - safe_div(future_high, close)

        close_ret = future_close / close - 1.0

        out[f"label_future_mfe_long_h{h}"] = mfe_long
        out[f"label_future_mae_long_h{h}"] = mae_long
        out[f"label_future_mfe_short_h{h}"] = mfe_short
        out[f"label_future_mae_short_h{h}"] = mae_short

        out[f"label_future_mfe_long_bps_h{h}"] = decimal_to_bps(mfe_long)
        out[f"label_future_mae_long_bps_h{h}"] = decimal_to_bps(mae_long)
        out[f"label_future_mfe_short_bps_h{h}"] = decimal_to_bps(mfe_short)
        out[f"label_future_mae_short_bps_h{h}"] = decimal_to_bps(mae_short)

        out[f"label_future_close_ret_bps_h{h}"] = decimal_to_bps(close_ret)
        out[f"label_future_range_bps_h{h}"] = decimal_to_bps(future_high / future_low - 1.0)

        meta.append({
            "label_name": f"label_future_mfe_long_h{h}",
            "label_config_id": config["label_config_id"],
            "label_type": "future_path_extremes",
            "horizon_bars": h,
            "uses_future_data": True,
        })

    return out, meta


# =============================================================================
# 12. LABELS: FUTURE REALIZED VOL
# =============================================================================

def compute_future_realized_vol(
    df: pd.DataFrame,
    config: Dict[str, Any],
    source: str,
    timeframe: str,
) -> Tuple[pd.DataFrame, List[Dict[str, Any]]]:

    close = df["Close"].astype(float)
    ret = np.log(close).diff()

    out = pd.DataFrame(index=df.index)
    meta: List[Dict[str, Any]] = []

    af = annualization_factor_from_timeframe(timeframe)
    window = int(config.get("rolling_window", 1000))

    for h in config.get("horizons", []):
        h = int(h)

        future_rv = ret.shift(-1)[::-1].rolling(h, min_periods=max(2, int(h / 3))).std()[::-1]
        future_rv_ann = future_rv * math.sqrt(af)

        out[f"label_future_rv_h{h}"] = future_rv
        out[f"label_future_rv_ann_h{h}"] = future_rv_ann
        out[f"label_future_rv_bps_h{h}"] = decimal_to_bps(future_rv)

        q75 = future_rv.rolling(window, min_periods=max(50, int(window * 0.2))).quantile(0.75)
        q90 = future_rv.rolling(window, min_periods=max(50, int(window * 0.2))).quantile(0.90)

        regime = pd.Series(0, index=df.index, dtype="float64")
        regime[future_rv > q75] = 1.0
        regime[future_rv > q90] = 2.0

        out[f"label_future_vol_regime_h{h}"] = regime

        meta.append({
            "label_name": f"label_future_vol_regime_h{h}",
            "label_config_id": config["label_config_id"],
            "label_type": "future_realized_vol",
            "horizon_bars": h,
            "uses_future_data": True,
        })

    return out, meta


# =============================================================================
# 13. LABELS: FUTURE TREND REGIME
# =============================================================================

def compute_trend_regime_forward(
    df: pd.DataFrame,
    config: Dict[str, Any],
    source: str,
    timeframe: str,
) -> Tuple[pd.DataFrame, List[Dict[str, Any]]]:

    close = df["Close"].astype(float)
    out = pd.DataFrame(index=df.index)
    meta: List[Dict[str, Any]] = []

    thresholds = config.get("trend_bps_thresholds", [25.0, 50.0, 100.0])

    for h in config.get("horizons", []):
        h = int(h)
        fwd = close.shift(-h) / close - 1.0
        fwd_bps = decimal_to_bps(fwd)

        out[f"label_future_trend_ret_bps_h{h}"] = fwd_bps

        for thr_bps in thresholds:
            thr_bps = float(thr_bps)
            token = int(thr_bps)

            regime = pd.Series(0, index=df.index, dtype="float64")
            regime[fwd_bps > thr_bps] = 1.0
            regime[fwd_bps < -thr_bps] = -1.0

            col = f"label_future_trend_regime_h{h}_thr{token}bps"
            out[col] = regime

            meta.append({
                "label_name": col,
                "label_config_id": config["label_config_id"],
                "label_type": "trend_regime_forward",
                "horizon_bars": h,
                "threshold_bps": thr_bps,
                "uses_future_data": True,
            })

    return out, meta


# =============================================================================
# 14. LABELS: RISK/REWARD FORWARD
# =============================================================================

def compute_risk_reward_forward(
    df: pd.DataFrame,
    config: Dict[str, Any],
    source: str,
    timeframe: str,
) -> Tuple[pd.DataFrame, List[Dict[str, Any]]]:

    close = df["Close"].astype(float)
    high = df["High"].astype(float)
    low = df["Low"].astype(float)

    out = pd.DataFrame(index=df.index)
    meta: List[Dict[str, Any]] = []

    rr_thresholds = config.get("rr_thresholds", [1.0, 1.5, 2.0])

    for h in config.get("horizons", []):
        h = int(h)

        future_high = rolling_future_max(high, h)
        future_low = rolling_future_min(low, h)

        mfe = future_high / close - 1.0
        mae_abs = (future_low / close - 1.0).abs()

        rr = mfe / mae_abs.replace(0, np.nan)

        out[f"label_future_rr_long_h{h}"] = rr

        for thr in rr_thresholds:
            thr = float(thr)
            token = str(thr).replace(".", "_")

            col = f"label_future_rr_long_h{h}_gt_{token}"
            out[col] = (rr >= thr).astype("float64")

            meta.append({
                "label_name": col,
                "label_config_id": config["label_config_id"],
                "label_type": "risk_reward_forward",
                "horizon_bars": h,
                "rr_threshold": thr,
                "uses_future_data": True,
            })

    return out, meta


# =============================================================================
# 15. LABELS: TAIL EVENTS
# =============================================================================

def compute_tail_event_forward(
    df: pd.DataFrame,
    config: Dict[str, Any],
    source: str,
    timeframe: str,
) -> Tuple[pd.DataFrame, List[Dict[str, Any]]]:

    close = df["Close"].astype(float)
    high = df["High"].astype(float)
    low = df["Low"].astype(float)

    out = pd.DataFrame(index=df.index)
    meta: List[Dict[str, Any]] = []

    for h in config.get("horizons", []):
        h = int(h)
        future_high = rolling_future_max(high, h)
        future_low = rolling_future_min(low, h)

        up_bps = decimal_to_bps(future_high / close - 1.0)
        down_bps = decimal_to_bps(future_low / close - 1.0)

        out[f"label_future_tail_up_bps_h{h}"] = up_bps
        out[f"label_future_tail_down_bps_h{h}"] = down_bps

        for tail_bps in config.get("tail_bps_list", []):
            tail_bps = float(tail_bps)
            token = int(tail_bps)

            up_col = f"label_tail_up_h{h}_gt_{token}bps"
            down_col = f"label_tail_down_h{h}_lt_{token}bps"

            out[up_col] = (up_bps >= tail_bps).astype("float64")
            out[down_col] = (down_bps <= -tail_bps).astype("float64")

            meta.append({
                "label_name": up_col,
                "label_config_id": config["label_config_id"],
                "label_type": "tail_event_forward",
                "horizon_bars": h,
                "tail_bps": tail_bps,
                "side": "up",
                "uses_future_data": True,
            })

            meta.append({
                "label_name": down_col,
                "label_config_id": config["label_config_id"],
                "label_type": "tail_event_forward",
                "horizon_bars": h,
                "tail_bps": tail_bps,
                "side": "down",
                "uses_future_data": True,
            })

    return out, meta


# =============================================================================
# 16. TRIPLE BARRIER
# =============================================================================

def choose_barrier_series(df: pd.DataFrame, config: Dict[str, Any]) -> Tuple[pd.Series, pd.Series, str]:
    barrier_type = config.get("barrier_type")
    min_barrier = bps_to_decimal(float(config.get("min_barrier_bps", 10.0)))

    if barrier_type == "volatility":
        w = int(config.get("vol_window", 20))
        col = f"feat_rv_std_{w}"

        if col in df.columns:
            base = df[col].astype(float)
            source_col = col
        else:
            close = df["Close"].astype(float)
            ret = np.log(close).diff()
            base = ret.rolling(w, min_periods=w).std()
            source_col = f"computed_rv_std_{w}"

        tp = base * float(config.get("tp_mult", 1.5))
        sl = base * float(config.get("sl_mult", 1.0))

    elif barrier_type == "atr":
        w = int(config.get("atr_window", 14))
        col = f"feat_atr_{w}_ratio"

        if col in df.columns:
            base = df[col].astype(float)
            source_col = col
        else:
            high = df["High"].astype(float)
            low = df["Low"].astype(float)
            close = df["Close"].astype(float)
            prev_close = close.shift(1)

            tr = pd.concat(
                [
                    high - low,
                    (high - prev_close).abs(),
                    (low - prev_close).abs(),
                ],
                axis=1,
            ).max(axis=1)

            base = tr.rolling(w, min_periods=w).mean() / close
            source_col = f"computed_atr_{w}_ratio"

        tp = base * float(config.get("tp_mult", 1.5))
        sl = base * float(config.get("sl_mult", 1.0))

    elif barrier_type == "fixed_bps":
        tp = pd.Series(bps_to_decimal(float(config.get("tp_bps", 40.0))), index=df.index)
        sl = pd.Series(bps_to_decimal(float(config.get("sl_bps", 25.0))), index=df.index)
        source_col = "fixed_bps"

    else:
        raise ValueError(f"barrier_type inválido: {barrier_type}")

    tp = tp.clip(lower=min_barrier)
    sl = sl.clip(lower=min_barrier)

    return tp, sl, source_col


def triple_barrier_loop(
    close: np.ndarray,
    high: np.ndarray,
    low: np.ndarray,
    datetimes: np.ndarray,
    tp_arr: np.ndarray,
    sl_arr: np.ndarray,
    horizon: int,
    side: str,
    total_cost_decimal: float,
) -> Dict[str, np.ndarray]:

    n = len(close)

    label = np.full(n, np.nan, dtype="float64")
    event_type = np.full(n, "", dtype=object)
    gross_ret = np.full(n, np.nan, dtype="float64")
    net_ret = np.full(n, np.nan, dtype="float64")
    holding_bars = np.full(n, np.nan, dtype="float64")
    exit_price = np.full(n, np.nan, dtype="float64")
    exit_datetime = np.full(n, pd.NaT, dtype="datetime64[ns]")
    mfe = np.full(n, np.nan, dtype="float64")
    mae = np.full(n, np.nan, dtype="float64")
    tp_level = np.full(n, np.nan, dtype="float64")
    sl_level = np.full(n, np.nan, dtype="float64")

    for i in range(n):
        entry = close[i]

        if not np.isfinite(entry) or entry <= 0:
            continue

        tp = tp_arr[i]
        sl = sl_arr[i]

        if not np.isfinite(tp) or not np.isfinite(sl) or tp <= 0 or sl <= 0:
            continue

        end = min(i + horizon, n - 1)

        if end <= i:
            continue

        if side == "long":
            upper = entry * (1.0 + tp)
            lower = entry * (1.0 - sl)
        else:
            upper = entry * (1.0 + sl)
            lower = entry * (1.0 - tp)

        tp_level[i] = upper if side == "long" else lower
        sl_level[i] = lower if side == "long" else upper

        path_high = high[i + 1:end + 1]
        path_low = low[i + 1:end + 1]
        path_close = close[i + 1:end + 1]

        if len(path_close) == 0:
            continue

        if side == "long":
            rel_high = path_high / entry - 1.0
            rel_low = path_low / entry - 1.0

            mfe[i] = np.nanmax(rel_high)
            mae[i] = np.nanmin(rel_low)

            exit_idx = None
            exit_ev = "time"

            for j in range(len(path_close)):
                if path_low[j] <= lower:
                    exit_idx = i + 1 + j
                    exit_ev = "stop_loss"
                    break
                if path_high[j] >= upper:
                    exit_idx = i + 1 + j
                    exit_ev = "take_profit"
                    break

            if exit_idx is None:
                exit_idx = end

            px = close[exit_idx]

            if exit_ev == "take_profit":
                px = upper
                y = 1.0
            elif exit_ev == "stop_loss":
                px = lower
                y = -1.0
            else:
                y = 0.0

            gr = px / entry - 1.0

        else:
            rel_low_short = 1.0 - safe_div(path_low, entry)
            rel_high_short = 1.0 - safe_div(path_high, entry)

            mfe[i] = np.nanmax(rel_low_short)
            mae[i] = np.nanmin(rel_high_short)

            exit_idx = None
            exit_ev = "time"

            for j in range(len(path_close)):
                if path_high[j] >= upper:
                    exit_idx = i + 1 + j
                    exit_ev = "stop_loss"
                    break
                if path_low[j] <= lower:
                    exit_idx = i + 1 + j
                    exit_ev = "take_profit"
                    break

            if exit_idx is None:
                exit_idx = end

            px = close[exit_idx]

            if exit_ev == "take_profit":
                px = lower
                y = 1.0
            elif exit_ev == "stop_loss":
                px = upper
                y = -1.0
            else:
                y = 0.0

            gr = 1.0 - safe_div(px, entry)

        label[i] = y
        event_type[i] = exit_ev
        gross_ret[i] = gr
        net_ret[i] = gr - total_cost_decimal
        holding_bars[i] = exit_idx - i
        exit_price[i] = px
        exit_datetime[i] = datetimes[exit_idx]

    return {
        "label": label,
        "event_type": event_type,
        "gross_return": gross_ret,
        "net_return": net_ret,
        "holding_bars": holding_bars,
        "exit_price": exit_price,
        "exit_datetime": exit_datetime,
        "mfe": mfe,
        "mae": mae,
        "tp_level": tp_level,
        "sl_level": sl_level,
    }


def compute_single_triple_barrier(
    df: pd.DataFrame,
    config: Dict[str, Any],
    source: str,
    timeframe: str,
) -> Tuple[pd.DataFrame, List[Dict[str, Any]]]:

    close = df["Close"].astype(float).to_numpy()
    high = df["High"].astype(float).to_numpy()
    low = df["Low"].astype(float).to_numpy()
    datetimes = pd.to_datetime(df["DateTime"]).to_numpy(dtype="datetime64[ns]")

    horizon = int(config["horizon_bars"])
    tp_series, sl_series, barrier_source = choose_barrier_series(df, config)

    cost_cfg = get_cost_config(source)
    fee = bps_to_decimal(cost_cfg["fee_bps_round_trip"])
    slippage = bps_to_decimal(cost_cfg["slippage_bps_round_trip"])

    tf_seconds = infer_timeframe_seconds(timeframe)
    funding_cost = 0.0

    if tf_seconds is not None:
        holding_days_est = (horizon * tf_seconds) / (24 * 3600)
        funding_cost = bps_to_decimal(cost_cfg["funding_bps_per_day"]) * holding_days_est

    total_cost = fee + slippage + funding_cost

    tp_arr = tp_series.to_numpy(dtype="float64")
    sl_arr = sl_series.to_numpy(dtype="float64")

    config_id = config["label_config_id"]

    out = pd.DataFrame(index=df.index)
    meta: List[Dict[str, Any]] = []

    for side in ["long", "short"]:
        res = triple_barrier_loop(
            close=close,
            high=high,
            low=low,
            datetimes=datetimes,
            tp_arr=tp_arr,
            sl_arr=sl_arr,
            horizon=horizon,
            side=side,
            total_cost_decimal=total_cost,
        )

        prefix = f"label_{config_id}_{side}"

        out[f"{prefix}_tb"] = res["label"]
        out[f"{prefix}_event"] = res["event_type"]
        out[f"{prefix}_gross_return"] = res["gross_return"]
        out[f"{prefix}_net_return"] = res["net_return"]
        out[f"{prefix}_gross_return_bps"] = decimal_to_bps(res["gross_return"])
        out[f"{prefix}_net_return_bps"] = decimal_to_bps(res["net_return"])
        out[f"{prefix}_holding_bars"] = res["holding_bars"]

        # Meta-label: 1 se TB foi positivo e líquido > 0; 0 caso contrário.
        out[f"{prefix}_meta_success_net"] = (
            (pd.Series(res["label"], index=df.index) == 1.0)
            & (pd.Series(res["net_return"], index=df.index) > 0)
        ).astype("float64")

        if SAVE_RAW_LABEL_COMPONENTS:
            out[f"{prefix}_exit_price"] = res["exit_price"]
            out[f"{prefix}_exit_datetime"] = res["exit_datetime"]
            out[f"{prefix}_mfe"] = res["mfe"]
            out[f"{prefix}_mae"] = res["mae"]
            out[f"{prefix}_mfe_bps"] = decimal_to_bps(res["mfe"])
            out[f"{prefix}_mae_bps"] = decimal_to_bps(res["mae"])
            out[f"{prefix}_tp_level"] = res["tp_level"]
            out[f"{prefix}_sl_level"] = res["sl_level"]

        out[f"{prefix}_tp_barrier_bps"] = decimal_to_bps(tp_series)
        out[f"{prefix}_sl_barrier_bps"] = decimal_to_bps(sl_series)

        meta.append({
            "label_name": f"{prefix}_tb",
            "label_config_id": config_id,
            "label_type": "triple_barrier",
            "side": side,
            "barrier_type": config.get("barrier_type"),
            "barrier_source": barrier_source,
            "horizon_bars": horizon,
            "tp_mult": config.get("tp_mult"),
            "sl_mult": config.get("sl_mult"),
            "tp_bps": config.get("tp_bps"),
            "sl_bps": config.get("sl_bps"),
            "min_barrier_bps": config.get("min_barrier_bps"),
            "estimated_total_cost_bps": total_cost * 10000.0,
            "uses_future_data": True,
        })

    return out, meta


def expand_triple_barrier_grid(config: Dict[str, Any]) -> List[Dict[str, Any]]:
    expanded: List[Dict[str, Any]] = []

    barrier_type = config.get("barrier_type")
    base_id = config["label_config_id"]

    if barrier_type == "volatility":
        for w in config.get("vol_windows", [20]):
            for h in config.get("horizon_bars_list", [16]):
                for tp in config.get("tp_mult_list", [1.5]):
                    for sl in config.get("sl_mult_list", [1.0]):
                        expanded.append({
                            "label_config_id": f"{base_id}_VOL{w}_TP{str(tp).replace('.', '')}_SL{str(sl).replace('.', '')}_H{h}",
                            "label_type": "triple_barrier",
                            "enabled": True,
                            "barrier_type": "volatility",
                            "vol_window": int(w),
                            "tp_mult": float(tp),
                            "sl_mult": float(sl),
                            "horizon_bars": int(h),
                            "min_barrier_bps": config.get("min_barrier_bps", 10.0),
                            "description": config.get("description"),
                        })

    elif barrier_type == "atr":
        for w in config.get("atr_windows", [14]):
            for h in config.get("horizon_bars_list", [16]):
                for tp in config.get("tp_mult_list", [1.5]):
                    for sl in config.get("sl_mult_list", [1.0]):
                        expanded.append({
                            "label_config_id": f"{base_id}_ATR{w}_TP{str(tp).replace('.', '')}_SL{str(sl).replace('.', '')}_H{h}",
                            "label_type": "triple_barrier",
                            "enabled": True,
                            "barrier_type": "atr",
                            "atr_window": int(w),
                            "tp_mult": float(tp),
                            "sl_mult": float(sl),
                            "horizon_bars": int(h),
                            "min_barrier_bps": config.get("min_barrier_bps", 10.0),
                            "description": config.get("description"),
                        })

    elif barrier_type == "fixed_bps":
        for h in config.get("horizon_bars_list", [20]):
            for tp_bps in config.get("tp_bps_list", [40.0]):
                for sl_bps in config.get("sl_bps_list", [25.0]):
                    if float(tp_bps) <= float(sl_bps) * 0.5:
                        continue
                    expanded.append({
                        "label_config_id": f"{base_id}_TP{int(tp_bps)}bps_SL{int(sl_bps)}bps_H{h}",
                        "label_type": "triple_barrier",
                        "enabled": True,
                        "barrier_type": "fixed_bps",
                        "tp_bps": float(tp_bps),
                        "sl_bps": float(sl_bps),
                        "horizon_bars": int(h),
                        "min_barrier_bps": config.get("min_barrier_bps", 10.0),
                        "description": config.get("description"),
                    })

    return expanded


def compute_triple_barrier_grid(
    df: pd.DataFrame,
    config: Dict[str, Any],
    source: str,
    timeframe: str,
) -> Tuple[pd.DataFrame, List[Dict[str, Any]]]:

    frames: List[pd.DataFrame] = []
    meta_all: List[Dict[str, Any]] = []

    expanded_configs = expand_triple_barrier_grid(config)

    for cfg in expanded_configs:
        labels, meta = compute_single_triple_barrier(df, cfg, source, timeframe)
        frames.append(labels)
        meta_all.extend(meta)

    if frames:
        return pd.concat(frames, axis=1), meta_all

    return pd.DataFrame(index=df.index), meta_all


# =============================================================================
# 17. GERAÇÃO DE LABELS POR SÉRIE
# =============================================================================

def generate_labels_for_df(
    df: pd.DataFrame,
    source: str,
    timeframe: str,
) -> Tuple[pd.DataFrame, List[Dict[str, Any]], List[Dict[str, Any]]]:

    label_frames: List[pd.DataFrame] = []
    label_metadata: List[Dict[str, Any]] = []
    execution_steps: List[Dict[str, Any]] = []

    for config in LABEL_CONFIGS:
        if not config.get("enabled", False):
            continue

        start = time.time()
        label_type = config.get("label_type")
        status = "OK"

        try:
            if label_type == "forward_returns":
                labels, meta = compute_forward_returns(df, config, source, timeframe)

            elif label_type == "directional_threshold":
                labels, meta = compute_directional_labels(df, config, source, timeframe)

            elif label_type == "quantile_forward_return":
                labels, meta = compute_quantile_forward_return_labels(df, config, source, timeframe)

            elif label_type == "future_path_extremes":
                labels, meta = compute_future_path_extremes(df, config, source, timeframe)

            elif label_type == "future_realized_vol":
                labels, meta = compute_future_realized_vol(df, config, source, timeframe)

            elif label_type == "trend_regime_forward":
                labels, meta = compute_trend_regime_forward(df, config, source, timeframe)

            elif label_type == "risk_reward_forward":
                labels, meta = compute_risk_reward_forward(df, config, source, timeframe)

            elif label_type == "tail_event_forward":
                labels, meta = compute_tail_event_forward(df, config, source, timeframe)

            elif label_type == "triple_barrier":
                labels, meta = compute_single_triple_barrier(df, config, source, timeframe)

            elif label_type == "triple_barrier_grid":
                labels, meta = compute_triple_barrier_grid(df, config, source, timeframe)

            elif label_type in {"relative_to_benchmark", "cross_sectional_rank"}:
                # Estes são criados em etapa posterior, painel cross-asset.
                labels = pd.DataFrame(index=df.index)
                meta = []

            else:
                raise ValueError(f"label_type desconhecido: {label_type}")

            if labels is not None and len(labels.columns) > 0:
                label_frames.append(labels)

            label_metadata.extend(meta)

        except Exception as exc:
            status = "ERROR"
            execution_steps.append({
                "label_config_id": config.get("label_config_id"),
                "label_type": label_type,
                "status": status,
                "error": str(exc),
                "traceback": traceback.format_exc(),
                "elapsed_seconds": round(time.time() - start, 6),
            })
            continue

        execution_steps.append({
            "label_config_id": config.get("label_config_id"),
            "label_type": label_type,
            "status": status,
            "elapsed_seconds": round(time.time() - start, 6),
        })

    if label_frames:
        labels_df = pd.concat(label_frames, axis=1)
    else:
        labels_df = pd.DataFrame(index=df.index)

    labels_df = labels_df.replace([np.inf, -np.inf], np.nan)

    return labels_df, label_metadata, execution_steps


# =============================================================================
# 18. AUDITORIA
# =============================================================================

def compute_label_distribution(labels_df: pd.DataFrame) -> Dict[str, Any]:
    dist: Dict[str, Any] = {}

    class_cols = [
        c for c in labels_df.columns
        if c.endswith("_tb")
        or c.startswith("label_dir_")
        or "regime" in c
        or "bucket" in c
        or "meta_success" in c
        or c.startswith("label_tail_")
    ]

    for col in class_cols:
        s = labels_df[col]
        vc = s.value_counts(dropna=False).to_dict()

        dist[col] = {
            "nan_count": int(s.isna().sum()),
            "class_counts": {str(k): int(v) for k, v in vc.items()},
            "class_ratio_pos": None if len(s.dropna()) == 0 else float((s == 1).sum() / max(1, s.notna().sum())),
            "class_ratio_neg": None if len(s.dropna()) == 0 else float((s == -1).sum() / max(1, s.notna().sum())),
            "class_ratio_zero": None if len(s.dropna()) == 0 else float((s == 0).sum() / max(1, s.notna().sum())),
        }

    return dist


def audit_labels_output(
    feature_df: pd.DataFrame,
    labels_df: pd.DataFrame,
    output_df: pd.DataFrame,
    timeframe: str,
) -> Dict[str, Any]:

    audit = {
        "audit_status": "PENDING",
        "checks": {},
        "warnings": [],
        "errors": [],
    }

    try:
        dt = pd.to_datetime(feature_df["DateTime"])
        out_dt = pd.to_datetime(output_df["DateTime"])

        audit["checks"]["feature_rows"] = int(len(feature_df))
        audit["checks"]["labels_rows"] = int(len(labels_df))
        audit["checks"]["output_rows"] = int(len(output_df))
        audit["checks"]["row_count_match"] = bool(len(feature_df) == len(labels_df) == len(output_df))

        audit["checks"]["datetime_exact_match"] = bool(
            len(dt) == len(out_dt) and np.array_equal(dt.to_numpy(), out_dt.to_numpy())
        )

        audit["checks"]["datetime_monotonic"] = bool(out_dt.is_monotonic_increasing)
        audit["checks"]["datetime_duplicates"] = int(out_dt.duplicated().sum())

        tf_seconds = infer_timeframe_seconds(timeframe)

        if tf_seconds and len(out_dt) > 1:
            deltas = out_dt.diff().dropna().dt.total_seconds()
            audit["checks"]["expected_timeframe_seconds"] = int(tf_seconds)
            audit["checks"]["median_delta_seconds"] = float(deltas.median()) if not deltas.empty else None
            audit["checks"]["max_delta_seconds"] = float(deltas.max()) if not deltas.empty else None
            audit["checks"]["large_gap_count"] = int((deltas > tf_seconds * 3.5).sum())
        else:
            audit["checks"]["expected_timeframe_seconds"] = tf_seconds
            audit["checks"]["median_delta_seconds"] = None
            audit["checks"]["max_delta_seconds"] = None
            audit["checks"]["large_gap_count"] = None

        label_cols = [c for c in output_df.columns if str(c).startswith("label_")]
        audit["checks"]["label_columns_count"] = int(len(label_cols))

        if label_cols:
            total_values = len(output_df) * len(label_cols)
            total_nulls = int(output_df[label_cols].isna().sum().sum())
            audit["checks"]["label_null_ratio"] = round(total_nulls / max(1, total_values), 8)
        else:
            audit["checks"]["label_null_ratio"] = None

        audit["checks"]["label_distribution"] = compute_label_distribution(labels_df)

        if not audit["checks"]["row_count_match"]:
            audit["errors"].append("Row count mismatch between features, labels and output.")

        if not audit["checks"]["datetime_exact_match"]:
            audit["errors"].append("DateTime mismatch between feature input and label output.")

        if not audit["checks"]["datetime_monotonic"]:
            audit["errors"].append("Output DateTime is not monotonic increasing.")

        if audit["checks"]["datetime_duplicates"] > 0:
            audit["errors"].append("Output DateTime contains duplicates.")

        if audit["checks"]["label_columns_count"] == 0:
            audit["errors"].append("No label columns generated.")

        if audit["checks"].get("large_gap_count") not in [None, 0]:
            audit["warnings"].append("Large DateTime gaps detected.")

        if audit["errors"]:
            audit["audit_status"] = "FAIL"
        elif audit["warnings"]:
            audit["audit_status"] = "WARNING"
        else:
            audit["audit_status"] = "PASS"

        return audit

    except Exception as exc:
        audit["audit_status"] = "ERROR"
        audit["errors"].append(str(exc))
        audit["traceback"] = traceback.format_exc()
        return audit


def add_label_metadata_columns(base_df: pd.DataFrame, item: Dict[str, Any]) -> pd.DataFrame:
    df = pd.DataFrame(index=base_df.index)

    df["DateTime"] = base_df["DateTime"]
    if "timestamp_utc_ms" in base_df.columns:
        df["timestamp_utc_ms"] = base_df["timestamp_utc_ms"]

    df["meta_asset"] = item.get("asset") or base_df.get("meta_asset", pd.Series(["UNKNOWN"] * len(base_df))).iloc[0]
    df["meta_symbol"] = item.get("symbol") or base_df.get("meta_symbol", pd.Series(["UNKNOWN"] * len(base_df))).iloc[0]
    df["meta_source"] = item.get("source") or base_df.get("meta_source", pd.Series(["unknown_source"] * len(base_df))).iloc[0]
    df["meta_timeframe"] = item.get("timeframe") or base_df.get("meta_timeframe", pd.Series(["unknown_timeframe"] * len(base_df))).iloc[0]

    if "meta_series_id" in base_df.columns and len(base_df) > 0:
        default_series_id = base_df["meta_series_id"].iloc[0]
    else:
        default_series_id = None

    df["meta_series_id"] = item.get("series_id") or default_series_id
    df["meta_label_schema_version"] = SCHEMA_VERSION
    df["meta_label_run_id"] = RUN_ID
    df["meta_label_generated_at"] = now_iso()
    df["meta_uses_future_data"] = True

    return df


# =============================================================================
# 19. PROCESSAMENTO DE UMA SÉRIE
# =============================================================================

def process_one_feature_file(item: Dict[str, Any]) -> Dict[str, Any]:
    start = time.time()

    feature_path = Path(str(item["feature_path"]))
    output_path = build_label_output_path(item)

    source = str(item.get("source") or "unknown_source")
    timeframe = str(item.get("timeframe") or "unknown_timeframe")
    asset = str(item.get("asset") or "UNKNOWN")
    symbol = str(item.get("symbol") or "UNKNOWN")

    result = {
        "run_id": RUN_ID,
        "status": "PENDING",
        "asset": asset,
        "symbol": symbol,
        "source": source,
        "timeframe": timeframe,
        "timeframe_seconds": infer_timeframe_seconds(timeframe),
        "series_id": item.get("series_id"),
        "feature_path": path_to_str(feature_path),
        "output_path": path_to_str(output_path),
        "discovery_source": item.get("discovery_source"),
        "memory_mb_start": get_process_memory_mb(),
    }

    try:
        if not feature_path.exists():
            raise FileNotFoundError(f"Feature file não encontrado: {feature_path}")

        if output_path.exists() and not OVERWRITE_EXISTING:
            result["status"] = "SKIPPED_EXISTS"
            result["elapsed_seconds"] = round(time.time() - start, 6)
            return result

        df_features = read_feature_file(feature_path)
        df_features = normalize_datetime_column(df_features)
        validate_feature_df(df_features)
        result["memory_mb_after_read"] = get_process_memory_mb()

        result["input_rows"] = int(len(df_features))
        result["input_columns"] = int(len(df_features.columns))
        result["input_first_datetime"] = str(df_features["DateTime"].min())
        result["input_last_datetime"] = str(df_features["DateTime"].max())

        labels_df, label_metadata, execution_steps = generate_labels_for_df(
            df=df_features,
            source=source,
            timeframe=timeframe,
        )
        result["memory_mb_after_label_generation"] = get_process_memory_mb()

        meta_df = add_label_metadata_columns(df_features, item)

        output_df = pd.concat([meta_df, labels_df], axis=1)
        output_df = output_df.replace([np.inf, -np.inf], np.nan)
        result["memory_mb_after_output_concat"] = get_process_memory_mb()

        audit = audit_labels_output(
            feature_df=df_features,
            labels_df=labels_df,
            output_df=output_df,
            timeframe=timeframe,
        )
        result["memory_mb_after_audit"] = get_process_memory_mb()

        if SAVE_AUDIT_COLUMNS_IN_PARQUET:
            output_df["audit_status"] = audit.get("audit_status")
            output_df["audit_label_columns_count"] = audit.get("checks", {}).get("label_columns_count")
            output_df["audit_datetime_exact_match"] = audit.get("checks", {}).get("datetime_exact_match")
            output_df["audit_row_count_match"] = audit.get("checks", {}).get("row_count_match")

        output_df = downcast_float_columns(output_df)
        result["memory_mb_after_downcast"] = get_process_memory_mb()

        ensure_dir(output_path.parent)
        tmp_path = output_path.with_suffix(".tmp.parquet")

        output_df.to_parquet(
            tmp_path,
            index=False,
            engine=PARQUET_ENGINE,
            compression=PARQUET_COMPRESSION,
        )

        tmp_path.replace(output_path)
        result["memory_mb_after_write"] = get_process_memory_mb()

        label_cols = [c for c in output_df.columns if str(c).startswith("label_")]

        result.update({
            "status": "OK",
            "output_rows": int(len(output_df)),
            "output_columns": int(len(output_df.columns)),
            "label_columns_count": int(len(label_cols)),
            "label_metadata_count": int(len(label_metadata)),
            "label_metadata": label_metadata,
            "execution_steps": execution_steps,
            "audit": audit,
            "output_first_datetime": str(output_df["DateTime"].min()),
            "output_last_datetime": str(output_df["DateTime"].max()),
            "sample_hash": make_hash_from_dataframe(output_df),
            "elapsed_seconds": round(time.time() - start, 6),
        })

        del df_features
        del labels_df
        del output_df

        if FORCE_GC_EACH_SERIES:
            gc.collect()

        return result

    except Exception as exc:
        result.update({
            "status": "ERROR",
            "error": str(exc),
            "traceback": traceback.format_exc(),
            "elapsed_seconds": round(time.time() - start, 6),
        })

        if FORCE_GC_EACH_SERIES:
            gc.collect()

        return result


# =============================================================================
# 20. RELATÓRIOS
# =============================================================================

def build_labels_catalog(results: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    seen = set()
    catalog: List[Dict[str, Any]] = []

    for r in results:
        meta = r.get("label_metadata", [])
        if not isinstance(meta, list):
            continue

        for item in meta:
            label_name = item.get("label_name")
            label_config_id = item.get("label_config_id")
            key = (label_name, label_config_id)

            if key in seen:
                continue

            seen.add(key)
            catalog.append(item)

    return catalog


def build_master_json_payload(
    mapa_ativos: Dict[str, Any],
    features_json: Dict[str, Any],
    base_arquivos: Dict[str, Any],
    machine_profile: Dict[str, Any],
    feature_files: List[Dict[str, Any]],
    results: List[Dict[str, Any]],
) -> Dict[str, Any]:

    ok = [r for r in results if r.get("status") == "OK"]
    errors = [r for r in results if r.get("status") == "ERROR"]
    skipped = [r for r in results if str(r.get("status", "")).startswith("SKIPPED")]

    labels_catalog = build_labels_catalog(results)

    assets = sorted({r.get("asset") for r in ok if r.get("asset")})
    sources = sorted({r.get("source") for r in ok if r.get("source")})
    timeframes = sorted({r.get("timeframe") for r in ok if r.get("timeframe")})

    audit_pass = sum(1 for r in ok if (r.get("audit") or {}).get("audit_status") == "PASS")
    audit_warning = sum(1 for r in ok if (r.get("audit") or {}).get("audit_status") == "WARNING")
    audit_fail = sum(1 for r in ok if (r.get("audit") or {}).get("audit_status") in {"FAIL", "ERROR"})

    return {
        "schema_version": SCHEMA_VERSION,
        "system": {
            "name": SYSTEM_NAME,
            "version": SYSTEM_VERSION,
            "layer": "4_LABELS",
            "script": SCRIPT_NAME,
            "run_id": RUN_ID,
            "generated_at": now_iso(),
        },
        "paths": {
            "root_dir": path_to_str(ROOT_DIR),
            "labels_dir": path_to_str(LABELS_DIR),
            "labels_parquet_dir": path_to_str(LABELS_PARQUET_DIR),
            "labels_json_path": path_to_str(LABELS_JSON_PATH),
            "labels_excel_path": path_to_str(LABELS_EXCEL_PATH),
            "features_json_path": path_to_str(FEATURES_JSON_PATH),
            "mapa_ativos_path": path_to_str(MAPA_ATIVOS_PATH),
            "base_arquivos_path": path_to_str(BASE_ARQUIVOS_PATH),
            "machine_profile_path": path_to_str(MACHINE_PROFILE_PATH),
            "run_report_path": path_to_str(RUN_REPORT_PATH),
            "incremental_audit_path": path_to_str(INCREMENTAL_AUDIT_PATH),
        },
        "timezone_policy": TIMEZONE_POLICY,
        "input_dependencies": {
            "mapa_ativos_loaded": bool(mapa_ativos),
            "features_json_loaded": bool(features_json),
            "base_arquivos_loaded": bool(base_arquivos),
            "machine_profile_loaded": bool(machine_profile),
        },
        "label_policy": {
            "uses_future_data": True,
            "labels_are_targets_not_features": True,
            "recommended_ml_join_keys": [
                "DateTime",
                "meta_asset",
                "meta_symbol",
                "meta_source",
                "meta_timeframe",
            ],
            "anti_leakage_warning": (
                "Labels usam dados futuros por definição. Nunca incluir colunas label_* "
                "como features no treino. Usar purging e embargo em walk-forward."
            ),
        },
        "cost_model_path": path_to_str(COST_MODEL_PATH),
        "cost_config": COST_CONFIG,
        "label_configs": LABEL_CONFIGS,
        "performance_policy": {
            "enable_parallel_processing": ENABLE_PARALLEL_PROCESSING,
            "max_workers_labels": MAX_WORKERS_LABELS,
            "label_resource_profile": LABEL_RESOURCE_PROFILE,
            "force_gc_each_series": FORCE_GC_EACH_SERIES,
            "read_only_needed_columns": READ_ONLY_NEEDED_COLUMNS,
            "downcast_floats_to_float32": DOWNCAST_FLOATS_TO_FLOAT32,
            "save_raw_label_components": SAVE_RAW_LABEL_COMPONENTS,
            "save_audit_columns_in_parquet": SAVE_AUDIT_COLUMNS_IN_PARQUET,
            "parquet_engine": PARQUET_ENGINE,
            "parquet_compression": PARQUET_COMPRESSION,
            "resource_telemetry_enabled": RESOURCE_TELEMETRY_ENABLED,
            "resource_telemetry_interval_seconds": RESOURCE_TELEMETRY_INTERVAL_SECONDS,
            "resource_telemetry_max_samples": RESOURCE_TELEMETRY_MAX_SAMPLES,
        },
        "hardware_telemetry": build_resource_telemetry_payload(),
        "summary": {
            "feature_files_discovered": len(feature_files),
            "series_attempted": len(results),
            "series_ok": len(ok),
            "series_error": len(errors),
            "series_skipped": len(skipped),
            "assets_ok": assets,
            "sources_ok": sources,
            "timeframes_ok": timeframes,
            "labels_catalog_count": len(labels_catalog),
            "audit_pass": audit_pass,
            "audit_warning": audit_warning,
            "audit_fail_or_error": audit_fail,
            "total_label_rows": int(sum(r.get("output_rows", 0) or 0 for r in ok)),
            "total_label_files": len(ok),
            "max_memory_mb_after_label_generation": max(
                [float(r.get("memory_mb_after_label_generation") or 0.0) for r in ok] or [0.0]
            ),
            "max_memory_mb_after_output_concat": max(
                [float(r.get("memory_mb_after_output_concat") or 0.0) for r in ok] or [0.0]
            ),
            "max_memory_mb_after_write": max(
                [float(r.get("memory_mb_after_write") or 0.0) for r in ok] or [0.0]
            ),
            "avg_label_columns_per_file": (
                float(np.mean([r.get("label_columns_count", 0) or 0 for r in ok])) if ok else 0.0
            ),
        },
        "labels_catalog": labels_catalog,
        "series_outputs": [
            {
                "status": r.get("status"),
                "asset": r.get("asset"),
                "symbol": r.get("symbol"),
                "source": r.get("source"),
                "timeframe": r.get("timeframe"),
                "timeframe_seconds": r.get("timeframe_seconds"),
                "series_id": r.get("series_id"),
                "feature_path": r.get("feature_path"),
                "output_path": r.get("output_path"),
                "output_rows": r.get("output_rows"),
                "output_columns": r.get("output_columns"),
                "label_columns_count": r.get("label_columns_count"),
                "audit_status": (r.get("audit") or {}).get("audit_status") if isinstance(r.get("audit"), dict) else None,
                "elapsed_seconds": r.get("elapsed_seconds"),
                "memory_mb_start": r.get("memory_mb_start"),
                "memory_mb_after_read": r.get("memory_mb_after_read"),
                "memory_mb_after_label_generation": r.get("memory_mb_after_label_generation"),
                "memory_mb_after_output_concat": r.get("memory_mb_after_output_concat"),
                "memory_mb_after_audit": r.get("memory_mb_after_audit"),
                "memory_mb_after_downcast": r.get("memory_mb_after_downcast"),
                "memory_mb_after_write": r.get("memory_mb_after_write"),
                "error": r.get("error"),
            }
            for r in results
        ],
        "recommended_next_steps": {
            "dataset_builder": "Criar etapa 5 para juntar features + labels por DateTime/source/asset/timeframe.",
            "panel_labels": "Criar etapa separada para labels cross-sectional e relative-to-benchmark usando painel completo.",
            "walk_forward": "Usar purging e embargo para evitar leakage em labels com horizonte futuro.",
            "risk_engine": "Modelo gera sinal; risk engine independente decide se opera e tamanho.",
            "meta_labeling": "Usar meta_success_net dos Triple Barriers para filtrar sinais primários.",
        },
    }


def save_run_report(results: List[Dict[str, Any]], feature_files: List[Dict[str, Any]]) -> Path:
    ok = [r for r in results if r.get("status") == "OK"]
    errors = [r for r in results if r.get("status") == "ERROR"]
    skipped = [r for r in results if str(r.get("status", "")).startswith("SKIPPED")]
    audit_pass = sum(1 for r in ok if (r.get("audit") or {}).get("audit_status") == "PASS")
    audit_warning = sum(1 for r in ok if (r.get("audit") or {}).get("audit_status") == "WARNING")
    audit_fail = sum(1 for r in ok if (r.get("audit") or {}).get("audit_status") in {"FAIL", "ERROR"})

    payload = {
        "schema_version": "ARCHANGEL_LABELS_RUN_REPORT_2.0",
        "generated_at": now_iso(),
        "run_id": RUN_ID,
        "script": SCRIPT_NAME,
        "summary": {
            "feature_files_discovered": len(feature_files),
            "series_attempted": len(results),
            "series_ok": sum(1 for r in results if r.get("status") == "OK"),
            "series_error": sum(1 for r in results if r.get("status") == "ERROR"),
            "series_skipped": sum(1 for r in results if str(r.get("status", "")).startswith("SKIPPED")),
            "audit_pass": audit_pass,
            "audit_warning": audit_warning,
            "audit_fail_or_error": audit_fail,
            "total_label_rows": int(sum(r.get("output_rows", 0) or 0 for r in ok)),
            "total_label_files": len(ok),
        },
        "paths": {
            "run_report_path": path_to_str(RUN_REPORT_PATH),
            "run_report_base_json_path": path_to_str(RUN_REPORT_BASE_JSON_PATH),
            "run_report_latest_path": path_to_str(RUN_REPORT_LATEST_PATH),
            "labels_json_path": path_to_str(LABELS_JSON_PATH),
            "labels_excel_path": path_to_str(LABELS_EXCEL_PATH),
            "labels_parquet_dir": path_to_str(LABELS_PARQUET_DIR),
            "features_json_path": path_to_str(FEATURES_JSON_PATH),
            "mapa_ativos_path": path_to_str(MAPA_ATIVOS_PATH),
            "incremental_audit_path": path_to_str(INCREMENTAL_AUDIT_PATH),
        },
        "label_configs": LABEL_CONFIGS,
        "performance_policy": {
            "enable_parallel_processing": ENABLE_PARALLEL_PROCESSING,
            "max_workers_labels": MAX_WORKERS_LABELS,
            "label_resource_profile": LABEL_RESOURCE_PROFILE,
            "resource_telemetry_enabled": RESOURCE_TELEMETRY_ENABLED,
            "resource_telemetry_interval_seconds": RESOURCE_TELEMETRY_INTERVAL_SECONDS,
        },
        "hardware_telemetry": build_resource_telemetry_payload(),
        "series_outputs": [
            {
                "status": r.get("status"),
                "asset": r.get("asset"),
                "symbol": r.get("symbol"),
                "source": r.get("source"),
                "timeframe": r.get("timeframe"),
                "timeframe_seconds": r.get("timeframe_seconds"),
                "series_id": r.get("series_id"),
                "feature_path": r.get("feature_path"),
                "output_path": r.get("output_path"),
                "output_rows": r.get("output_rows"),
                "output_columns": r.get("output_columns"),
                "label_columns_count": r.get("label_columns_count"),
                "audit_status": (r.get("audit") or {}).get("audit_status") if isinstance(r.get("audit"), dict) else None,
                "elapsed_seconds": r.get("elapsed_seconds"),
                "memory_mb_start": r.get("memory_mb_start"),
                "memory_mb_after_read": r.get("memory_mb_after_read"),
                "memory_mb_after_label_generation": r.get("memory_mb_after_label_generation"),
                "memory_mb_after_output_concat": r.get("memory_mb_after_output_concat"),
                "memory_mb_after_audit": r.get("memory_mb_after_audit"),
                "memory_mb_after_downcast": r.get("memory_mb_after_downcast"),
                "memory_mb_after_write": r.get("memory_mb_after_write"),
                "error": r.get("error"),
            }
            for r in results
        ],
        "error_outputs": [
            {
                "status": r.get("status"),
                "series_id": r.get("series_id"),
                "asset": r.get("asset"),
                "symbol": r.get("symbol"),
                "source": r.get("source"),
                "timeframe": r.get("timeframe"),
                "feature_path": r.get("feature_path"),
                "output_path": r.get("output_path"),
                "error": r.get("error"),
            }
            for r in errors + skipped
        ],
    }

    write_json_atomic(payload, RUN_REPORT_PATH)
    write_json_atomic(payload, RUN_REPORT_BASE_JSON_PATH)
    write_json_atomic(payload, RUN_REPORT_LATEST_PATH)
    return RUN_REPORT_PATH


# =============================================================================
# 21. EXCEL REPORT
# =============================================================================

def build_dashboard_df(results: List[Dict[str, Any]], feature_files: List[Dict[str, Any]]) -> pd.DataFrame:
    ok = [r for r in results if r.get("status") == "OK"]
    errors = [r for r in results if r.get("status") == "ERROR"]
    skipped = [r for r in results if str(r.get("status", "")).startswith("SKIPPED")]

    audit_pass = sum(1 for r in ok if (r.get("audit") or {}).get("audit_status") == "PASS")
    audit_warning = sum(1 for r in ok if (r.get("audit") or {}).get("audit_status") == "WARNING")
    audit_fail = sum(1 for r in ok if (r.get("audit") or {}).get("audit_status") in {"FAIL", "ERROR"})

    total_elapsed = sum(float(r.get("elapsed_seconds") or 0) for r in results)
    avg_elapsed = total_elapsed / max(1, len(results))

    avg_label_cols = float(np.mean([r.get("label_columns_count", 0) or 0 for r in ok])) if ok else 0.0

    rows = [
        {"section": "RUN", "metric": "Run ID", "value": RUN_ID, "interpretation": "Unique run identifier."},
        {"section": "RUN", "metric": "Generated At", "value": now_iso(), "interpretation": "Local generation timestamp."},
        {"section": "RUN", "metric": "Schema Version", "value": SCHEMA_VERSION, "interpretation": "Label store schema version."},

        {"section": "COVERAGE", "metric": "Feature Files Discovered", "value": len(feature_files), "interpretation": "Input feature files found."},
        {"section": "COVERAGE", "metric": "Series OK", "value": len(ok), "interpretation": "Label files successfully created."},
        {"section": "COVERAGE", "metric": "Series Error", "value": len(errors), "interpretation": "Series with processing errors."},
        {"section": "COVERAGE", "metric": "Series Skipped", "value": len(skipped), "interpretation": "Series skipped."},

        {"section": "LABELS", "metric": "Enabled Label Configs", "value": sum(1 for c in LABEL_CONFIGS if c.get("enabled")), "interpretation": "Number of active label configurations."},
        {"section": "LABELS", "metric": "Total Label Rows", "value": int(sum(r.get("output_rows", 0) or 0 for r in ok)), "interpretation": "Total rows produced across label files."},
        {"section": "LABELS", "metric": "Total Label Files", "value": len(ok), "interpretation": "Number of Parquet label files generated."},
        {"section": "LABELS", "metric": "Average Label Columns/File", "value": round(avg_label_cols, 2), "interpretation": "Average number of label columns per output file."},

        {"section": "AUDIT", "metric": "Audit PASS", "value": audit_pass, "interpretation": "Files with clean label audit."},
        {"section": "AUDIT", "metric": "Audit WARNING", "value": audit_warning, "interpretation": "Files with non-critical warnings."},
        {"section": "AUDIT", "metric": "Audit FAIL/ERROR", "value": audit_fail, "interpretation": "Files with critical audit issues."},

        {"section": "PERFORMANCE", "metric": "Parallel Processing", "value": ENABLE_PARALLEL_PROCESSING, "interpretation": "Parallel label generation enabled."},
        {"section": "PERFORMANCE", "metric": "Max Workers", "value": MAX_WORKERS_LABELS, "interpretation": "Maximum parallel workers."},
        {"section": "PERFORMANCE", "metric": "Average Seconds / Series", "value": round(avg_elapsed, 3), "interpretation": "Average processing time per feature file."},

        {"section": "GOVERNANCE", "metric": "Uses Future Data", "value": True, "interpretation": "Labels use future path by definition."},
        {"section": "GOVERNANCE", "metric": "Use As Feature?", "value": "NO", "interpretation": "Never use label_* columns as input features."},
    ]

    return pd.DataFrame(rows)


def build_series_df(results: List[Dict[str, Any]]) -> pd.DataFrame:
    rows = []

    for r in results:
        audit = r.get("audit", {}) if isinstance(r.get("audit"), dict) else {}
        checks = audit.get("checks", {}) if isinstance(audit.get("checks"), dict) else {}

        rows.append({
            "status": r.get("status"),
            "audit_status": audit.get("audit_status"),
            "asset": r.get("asset"),
            "symbol": r.get("symbol"),
            "source": r.get("source"),
            "timeframe": r.get("timeframe"),
            "input_rows": r.get("input_rows"),
            "output_rows": r.get("output_rows"),
            "output_columns": r.get("output_columns"),
            "label_columns_count": r.get("label_columns_count"),
            "row_count_match": checks.get("row_count_match"),
            "datetime_exact_match": checks.get("datetime_exact_match"),
            "large_gap_count": checks.get("large_gap_count"),
            "label_null_ratio": checks.get("label_null_ratio"),
            "elapsed_seconds": r.get("elapsed_seconds"),
            "feature_path": r.get("feature_path"),
            "output_path": r.get("output_path"),
            "error": r.get("error"),
        })

    return pd.DataFrame(rows)


def build_label_configs_df() -> pd.DataFrame:
    return pd.DataFrame(LABEL_CONFIGS)


def build_label_catalog_df(results: List[Dict[str, Any]]) -> pd.DataFrame:
    return pd.DataFrame(build_labels_catalog(results))


def build_audit_df(results: List[Dict[str, Any]]) -> pd.DataFrame:
    rows = []

    for r in results:
        audit = r.get("audit", {}) if isinstance(r.get("audit"), dict) else {}
        checks = audit.get("checks", {}) if isinstance(audit.get("checks"), dict) else {}

        rows.append({
            "asset": r.get("asset"),
            "symbol": r.get("symbol"),
            "source": r.get("source"),
            "timeframe": r.get("timeframe"),
            "status": r.get("status"),
            "audit_status": audit.get("audit_status"),
            "row_count_match": checks.get("row_count_match"),
            "datetime_exact_match": checks.get("datetime_exact_match"),
            "datetime_monotonic": checks.get("datetime_monotonic"),
            "datetime_duplicates": checks.get("datetime_duplicates"),
            "expected_timeframe_seconds": checks.get("expected_timeframe_seconds"),
            "median_delta_seconds": checks.get("median_delta_seconds"),
            "max_delta_seconds": checks.get("max_delta_seconds"),
            "large_gap_count": checks.get("large_gap_count"),
            "label_columns_count": checks.get("label_columns_count"),
            "label_null_ratio": checks.get("label_null_ratio"),
            "warnings": " | ".join(audit.get("warnings", [])) if isinstance(audit.get("warnings"), list) else None,
            "errors": " | ".join(audit.get("errors", [])) if isinstance(audit.get("errors"), list) else None,
            "output_path": r.get("output_path"),
        })

    return pd.DataFrame(rows)


def build_errors_df(results: List[Dict[str, Any]]) -> pd.DataFrame:
    rows = []

    for r in results:
        if r.get("status") != "ERROR":
            continue

        rows.append({
            "asset": r.get("asset"),
            "symbol": r.get("symbol"),
            "source": r.get("source"),
            "timeframe": r.get("timeframe"),
            "feature_path": r.get("feature_path"),
            "output_path": r.get("output_path"),
            "error": r.get("error"),
            "elapsed_seconds": r.get("elapsed_seconds"),
        })

    return pd.DataFrame(rows)


def build_costs_df() -> pd.DataFrame:
    rows = []
    for source, cfg in COST_CONFIG.items():
        row = {"source": source}
        row.update(cfg)
        rows.append(row)
    return pd.DataFrame(rows)


def build_usage_guide_df() -> pd.DataFrame:
    rows = [
        {
            "topic": "Purpose",
            "description": "Labels are target variables for ML/DL and backtesting. They use future data by definition.",
            "action": "Never include label_* columns as model features.",
        },
        {
            "topic": "Expanded Labels",
            "description": "This version creates return, direction, quantile, path, vol, tail, trend and triple-barrier labels.",
            "action": "Start with a subset of labels, then compare performance by target family.",
        },
        {
            "topic": "Triple Barrier",
            "description": "Captures whether take profit, stop loss, or time barrier is reached first.",
            "action": "Use label_*_tb as classification target.",
        },
        {
            "topic": "Meta Labeling",
            "description": "Meta-success labels are included for triple-barrier outputs.",
            "action": "Train a second-stage model to decide whether to act on a primary signal.",
        },
        {
            "topic": "Walk-forward",
            "description": "Many labels overlap in time.",
            "action": "Use purging and embargo in train/test splits.",
        },
        {
            "topic": "Cross-Asset",
            "description": "Correlations are usually features, not labels. Relative future returns can be labels.",
            "action": "Use separate panel builder for benchmark-relative and cross-sectional targets.",
        },
    ]

    return pd.DataFrame(rows)


def style_excel_workbook(writer: pd.ExcelWriter) -> None:
    try:
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.formatting.rule import CellIsRule
        from openpyxl.utils import get_column_letter

        wb = writer.book

        dark_blue = "17365D"
        white = "FFFFFF"
        light_green = "E2F0D9"
        light_red = "F4CCCC"
        light_yellow = "FFF2CC"
        thin = Side(style="thin", color="D9E2F3")

        for ws in wb.worksheets:
            ws.sheet_view.showGridLines = False

            if ws.max_row >= 1:
                for cell in ws[1]:
                    cell.fill = PatternFill("solid", fgColor=dark_blue)
                    cell.font = Font(color=white, bold=True)
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    cell.border = Border(bottom=thin)

                ws.freeze_panes = "A2"
                ws.auto_filter.ref = ws.dimensions

            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(vertical="top", wrap_text=False)
                    cell.border = Border(bottom=thin)

            for idx, col_cells in enumerate(ws.columns, start=1):
                max_len = 10
                for cell in list(col_cells)[:400]:
                    value = "" if cell.value is None else str(cell.value)
                    max_len = max(max_len, len(value))
                ws.column_dimensions[get_column_letter(idx)].width = min(max_len + 2, 80)

            if ws.title in {"SERIES", "AUDIT"}:
                header = [c.value for c in ws[1]]
                for col_idx, name in enumerate(header, start=1):
                    col_letter = get_column_letter(col_idx)

                    if name in {"status", "audit_status"}:
                        ws.conditional_formatting.add(
                            f"{col_letter}2:{col_letter}{ws.max_row}",
                            CellIsRule(operator="equal", formula=['"OK"'], fill=PatternFill("solid", fgColor=light_green)),
                        )
                        ws.conditional_formatting.add(
                            f"{col_letter}2:{col_letter}{ws.max_row}",
                            CellIsRule(operator="equal", formula=['"PASS"'], fill=PatternFill("solid", fgColor=light_green)),
                        )
                        ws.conditional_formatting.add(
                            f"{col_letter}2:{col_letter}{ws.max_row}",
                            CellIsRule(operator="equal", formula=['"WARNING"'], fill=PatternFill("solid", fgColor=light_yellow)),
                        )
                        ws.conditional_formatting.add(
                            f"{col_letter}2:{col_letter}{ws.max_row}",
                            CellIsRule(operator="equal", formula=['"ERROR"'], fill=PatternFill("solid", fgColor=light_red)),
                        )
                        ws.conditional_formatting.add(
                            f"{col_letter}2:{col_letter}{ws.max_row}",
                            CellIsRule(operator="equal", formula=['"FAIL"'], fill=PatternFill("solid", fgColor=light_red)),
                        )

    except Exception:
        pass


def save_excel_report(results: List[Dict[str, Any]], feature_files: List[Dict[str, Any]]) -> Path:
    ensure_dir(BASE_EXCEL_DIR)
    ensure_dir(LABELS_EXCEL_DIR)

    dashboard_df = build_dashboard_df(results, feature_files)
    series_df = build_series_df(results)
    configs_df = build_label_configs_df()
    catalog_df = build_label_catalog_df(results)
    audit_df = build_audit_df(results)
    costs_df = build_costs_df()
    usage_df = build_usage_guide_df()
    errors_df = build_errors_df(results)

    with pd.ExcelWriter(LABELS_EXCEL_PATH, engine="openpyxl") as writer:
        dashboard_df.to_excel(writer, sheet_name="DASHBOARD", index=False)
        series_df.to_excel(writer, sheet_name="SERIES", index=False)
        configs_df.to_excel(writer, sheet_name="LABEL_CONFIGS", index=False)
        catalog_df.to_excel(writer, sheet_name="LABEL_CATALOG", index=False)
        audit_df.to_excel(writer, sheet_name="AUDIT", index=False)
        costs_df.to_excel(writer, sheet_name="COSTS", index=False)
        usage_df.to_excel(writer, sheet_name="USAGE_GUIDE", index=False)
        errors_df.to_excel(writer, sheet_name="ERRORS", index=False)

        style_excel_workbook(writer)

    try:
        local_copy = LABELS_EXCEL_DIR / f"4_LABELS_{RUN_ID}.xlsx"
        import shutil
        shutil.copy2(LABELS_EXCEL_PATH, local_copy)
    except Exception:
        pass

    return LABELS_EXCEL_PATH


# =============================================================================
# 22. MAIN
# =============================================================================

def main() -> None:
    total_start = time.time()
    telemetry_thread = start_resource_telemetry()

    ensure_dir(BASE_JSON_DIR)
    ensure_dir(BASE_EXCEL_DIR)
    ensure_dir(LABELS_DIR)
    ensure_dir(LABELS_PARQUET_DIR)
    ensure_dir(LABELS_LOG_DIR)
    ensure_dir(LABELS_AUDIT_DIR)
    ensure_dir(LABELS_EXCEL_DIR)

    print("=" * 120)
    print("ARCHANGEL v1 | 4_LABELS | EXTENDED LABEL FACTORY")
    print("=" * 120)
    print(f"[SCRIPT] {SCRIPT_NAME}")
    print(f"[SCHEMA] {SCHEMA_VERSION}")
    print(f"[RUN_ID] {RUN_ID}")
    print(f"[ROOT_DIR] {ROOT_DIR}")
    print(f"[FEATURES_JSON] {FEATURES_JSON_PATH}")
    print(f"[MAPA_ATIVOS] {MAPA_ATIVOS_PATH}")
    print(f"[LABELS_PARQUET_DIR] {LABELS_PARQUET_DIR}")
    print(f"[LABELS_JSON_PATH] {LABELS_JSON_PATH}")
    print(f"[LABELS_EXCEL_PATH] {LABELS_EXCEL_PATH}")
    print(f"[COST_MODEL_PATH] {COST_MODEL_PATH}")
    print(f"[INÍCIO] {now_iso()}")
    print("=" * 120)

    mapa_ativos = load_json(MAPA_ATIVOS_PATH, required=True)
    features_json = load_json(FEATURES_JSON_PATH, required=True)
    base_arquivos = load_json(BASE_ARQUIVOS_PATH, required=False)
    machine_profile = load_json(MACHINE_PROFILE_PATH, required=False)

    feature_files = discover_feature_files(features_json)

    print(f"[FEATURE FILES DESCOBERTOS] {len(feature_files)}")
    print(f"[LABEL CONFIGS ENABLED] {sum(1 for c in LABEL_CONFIGS if c.get('enabled'))}")
    print(f"[PARALELISMO] {ENABLE_PARALLEL_PROCESSING}")
    print(f"[MAX_WORKERS_LABELS] {MAX_WORKERS_LABELS}")
    print(f"[RESOURCE_PROFILE] {LABEL_RESOURCE_PROFILE}")
    print(f"[TELEMETRIA HW] {RESOURCE_TELEMETRY_ENABLED} | intervalo={RESOURCE_TELEMETRY_INTERVAL_SECONDS}s")

    if not feature_files:
        raise RuntimeError(
            "Nenhum arquivo de features encontrado. Verifique 03_FEATURES_CATALOG_LATEST.json e FEATURES_PARQUET_DIR."
        )

    append_incremental_audit({
        "event": "RUN_STARTED",
        "feature_files_discovered": len(feature_files),
        "enabled_label_configs": sum(1 for c in LABEL_CONFIGS if c.get("enabled")),
        "schema_version": SCHEMA_VERSION,
        "labels_dir": path_to_str(LABELS_DIR),
        "labels_parquet_dir": path_to_str(LABELS_PARQUET_DIR),
    })

    results: List[Dict[str, Any]] = []

    try:
        if ENABLE_PARALLEL_PROCESSING and len(feature_files) > 1:
            workers = min(MAX_WORKERS_LABELS, os.cpu_count() or 1, len(feature_files))
            print(f"[PARALELISMO] Ativado com {workers} workers")

            with ProcessPoolExecutor(max_workers=workers) as executor:
                future_to_idx = {
                    executor.submit(process_one_feature_file, item): idx
                    for idx, item in enumerate(feature_files, start=1)
                }

                completed = 0

                for future in as_completed(future_to_idx):
                    idx = future_to_idx[future]
                    completed += 1

                    try:
                        result = future.result()
                    except Exception as exc:
                        result = {
                            "run_id": RUN_ID,
                            "status": "ERROR",
                            "error": str(exc),
                            "traceback": traceback.format_exc(),
                            "progress_index": idx,
                            "progress_total": len(feature_files),
                        }

                    result["progress_index"] = idx
                    result["progress_total"] = len(feature_files)

                    results.append(result)
                    append_incremental_audit(result)

                    print(
                        f"[PROGRESSO] {completed}/{len(feature_files)} | "
                        f"idx={idx} | status={result.get('status')} | "
                        f"audit={(result.get('audit') or {}).get('audit_status') if isinstance(result.get('audit'), dict) else None} | "
                        f"{result.get('source')} | {result.get('asset')} | {result.get('timeframe')} | "
                        f"labels={result.get('label_columns_count')} | "
                        f"{result.get('elapsed_seconds')}s"
                    )

        else:
            for idx, item in enumerate(feature_files, start=1):
                print(f"\n[PROGRESSO] {idx}/{len(feature_files)}")
                result = process_one_feature_file(item)
                result["progress_index"] = idx
                result["progress_total"] = len(feature_files)

                results.append(result)
                append_incremental_audit(result)

    except KeyboardInterrupt:
        print("\n[INTERRUPÇÃO] Execução interrompida manualmente. Salvando relatórios parciais...")
        append_incremental_audit({
            "event": "RUN_INTERRUPTED_KEYBOARD",
            "processed_so_far": len(results),
            "feature_files_discovered": len(feature_files),
        })

    except Exception as exc:
        print("\n[ERRO FATAL] Salvando relatórios parciais...")
        print(str(exc))
        append_incremental_audit({
            "event": "RUN_FATAL_ERROR",
            "processed_so_far": len(results),
            "feature_files_discovered": len(feature_files),
            "error": str(exc),
            "traceback": traceback.format_exc(),
        })

    finally:
        results = sorted(
            results,
            key=lambda r: (
                str(r.get("source", "")),
                str(r.get("asset", "")),
                str(r.get("symbol", "")),
                str(r.get("timeframe", "")),
                str(r.get("feature_path", "")),
            ),
        )

        stop_resource_telemetry(telemetry_thread)
        telemetry_thread = None

        run_report_path = save_run_report(results, feature_files)

        if ENABLE_JSON_MASTER_REPORT:
            labels_json_payload = build_master_json_payload(
                mapa_ativos=mapa_ativos,
                features_json=features_json,
                base_arquivos=base_arquivos,
                machine_profile=machine_profile,
                feature_files=feature_files,
                results=results,
            )

            write_json_atomic(labels_json_payload, LABELS_JSON_PATH)

        if ENABLE_EXCEL_REPORT:
            excel_path = save_excel_report(results, feature_files)
        else:
            excel_path = None

        total_elapsed = round(time.time() - total_start, 6)

        ok_count = sum(1 for r in results if r.get("status") == "OK")
        error_count = sum(1 for r in results if r.get("status") == "ERROR")
        skipped_count = sum(1 for r in results if str(r.get("status", "")).startswith("SKIPPED"))

        audit_pass = sum(1 for r in results if (r.get("audit") or {}).get("audit_status") == "PASS")
        audit_warning = sum(1 for r in results if (r.get("audit") or {}).get("audit_status") == "WARNING")
        audit_fail = sum(1 for r in results if (r.get("audit") or {}).get("audit_status") in {"FAIL", "ERROR"})

        append_incremental_audit({
            "event": "RUN_FINISHED_OR_PARTIAL_SAVED",
            "processed_series": len(results),
            "feature_files_discovered": len(feature_files),
            "ok": ok_count,
            "error": error_count,
            "skipped": skipped_count,
            "audit_pass": audit_pass,
            "audit_warning": audit_warning,
            "audit_fail": audit_fail,
            "elapsed_seconds": total_elapsed,
            "labels_json_path": path_to_str(LABELS_JSON_PATH),
            "labels_excel_path": path_to_str(LABELS_EXCEL_PATH),
            "run_report_path": path_to_str(run_report_path),
        })

        print("=" * 120)
        print("[FINALIZADO / RELATÓRIOS SALVOS]")
        print(f"[FEATURE FILES DESCOBERTOS] {len(feature_files)}")
        print(f"[PROCESSADAS] {len(results)} / {len(feature_files)}")
        print(f"[OK] {ok_count}")
        print(f"[ERROS] {error_count}")
        print(f"[SKIPPED] {skipped_count}")
        print(f"[AUDIT PASS] {audit_pass}")
        print(f"[AUDIT WARNING] {audit_warning}")
        print(f"[AUDIT FAIL/ERROR] {audit_fail}")
        print(f"[TEMPO TOTAL] {total_elapsed}s")
        print(f"[LABELS JSON] {LABELS_JSON_PATH}")
        print(f"[LABELS EXCEL] {excel_path}")
        print(f"[RUN REPORT] {run_report_path}")
        print(f"[AUDITORIA INCREMENTAL] {INCREMENTAL_AUDIT_PATH}")
        print(f"[FIM] {now_iso()}")
        print("=" * 120)


if __name__ == "__main__":
    main()
